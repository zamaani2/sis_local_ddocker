from shs_system.models import (
    SchoolInformation,
    Assessment,
    Teacher,
    Class,
    Subject,
    TeacherSubjectAssignment,
    Term,
    ClassTeacher,
    GradingSystem,
    Form,
    Student,
    AttendanceRecord,
    SchoolInformation,
    Department,
    LearningArea,
    StudentClass,
    Assessment,
    ClassSubject,
    AcademicYear,
    ScoringConfiguration,
)
import csv
import datetime
import logging
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_POST, require_http_methods

logger = logging.getLogger(__name__)

from django.db.models import Q, Avg, Count, Case, When, IntegerField, F
import openpyxl
import pandas as pd
from io import StringIO
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from email.mime.text import MIMEText


# IMPROVEMENT 6: Add a helper function for optimized position calculation
def calculate_positions_optimized(class_subject, term):
    """
    Optimized version of position calculation to reduce database contention.

    This implementation:
    1. Fetches all assessments at once to reduce db queries
    2. Calculates positions in memory
    3. Uses bulk_update to reduce number of db operations
    """
    # Fetch all assessments for this class subject and term with a single query
    assessments = list(
        Assessment.objects.filter(
            class_subject=class_subject, assessment_type="exam_score", term=term
        ).select_related("student")
    )

    # Sort by total score in descending order
    assessments.sort(
        key=lambda a: a.total_score if a.total_score is not None else -1, reverse=True
    )

    # Assign positions
    current_position = 1
    last_score = None
    updates_needed = []

    for i, assessment in enumerate(assessments):
        if assessment.total_score is None:
            # Skip students with no scores
            assessment.position = None
            continue

        # Handle ties (same score = same position)
        if last_score is not None and assessment.total_score != last_score:
            current_position = i + 1

        assessment.position = current_position
        last_score = assessment.total_score
        updates_needed.append(assessment)

    # Bulk update to minimize database operations
    if updates_needed:
        # Use batch size to avoid oversized queries
        batch_size = 100
        for i in range(0, len(updates_needed), batch_size):
            batch = updates_needed[i : i + batch_size]
            Assessment.objects.bulk_update(batch, ["position"])

    return len(updates_needed)  # Return the number of updated positions


# Helper function to get user's school
def get_user_school(user):
    """
    Get the school associated with the current user.
    For superadmins, this will return None (they can access all schools).
    For other users, it will return their associated school.
    """
    if user.is_superadmin:
        return None  # Superadmins can access all schools
    return user.school


@login_required
def enter_scores(request):
    # Ensure only teachers can access this view
    if request.user.role != "teacher":
        return render(request, "errors/403.html", status=403)

    teacher = getattr(request.user, "teacher_profile", None)
    if not teacher:
        messages.error(request, "Teacher profile is not linked to this user.")
        return redirect("login")

    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    # Get the current academic year and term for the user's school
    if user_school:
        current_academic_year = AcademicYear.objects.filter(
            is_current=True, school=user_school
        ).first()
        current_term = Term.objects.filter(is_current=True, school=user_school).first()
    else:
        # For superadmins, get any current academic year/term
        current_academic_year = AcademicYear.objects.filter(is_current=True).first()
        current_term = Term.objects.filter(is_current=True).first()

    if not current_academic_year or not current_term:
        messages.error(request, "No active academic year or term is set.")
        return redirect("dashboard")

    # Fetch only assignments for the current academic year and user's school
    assignments = []
    if current_academic_year and current_term:
        assignments_query = teacher.teachersubjectassignment_set.filter(
            is_active=True, academic_year=current_academic_year
        )

        # Apply school filter for multi-tenancy if not superadmin
        if user_school:
            assignments_query = assignments_query.filter(school=user_school)


        # Filter out assignments where ClassSubject is not active
        # Get active class-subject combinations
        active_class_subjects = ClassSubject.objects.filter(
            academic_year=current_academic_year, is_active=True
        ).values_list('class_name_id', 'subject_id')
        
        # Filter assignments to only include those with active ClassSubject
        filtered_assignments = []
        for assignment in assignments_query.select_related("class_assigned", "subject"):
            if (assignment.class_assigned.id, assignment.subject.id) in active_class_subjects:
                filtered_assignments.append(assignment)
        
        assignments = filtered_assignments


    # Initialize context with common data
    context = {
        "assignments": assignments,
        "current_academic_year": current_academic_year,
        "current_term": current_term,
        "user_school": user_school,  # Add user's school to context
    }

    # Get scoring configuration for dynamic display
    if user_school:
        scoring_config = ScoringConfiguration.get_active_config(user_school)
        if scoring_config:
            context["scoring_config"] = scoring_config
            context["class_score_percentage"] = scoring_config.class_score_percentage
            context["exam_score_percentage"] = scoring_config.exam_score_percentage
            context["max_class_score"] = scoring_config.max_class_score
            context["max_exam_score"] = scoring_config.max_exam_score
        else:
            # Use defaults if no configuration exists
            context["class_score_percentage"] = 30
            context["exam_score_percentage"] = 70
            context["max_class_score"] = 30
            context["max_exam_score"] = 70
    else:
        # For superadmins, use defaults
        context["class_score_percentage"] = 30
        context["exam_score_percentage"] = 70
        context["max_class_score"] = 30
        context["max_exam_score"] = 70

    if request.method == "POST":
        try:
            assignment_id = request.POST.get("assignment_id")
            if not assignment_id:
                raise ValueError("No assignment ID provided")

            # Include school filter for multi-tenancy
            assignment_query = TeacherSubjectAssignment.objects.filter(
                id=assignment_id, teacher=teacher, is_active=True
            )

            if user_school:
                assignment_query = assignment_query.filter(school=user_school)

            assignment = get_object_or_404(assignment_query)

            if not current_term:
                messages.error(request, "No current term is set.")
                return redirect("enter_scores")

            # Get or create the ClassSubject instance with school context
            try:
                class_subject_query = ClassSubject.objects.filter(
                    subject=assignment.subject,
                    class_name=assignment.class_assigned,
                    academic_year=assignment.academic_year,

                    is_active=True

                )

                # Apply school filter through related objects
                if user_school:
                    class_subject_query = class_subject_query.filter(
                        subject__school=user_school,
                        class_name__school=user_school,
                        academic_year__school=user_school,
                    )

                class_subject = class_subject_query.first()

                if not class_subject:

                    # Check if there's an inactive ClassSubject that we can reactivate
                    inactive_class_subject_query = ClassSubject.objects.filter(
                        subject=assignment.subject,
                        class_name=assignment.class_assigned,
                        academic_year=assignment.academic_year,
                        is_active=False
                    )

                    if user_school:
                        inactive_class_subject_query = inactive_class_subject_query.filter(
                            subject__school=user_school,
                            class_name__school=user_school,
                            academic_year__school=user_school,
                        )

                    inactive_class_subject = inactive_class_subject_query.first()

                    if inactive_class_subject:
                        # Reactivate the existing assignment
                        inactive_class_subject.is_active = True
                        inactive_class_subject.save()
                        class_subject = inactive_class_subject
                    else:
                        # Don't create automatically - this should be done through admin interface
                        error_message = f"Subject '{assignment.subject.subject_name}' is not assigned to class '{assignment.class_assigned.name}'. Please assign this subject to the class first."
                        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                            return JsonResponse(
                                {
                                    "success": False,
                                    "message": error_message,
                                    "error_messages": [error_message],
                                }
                            )
                        messages.error(request, error_message)
                        return redirect("enter_scores")

                logger.debug(
                    f"ClassSubject {'reactivated' if class_subject and not class_subject_query.exists() else 'found'}: {class_subject}"

                )
            except Exception as e:
                error_message = f"Failed to create or retrieve class subject: {str(e)}"
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse(
                        {
                            "success": False,
                            "message": error_message,
                            "error_messages": [error_message],
                        }
                    )
                messages.error(request, error_message)
                return redirect("enter_scores")

            # Fetch all active students in the assigned class with school context
            student_classes_query = StudentClass.objects.filter(
                assigned_class=assignment.class_assigned, is_active=True
            )

            # Apply school filter for multi-tenancy
            if user_school:
                student_classes_query = student_classes_query.filter(school=user_school)

            student_classes = student_classes_query.select_related("student")

            # Sort students by ID to ensure consistent processing order
            student_classes = sorted(student_classes, key=lambda sc: sc.student.id)

            # Prepare collections for tracking operations
            error_messages = []
            success_messages = []
            updated_students = []

            # Process each student's data with individual saves
            for student_class in student_classes:
                student = student_class.student
                class_score = request.POST.get(f"class_score_{student.id}")
                exam_score = request.POST.get(f"exam_score_{student.id}")

                if not class_score or not exam_score:
                    continue  # Skip if no scores provided

                try:
                    class_score = float(class_score)
                    exam_score = float(exam_score)

                    # Get scoring configuration for dynamic validation and calculation
                    scoring_config = None
                    if user_school:
                        scoring_config = ScoringConfiguration.get_active_config(
                            user_school
                        )

                    # Use dynamic maximum values if configuration exists, otherwise use defaults
                    max_class_score = (
                        scoring_config.max_class_score if scoring_config else 30
                    )
                    max_exam_score = (
                        scoring_config.max_exam_score if scoring_config else 70
                    )

                    # Ensure scores are within valid ranges using dynamic values
                    if not (0 <= class_score <= max_class_score) or not (
                        0 <= exam_score <= max_exam_score
                    ):
                        error_msg = f"Invalid score range for {student}. Class score must be 0-{max_class_score}, Exam score must be 0-{max_exam_score}."
                        error_messages.append(error_msg)
                        continue

                    # Calculate total score using dynamic configuration
                    if scoring_config:
                        # Use dynamic configuration to calculate total score
                        class_score_decimal = Decimal(str(class_score))
                        exam_score_decimal = Decimal(str(exam_score))
                        max_class_score_decimal = Decimal(
                            str(scoring_config.max_class_score)
                        )
                        max_exam_score_decimal = Decimal(
                            str(scoring_config.max_exam_score)
                        )
                        class_score_percentage_decimal = Decimal(
                            str(scoring_config.class_score_percentage)
                        )
                        exam_score_percentage_decimal = Decimal(
                            str(scoring_config.exam_score_percentage)
                        )

                        class_score_weighted = (
                            class_score_decimal / max_class_score_decimal
                        ) * class_score_percentage_decimal
                        exam_score_weighted = (
                            exam_score_decimal / max_exam_score_decimal
                        ) * exam_score_percentage_decimal
                        total_score = float(class_score_weighted + exam_score_weighted)
                    else:
                        # Fallback to simple addition if no configuration
                        # Ensure both values are Decimal to avoid type mismatch
                        class_score_decimal = Decimal(str(class_score))
                        exam_score_decimal = Decimal(str(exam_score))
                        total_score = class_score_decimal + exam_score_decimal

                    # Use GradingSystem to get grade and remarks with school context
                    grade_info = GradingSystem.get_grade_for_score(
                        total_score, user_school
                    )

                    if not grade_info:
                        error_msg = f"No grading criteria found for score {total_score} for {student.full_name}"
                        error_messages.append(error_msg)
                        continue

                    grade = grade_info.grade_letter
                    remarks = grade_info.remarks

                    # Check if assessment exists and update or create accordingly - INDIVIDUAL SAVE
                    assessment_query = Assessment.objects.filter(
                        class_subject=class_subject,
                        student=student,
                        term=current_term,
                        assessment_type="exam_score",
                    )

                    assessment = assessment_query.first()

                    if assessment:
                        # Update existing assessment
                        assessment.class_score = class_score
                        assessment.exam_score = exam_score
                        # Don't set total_score here - let the model's save method calculate it
                        assessment.grade = grade
                        assessment.remarks = remarks
                        assessment.recorded_by = request.user
                        assessment.save()
                        created = False
                    else:
                        # Create new assessment
                        assessment = Assessment.objects.create(
                            class_subject=class_subject,
                            student=student,
                            term=current_term,
                            assessment_type="exam_score",
                            class_score=class_score,
                            exam_score=exam_score,
                            # Don't set total_score here - let the model's save method calculate it
                            grade=grade,
                            remarks=remarks,
                            recorded_by=request.user,
                        )
                        created = True

                    # Get the calculated total score from the saved assessment
                    total_score = (
                        float(assessment.total_score)
                        if assessment.total_score
                        else total_score
                    )

                    success_msg = f"Scores saved for {student}."
                    success_messages.append(success_msg)

                    # Add student data to updated_students list for AJAX
                    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                        updated_students.append(
                            {
                                "id": student.id,
                                "class_score": class_score,
                                "exam_score": exam_score,
                                "total_score": total_score,
                                "grade": grade,
                                "remarks": remarks,
                                "position": "TBD",  # Position will be updated later
                            }
                        )

                except ValueError:
                    error_msg = f"Invalid score format for {student}."
                    error_messages.append(error_msg)
                except Exception as e:
                    logger.error(
                        f"Error saving assessment for student {student.id}: {str(e)}"
                    )
                    error_msg = f"Error saving scores for {student}: {str(e)}"
                    error_messages.append(error_msg)

            # Calculate positions separately after all individual saves
            positions_calculated = False
            try:
                if (
                    success_messages
                ):  # Only calculate if at least one student was updated
                    calculate_positions_optimized(class_subject, current_term)
                    positions_calculated = True
                    logger.debug("Positions calculated successfully")
            except Exception as e:
                logger.error(f"Error calculating positions: {str(e)}")
                error_messages.append(f"Error calculating positions: {str(e)}")

            # Get updated positions for AJAX response
            if (
                positions_calculated
                and request.headers.get("X-Requested-With") == "XMLHttpRequest"
            ):
                try:
                    # Fetch the latest assessment data including positions
                    latest_assessments = Assessment.objects.filter(
                        class_subject=class_subject,
                        student__in=[
                            student_class.student for student_class in student_classes
                        ],
                        term=current_term,
                    )

                    # Update position information
                    position_map = {
                        assessment.student.id: assessment.position
                        for assessment in latest_assessments
                    }
                    for student_data in updated_students:
                        student_data["position"] = position_map.get(
                            student_data["id"], "N/A"
                        )
                except Exception as e:
                    logger.error(f"Error fetching updated positions: {str(e)}")
                    # Continue without accurate position data

            # If this is an AJAX request, return JSON response
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse(
                    {
                        "success": len(error_messages) == 0,
                        "message": (
                            "Scores saved successfully!"
                            if len(error_messages) == 0
                            else "There were errors saving some scores."
                        ),
                        "students": updated_students,
                        "error_messages": error_messages,
                        "success_messages": success_messages,
                    }
                )

            # For traditional form submission, set messages and redirect
            for message in error_messages:
                messages.error(request, message)

            for message in success_messages:
                messages.success(request, message)

            return redirect("enter_scores")

        except Exception as e:
            logger.error(f"Unexpected error in enter_scores: {str(e)}")
            error_message = f"An unexpected error occurred: {str(e)}"

            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse(
                    {
                        "success": False,
                        "message": error_message,
                        "error_messages": [error_message],
                    }
                )

            messages.error(request, error_message)
            return redirect("enter_scores")

    # Handle GET request - Code remains the same
    if request.GET.get("assignment_id"):
        assignment_id = request.GET["assignment_id"]

        # Include school filter for multi-tenancy
        assignment_query = TeacherSubjectAssignment.objects.filter(
            id=assignment_id, teacher=teacher, is_active=True
        )

        if user_school:
            assignment_query = assignment_query.filter(school=user_school)

        assignment = get_object_or_404(assignment_query)

        # Fetch all active students in the assigned class with school context
        student_classes_query = StudentClass.objects.filter(
            assigned_class=assignment.class_assigned, is_active=True
        )

        # Apply school filter for multi-tenancy
        if user_school:
            student_classes_query = student_classes_query.filter(school=user_school)

        student_classes = student_classes_query.select_related("student")

        # Fetch assessments for the selected assignment with school context

        # Exclude mock exam assessments - only get regular term assessments

        assessments_query = Assessment.objects.filter(
            class_subject__subject=assignment.subject,
            class_subject__class_name=assignment.class_assigned,
            class_subject__academic_year=assignment.academic_year,

            class_subject__is_active=True,
            term=current_term,
        ).exclude(assessment_type='mock_exam')


        # Apply school filter through related objects if needed
        if user_school:
            assessments_query = assessments_query.filter(
                class_subject__subject__school=user_school,
                class_subject__class_name__school=user_school,
                class_subject__academic_year__school=user_school,
            )

        assessments = assessments_query.select_related("student")

        # Map assessments by student ID
        student_assessments = {
            assessment.student.id: assessment for assessment in assessments
        }

        for student_class in student_classes:
            student = student_class.student
            assessment = student_assessments.get(student.id)
            if assessment:
                student.class_score = assessment.class_score
                student.exam_score = assessment.exam_score
                student.total_score = assessment.total_score
                student.grade = assessment.grade
                student.remarks = assessment.remarks
                student.position = assessment.position
            else:
                student.class_score = None
                student.exam_score = None
                student.total_score = None
                student.grade = None
                student.remarks = None
                student.position = None

        context["students_in_class"] = student_classes
        context["selected_assignment"] = assignment

    return render(request, "student/enter_scores.html", context)


@login_required
def export_scores(request):
    """Export scores for a teacher's assigned class and subject to Excel."""
    if request.user.role != "teacher":
        return render(request, "errors/403.html", status=403)

    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    # Get the assignment ID from the request
    assignment_id = request.GET.get("assignment_id")
    if not assignment_id:
        messages.error(request, "No assignment selected for export.")
        return redirect("enter_scores")

    # Get the teacher instance
    teacher = request.user.teacher_profile
    if not teacher:
        messages.error(request, "Teacher profile is not linked to this user.")
        return redirect("login")

    # Get the assignment with school context
    try:
        assignment_query = TeacherSubjectAssignment.objects.filter(
            id=assignment_id, teacher=teacher, is_active=True
        )

        # Apply school filter for multi-tenancy
        if user_school:
            assignment_query = assignment_query.filter(school=user_school)

        assignment = assignment_query.get()
    except TeacherSubjectAssignment.DoesNotExist:
        messages.error(request, "Invalid assignment selected.")
        return redirect("enter_scores")

    # Get current term with school context
    if user_school:
        current_term = Term.objects.filter(

            academic_year=assignment.academic_year,
            is_active=True, is_current=True, school=user_school
        ).first()
    else:
        current_term = Term.objects.filter(
            academic_year=assignment.academic_year,
            is_active=True, is_current=True

        ).first()

    if not current_term:
        messages.error(request, "No active term found for this academic year.")
        return redirect("enter_scores")

    # Get the ClassSubject instance with school context
    try:
        class_subject_query = ClassSubject.objects.filter(
            subject=assignment.subject,
            class_name=assignment.class_assigned,
            academic_year=assignment.academic_year,

            is_active=True

        )

        # Apply school filter through related objects
        if user_school:
            class_subject_query = class_subject_query.filter(
                subject__school=user_school,
                class_name__school=user_school,
                academic_year__school=user_school,
            )

        class_subject = class_subject_query.first()

        if not class_subject:

            # Check if there's an inactive ClassSubject that we can reactivate
            inactive_class_subject_query = ClassSubject.objects.filter(
                subject=assignment.subject,
                class_name=assignment.class_assigned,
                academic_year=assignment.academic_year,
                is_active=False
            )

            if user_school:
                inactive_class_subject_query = inactive_class_subject_query.filter(
                    subject__school=user_school,
                    class_name__school=user_school,
                    academic_year__school=user_school,
                )

            inactive_class_subject = inactive_class_subject_query.first()

            if inactive_class_subject:
                # Reactivate the existing assignment
                inactive_class_subject.is_active = True
                inactive_class_subject.save()
                class_subject = inactive_class_subject
            else:
                # Don't create automatically - this should be done through admin interface
                messages.error(request, f"Subject '{assignment.subject.subject_name}' is not assigned to class '{assignment.class_assigned.name}'. Please assign this subject to the class first.")
                return redirect("enter_scores")
    except ClassSubject.DoesNotExist:
        # Don't create automatically - this should be done through admin interface
        messages.error(request, f"Subject '{assignment.subject.subject_name}' is not assigned to class '{assignment.class_assigned.name}'. Please assign this subject to the class first.")
        return redirect("enter_scores")


    # Fetch all active students in the assigned class with school context
    student_classes_query = StudentClass.objects.filter(
        assigned_class=assignment.class_assigned, is_active=True
    )

    # Apply school filter for multi-tenancy
    if user_school:
        student_classes_query = student_classes_query.filter(school=user_school)

    student_classes = student_classes_query.select_related("student")

    # Fetch existing assessments for these students with school context

    # Exclude mock exam assessments from export
    assessments_query = Assessment.objects.filter(
        class_subject=class_subject, term=current_term
    ).exclude(assessment_type='mock_exam')

    assessments = assessments_query.select_related("student")

    # Create a dictionary for quick lookup of assessments by student ID
    student_assessments = {
        assessment.student.id: assessment for assessment in assessments
    }

    try:
        # Import necessary modules
        import tempfile
        from io import BytesIO

        # Create a new workbook and worksheet
        # Use xlsx.compat mode for better compatibility with Excel 2019
        wb = openpyxl.Workbook()
        wb.iso_dates = True  # Use ISO dates for better compatibility
        ws = wb.active

        # Clean sheet name to avoid Excel errors
        sheet_name = (
            f"{assignment.class_assigned.name}_{assignment.subject.subject_name}"
        )
        # Remove invalid characters for Excel sheet names
        invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, "_")
        # Limit sheet name length to 31 characters (Excel limit)
        ws.title = sheet_name[:31]

        # Define styles
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Get school information for the header
        if user_school:
            school_name = user_school.name
        else:
            # Try to get the active school or use a default
            active_school = SchoolInformation.get_active()
            school_name = active_school.name if active_school else "SCHOOL NAME"

        # Write header information
        ws["A1"] = f"School Name: {school_name}"
        ws["A2"] = f"Class: {assignment.class_assigned.name}"
        ws["A3"] = f"Subject: {assignment.subject.subject_name}"
        ws["A4"] = f"Term: {current_term.term_number}"
        ws["A5"] = f"Academic Year: {assignment.academic_year}"

        # Make headers bold
        for row in range(1, 6):
            ws.cell(row=row, column=1).font = Font(bold=True)

        # Get scoring configuration for dynamic headers
        scoring_config = None
        if user_school:
            scoring_config = ScoringConfiguration.get_active_config(user_school)

        # Use dynamic values if configuration exists, otherwise use defaults
        class_score_percentage = (
            scoring_config.class_score_percentage if scoring_config else 30
        )
        exam_score_percentage = (
            scoring_config.exam_score_percentage if scoring_config else 70
        )
        max_class_score = scoring_config.max_class_score if scoring_config else 30
        max_exam_score = scoring_config.max_exam_score if scoring_config else 70

        # Define column headers with dynamic percentages
        headers = [
            "Student ID",
            "Student Name",
            f"Class Score ({class_score_percentage}%)",
            f"Exam Score ({exam_score_percentage}%)",
            "Total Score",
            "Grade",
            "Remarks",
        ]

        # Apply headers to row 7
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # Add calculation instructions with dynamic max values
        ws["A6"] = (
            f"Instructions: Edit only the Class Score (max: {max_class_score}) and Exam Score (max: {max_exam_score}) columns."
        )
        ws.merge_cells("A6:G6")
        ws["A6"].font = Font(bold=True, color="FF0000")
        ws["A6"].alignment = Alignment(horizontal="center")

        # Add data rows
        row_num = 8
        for student_class in student_classes:
            student = student_class.student
            assessment = student_assessments.get(student.id)

            # Write student data
            ws.cell(row=row_num, column=1, value=student.admission_number).border = (
                border
            )
            ws.cell(row=row_num, column=2, value=student.full_name).border = border

            # Write assessment data if it exists
            class_score_cell = ws.cell(row=row_num, column=3)
            class_score_cell.value = assessment.class_score if assessment else ""
            class_score_cell.border = border
            # Set number format for better compatibility
            class_score_cell.number_format = "0.00"

            exam_score_cell = ws.cell(row=row_num, column=4)
            exam_score_cell.value = assessment.exam_score if assessment else ""
            exam_score_cell.border = border
            # Set number format for better compatibility
            exam_score_cell.number_format = "0.00"

            # Formula for total score - use string formula for compatibility
            total_cell = ws.cell(row=row_num, column=5)
            total_cell.value = (
                f'=IF(AND(C{row_num}<>"",D{row_num}<>""),C{row_num}+D{row_num},"")'
            )
            total_cell.border = border
            # Set number format for better compatibility
            total_cell.number_format = "0.00"

            # Formula for grade - use simpler formula for compatibility
            grade_cell = ws.cell(row=row_num, column=6)
            grade_cell.value = (
                f'=IF(E{row_num}="","",'
                f'IF(E{row_num}>=75,"A1",'
                f'IF(E{row_num}>=70,"B2",'
                f'IF(E{row_num}>=65,"B3",'
                f'IF(E{row_num}>=60,"C4",'
                f'IF(E{row_num}>=55,"C5",'
                f'IF(E{row_num}>=50,"C6",'
                f'IF(E{row_num}>=45,"D7",'
                f'IF(E{row_num}>=40,"E8","F9")))))))))'
            )
            grade_cell.border = border

            # Formula for remarks - use simpler formula for compatibility
            remarks_cell = ws.cell(row=row_num, column=7)
            remarks_cell.value = (
                f'=IF(E{row_num}="","",'
                f'IF(E{row_num}>=75,"EXCELLENT",'
                f'IF(E{row_num}>=70,"VERY GOOD",'
                f'IF(E{row_num}>=65,"GOOD",'
                f'IF(E{row_num}>=60,"CREDIT",'
                f'IF(E{row_num}>=55,"PASS",'
                f'IF(E{row_num}>=50,"WEAK PASS",'
                f'IF(E{row_num}>=45,"Needs Improvement",'
                f'IF(E{row_num}>=40,"Poor","FAIL")))))))))'
            )
            remarks_cell.border = border

            row_num += 1

        # Add hidden metadata for import
        ws["H1"] = "class_subject_id"
        ws["I1"] = class_subject.class_subject_id
        ws["H2"] = "assignment_id"
        ws["I2"] = assignment.assignment_id
        # Add school ID for multi-tenancy
        if user_school:
            ws["H3"] = "school_id"
            ws["I3"] = user_school.id

        # Hide the metadata columns
        ws.column_dimensions["H"].hidden = True
        ws.column_dimensions["I"].hidden = True

        # Set column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 20

        # Use BytesIO for in-memory file handling to avoid file corruption
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = output.getvalue()

        # Clean filename by removing special characters
        class_name = "".join(
            e
            for e in assignment.class_assigned.name
            if e.isalnum() or e in ["_", "-", " "]
        )
        subject_name = "".join(
            e
            for e in assignment.subject.subject_name
            if e.isalnum() or e in ["_", "-", " "]
        )

        # Create response with proper content type
        response = HttpResponse(
            file_content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Set content disposition with cleaned filename
        response["Content-Disposition"] = (
            f'attachment; filename="{class_name}_{subject_name}_scores.xlsx"'
        )

        return response

    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}", exc_info=True)
        messages.error(request, f"Error generating Excel file: {str(e)}")
        return redirect("enter_scores")


@login_required
@csrf_exempt
def import_scores(request):
    """Import scores from Excel file."""
    if request.method != "POST":
        return render(request, "import_scores.html")

    # Check if the user is a teacher
    if request.user.role != "teacher":
        messages.error(request, "Only teachers can import scores.")
        return render(request, "import_scores.html", status=403)

    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    # Get the assignment ID (for class and subject)
    assignment_id = request.POST.get("assignment_id")
    if not assignment_id:
        messages.error(request, "No assignment selected.")
        return render(request, "import_scores.html")

    # Get the teacher instance
    teacher = request.user.teacher_profile
    if not teacher:
        messages.error(request, "Teacher profile is not linked to your account.")
        return redirect("login")

    # Get the assignment with school context
    try:
        assignment_query = TeacherSubjectAssignment.objects.filter(
            id=assignment_id, teacher=teacher, is_active=True
        )

        # Apply school filter for multi-tenancy
        if user_school:
            assignment_query = assignment_query.filter(school=user_school)

        assignment = assignment_query.get()
    except TeacherSubjectAssignment.DoesNotExist:
        messages.error(request, "Invalid assignment selected.")
        return render(request, "import_scores.html")

    # Get current term with school context
    if user_school:
        current_term = Term.objects.filter(

            academic_year=assignment.academic_year,
            is_active=True, is_current=True, school=user_school
        ).first()
    else:
        current_term = Term.objects.filter(
            academic_year=assignment.academic_year,
            is_active=True, is_current=True

        ).first()

    if not current_term:
        messages.error(request, "No active term found for this academic year.")
        return render(request, "import_scores.html")

    # Get the ClassSubject instance with school context
    try:
        class_subject_query = ClassSubject.objects.filter(
            subject=assignment.subject,
            class_name=assignment.class_assigned,
            academic_year=assignment.academic_year,

            is_active=True

        )

        # Apply school filter through related objects
        if user_school:
            class_subject_query = class_subject_query.filter(
                subject__school=user_school,
                class_name__school=user_school,
                academic_year__school=user_school,
            )

        class_subject = class_subject_query.first()

        if not class_subject:
            # Create with proper school association
            class_subject = ClassSubject.objects.create(
                subject=assignment.subject,
                class_name=assignment.class_assigned,
                academic_year=assignment.academic_year,

            is_active=True

            )
    except ClassSubject.DoesNotExist:
        # Create the class subject if it doesn't exist
        class_subject = ClassSubject.objects.create(
            subject=assignment.subject,
            class_name=assignment.class_assigned,
            academic_year=assignment.academic_year,

            is_active=True

        )

    # Check if a file was uploaded
    if "file" not in request.FILES:
        messages.error(request, "No file was uploaded.")
        return render(request, "import_scores.html")

    file = request.FILES["file"]
    # Check if the file is an Excel file
    if not file.name.endswith((".xlsx", ".xls")):
        messages.error(request, "Please upload an Excel file (.xlsx or .xls).")
        return render(request, "import_scores.html")

    try:
        # Create a BytesIO object from the uploaded file
        from io import BytesIO

        file_content = BytesIO(file.read())

        # Load the workbook with data_only=True to get values instead of formulas
        wb = openpyxl.load_workbook(file_content, data_only=True)
        ws = wb.active

        # Verify the file is for the correct school (for multi-tenancy)
        if user_school and "H3" in ws and ws["H3"].value == "school_id":
            file_school_id = ws["I3"].value
            if file_school_id and int(file_school_id) != user_school.id:
                messages.error(request, "This file belongs to a different school.")
                return render(request, "import_scores.html")

        # Find the data rows (starting from row 8)
        start_row = 8
        updated_assessments = []
        updated_count = 0
        errors = []

        # Loop through the rows
        for row in range(start_row, ws.max_row + 1):
            # Get the student ID from column A
            student_id = ws.cell(row=row, column=1).value
            if not student_id:
                continue  # Skip empty rows

            # Find the student with school context
            try:
                student_query = Student.objects.filter(admission_number=student_id)

                # Apply school filter for multi-tenancy
                if user_school:
                    student_query = student_query.filter(school=user_school)

                student = student_query.get()
            except Student.DoesNotExist:
                errors.append(f"Student with ID {student_id} not found.")
                continue

            # Check if the student is in the class with school context
            student_class_query = StudentClass.objects.filter(
                student=student,
                assigned_class=assignment.class_assigned,
                is_active=True,
            )

            # Apply school filter for multi-tenancy
            if user_school:
                student_class_query = student_class_query.filter(school=user_school)

            if not student_class_query.exists():
                errors.append(
                    f"Student {student.full_name} ({student_id}) is not in this class."
                )
                continue

            # Get the scores
            try:
                class_score_cell = ws.cell(row=row, column=3)
                exam_score_cell = ws.cell(row=row, column=4)

                class_score_value = class_score_cell.value
                exam_score_value = exam_score_cell.value

                # Handle different data types for scores
                if isinstance(class_score_value, str):
                    class_score_value = class_score_value.strip()
                    class_score_value = (
                        float(class_score_value) if class_score_value else None
                    )

                if isinstance(exam_score_value, str):
                    exam_score_value = exam_score_value.strip()
                    exam_score_value = (
                        float(exam_score_value) if exam_score_value else None
                    )

                # Convert to decimal or None
                class_score = (
                    Decimal(str(class_score_value))
                    if class_score_value not in ("", None)
                    else None
                )
                exam_score = (
                    Decimal(str(exam_score_value))
                    if exam_score_value not in ("", None)
                    else None
                )

                # Validate score ranges using dynamic configuration
                scoring_config = None
                if user_school:
                    scoring_config = ScoringConfiguration.get_active_config(user_school)

                # Use dynamic maximum values if configuration exists, otherwise use defaults
                max_class_score = (
                    scoring_config.max_class_score if scoring_config else 30
                )
                max_exam_score = scoring_config.max_exam_score if scoring_config else 70

                if class_score is not None and (
                    class_score < 0 or class_score > max_class_score
                ):
                    errors.append(
                        f"Invalid class score for {student.full_name}: {class_score}. Must be between 0 and {max_class_score}."
                    )
                    continue

                if exam_score is not None and (
                    exam_score < 0 or exam_score > max_exam_score
                ):
                    errors.append(
                        f"Invalid exam score for {student.full_name}: {exam_score}. Must be between 0 and {max_exam_score}."
                    )
                    continue
            except (ValueError, InvalidOperation) as e:
                errors.append(f"Invalid score format for {student.full_name}: {str(e)}")
                continue

            # Update or create the assessment with school context
            assessment_query = Assessment.objects.filter(
                class_subject=class_subject,
                student=student,
                term=current_term,
                assessment_type="exam_score",
            )

            assessment = assessment_query.first()

            if assessment:
                # Update existing assessment
                # Clear individual components to prevent save method from overriding class_score
                assessment.individual_score = None
                assessment.class_test_score = None
                assessment.project_score = None
                assessment.group_work_score = None
                assessment.class_score = class_score
                assessment.exam_score = exam_score
                assessment.recorded_by = request.user
                assessment.save()
            else:
                # Create new assessment
                assessment = Assessment.objects.create(
                    class_subject=class_subject,
                    student=student,
                    term=current_term,
                    assessment_type="exam_score",
                    individual_score=None,  # Clear to prevent auto-calculation
                    class_test_score=None,  # Clear to prevent auto-calculation
                    project_score=None,  # Clear to prevent auto-calculation
                    group_work_score=None,  # Clear to prevent auto-calculation
                    class_score=class_score,
                    exam_score=exam_score,
                    recorded_by=request.user,
                )

            # The Assessment model's save method will automatically calculate total_score, grade, and remarks
            # No need for manual calculation here
            updated_assessments.append(assessment)
            updated_count += 1

        # Calculate positions for all students in this class_subject
        calculate_positions_optimized(class_subject, current_term)

        # Check if this is an AJAX request
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            if errors:
                return JsonResponse(
                    {
                        "success": False,
                        "message": "There were errors importing scores.",
                        "error_messages": errors[:10],  # First 10 errors
                        "total_errors": len(errors),
                    }
                )
            else:
                return JsonResponse(
                    {
                        "success": True,
                        "message": f"Successfully imported scores for {updated_count} students.",
                        "updated": updated_count,
                    }
                )

        # Show success or error message
        if updated_count > 0:
            messages.success(
                request, f"Successfully imported scores for {updated_count} students."
            )
        else:
            messages.warning(request, "No scores were imported.")

        if errors:
            for error in errors[:10]:  # Show only the first 10 errors
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(
                    request,
                    f"... and {len(errors) - 10} more errors. Check your file and try again.",
                )

        return render(request, "import_scores.html")
    except Exception as e:
        logger.error(f"Error importing scores: {str(e)}", exc_info=True)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse(
                {
                    "success": False,
                    "message": f"Error importing scores: {str(e)}",
                    "error_messages": [str(e)],
                }
            )
        messages.error(request, f"Error importing scores: {str(e)}")
        return render(request, "import_scores.html")


@login_required
def export_scores_batch(request):
    """Export scores for multiple classes assigned to a teacher."""
    if request.user.role != "teacher":
        return render(request, "errors/403.html", status=403)

    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    # Get the assignment IDs from the request
    assignment_ids_param = request.GET.get("assignment_ids", "")
    if not assignment_ids_param:
        messages.error(
            request,
            "Please select at least one class to export. Use the batch export option when no class is selected.",
        )
        return redirect("enter_scores")

    assignment_ids = assignment_ids_param.split(",")

    # Get the teacher instance
    teacher = request.user.teacher_profile
    if not teacher:
        messages.error(request, "Teacher profile is not linked to this user.")
        return redirect("login")

    try:
        # Import necessary modules
        import tempfile
        from io import BytesIO

        # Create a new workbook with better compatibility settings
        wb = openpyxl.Workbook()
        wb.iso_dates = True  # Use ISO dates for better compatibility

        # Remove the default worksheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Get current academic year and term with school context
        if user_school:
            current_academic_year = AcademicYear.objects.filter(
                is_current=True, school=user_school
            ).first()
            current_term = Term.objects.filter(
                academic_year=current_academic_year, is_current=True, school=user_school
            ).first()
        else:
            current_academic_year = AcademicYear.objects.filter(is_current=True).first()
            current_term = Term.objects.filter(
                academic_year=current_academic_year, is_current=True
            ).first()

        if not current_academic_year or not current_term:
            messages.error(request, "No active academic year or term found.")
            return redirect("enter_scores")

        # Define styles
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Process each assignment
        processed_assignments = 0

        for assignment_id in assignment_ids:
            try:
                # Get the assignment with school context
                assignment_query = TeacherSubjectAssignment.objects.filter(
                    id=assignment_id, teacher=teacher, is_active=True
                )

                # Apply school filter for multi-tenancy
                if user_school:
                    assignment_query = assignment_query.filter(school=user_school)

                assignment = assignment_query.get()

                # Get the ClassSubject instance with school context
                try:
                    class_subject_query = ClassSubject.objects.filter(
                        subject=assignment.subject,
                        class_name=assignment.class_assigned,
                        academic_year=current_academic_year,

                        is_active=True

                    )

                    # Apply school filter through related objects
                    if user_school:
                        class_subject_query = class_subject_query.filter(
                            subject__school=user_school,
                            class_name__school=user_school,
                            academic_year__school=user_school,
                        )

                    class_subject = class_subject_query.first()

                    if not class_subject:

                        # Check if there's an inactive ClassSubject that we can reactivate
                        inactive_class_subject_query = ClassSubject.objects.filter(
                            subject=assignment.subject,
                            class_name=assignment.class_assigned,
                            academic_year=current_academic_year,
                            is_active=False
                        )

                        if user_school:
                            inactive_class_subject_query = inactive_class_subject_query.filter(
                                subject__school=user_school,
                                class_name__school=user_school,
                                academic_year__school=user_school,
                            )

                        inactive_class_subject = inactive_class_subject_query.first()

                        if inactive_class_subject:
                            # Reactivate the existing assignment
                            inactive_class_subject.is_active = True
                            inactive_class_subject.save()
                            class_subject = inactive_class_subject
                        else:
                            # Skip this assignment - don't create automatically
                            continue
                except ClassSubject.DoesNotExist:
                    # Skip this assignment - don't create automatically
                    continue


                # Fetch all active students in the assigned class with school context
                student_classes_query = StudentClass.objects.filter(
                    assigned_class=assignment.class_assigned, is_active=True
                )

                # Apply school filter for multi-tenancy
                if user_school:
                    student_classes_query = student_classes_query.filter(
                        school=user_school
                    )

                student_classes = student_classes_query.select_related("student")

                # Skip if no students found
                if not student_classes.exists():
                    continue

                # Fetch existing assessments for these students

                # Exclude mock exam assessments from batch export
                assessments = Assessment.objects.filter(
                    class_subject=class_subject, term=current_term
                ).exclude(assessment_type='mock_exam').select_related("student")


                # Create a dictionary for quick lookup of assessments by student ID
                student_assessments = {
                    assessment.student.id: assessment for assessment in assessments
                }

                # Create a worksheet for this class/subject
                sheet_name = f"{assignment.class_assigned.name}_{assignment.subject.subject_name}"

                # Remove invalid characters for Excel sheet names
                invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
                for char in invalid_chars:
                    sheet_name = sheet_name.replace(char, "_")

                # Truncate sheet name if too long (Excel limit is 31 chars)
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]

                # Create sheet and add to workbook
                ws = wb.create_sheet(title=sheet_name)

                # Get scoring configuration for dynamic headers
                scoring_config = None
                if user_school:
                    scoring_config = ScoringConfiguration.get_active_config(user_school)

                # Use dynamic values if configuration exists, otherwise use defaults
                class_score_percentage = (
                    scoring_config.class_score_percentage if scoring_config else 30
                )
                exam_score_percentage = (
                    scoring_config.exam_score_percentage if scoring_config else 70
                )
                max_class_score = (
                    scoring_config.max_class_score if scoring_config else 30
                )
                max_exam_score = scoring_config.max_exam_score if scoring_config else 70

                # Get school information for the header
                if user_school:
                    school_name = user_school.name
                else:
                    # Try to get the active school or use a default
                    active_school = SchoolInformation.get_active()
                    school_name = active_school.name if active_school else "SCHOOL NAME"

                # Write header information
                ws["A1"] = f"School Name: {school_name}"
                ws["A2"] = f"Class: {assignment.class_assigned.name}"
                ws["A3"] = f"Subject: {assignment.subject.subject_name}"
                ws["A4"] = f"Term: {current_term.term_number}"
                ws["A5"] = f"Academic Year: {current_academic_year}"

                # Make headers bold
                for row in range(1, 6):
                    ws.cell(row=row, column=1).font = Font(bold=True)

                # Define column headers with dynamic percentages
                headers = [
                    "Student ID",
                    "Student Name",
                    f"Class Score ({class_score_percentage}%)",
                    f"Exam Score ({exam_score_percentage}%)",
                    "Total Score",
                    "Grade",
                    "Remarks",
                ]

                # Apply headers to row 7
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=7, column=col_num)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

                # Add calculation instructions with dynamic max values
                ws["A6"] = (
                    f"Instructions: Edit only the Class Score (max: {max_class_score}) and Exam Score (max: {max_exam_score}) columns."
                )
                ws.merge_cells("A6:G6")
                ws["A6"].font = Font(bold=True, color="FF0000")
                ws["A6"].alignment = Alignment(horizontal="center")

                # Add data rows
                row_num = 8
                for student_class in student_classes:
                    student = student_class.student
                    assessment = student_assessments.get(student.id)

                    # Write student data
                    ws.cell(
                        row=row_num, column=1, value=student.admission_number
                    ).border = border
                    ws.cell(row=row_num, column=2, value=student.full_name).border = (
                        border
                    )

                    # Write assessment data if it exists
                    class_score_cell = ws.cell(row=row_num, column=3)
                    class_score_cell.value = (
                        assessment.class_score if assessment else ""
                    )
                    class_score_cell.border = border
                    # Set number format for better compatibility
                    class_score_cell.number_format = "0.00"

                    exam_score_cell = ws.cell(row=row_num, column=4)
                    exam_score_cell.value = assessment.exam_score if assessment else ""
                    exam_score_cell.border = border
                    # Set number format for better compatibility
                    exam_score_cell.number_format = "0.00"

                    # Formula for total score
                    total_cell = ws.cell(row=row_num, column=5)
                    total_cell.value = f'=IF(AND(C{row_num}<>"",D{row_num}<>""),C{row_num}+D{row_num},"")'
                    total_cell.border = border
                    # Set number format for better compatibility
                    total_cell.number_format = "0.00"

                    # Formula for grade - use simpler formula for compatibility
                    grade_cell = ws.cell(row=row_num, column=6)
                    grade_cell.value = (
                        f'=IF(E{row_num}="","",'
                        f'IF(E{row_num}>=75,"A1",'
                        f'IF(E{row_num}>=70,"B2",'
                        f'IF(E{row_num}>=65,"B3",'
                        f'IF(E{row_num}>=60,"C4",'
                        f'IF(E{row_num}>=55,"C5",'
                        f'IF(E{row_num}>=50,"C6",'
                        f'IF(E{row_num}>=45,"D7",'
                        f'IF(E{row_num}>=40,"E8","F9")))))))))'
                    )
                    grade_cell.border = border

                    # Formula for remarks - use simpler formula for compatibility
                    remarks_cell = ws.cell(row=row_num, column=7)
                    remarks_cell.value = (
                        f'=IF(E{row_num}="","",'
                        f'IF(E{row_num}>=75,"EXCELLENT",'
                        f'IF(E{row_num}>=70,"VERY GOOD",'
                        f'IF(E{row_num}>=65,"GOOD",'
                        f'IF(E{row_num}>=60,"CREDIT",'
                        f'IF(E{row_num}>=55,"PASS",'
                        f'IF(E{row_num}>=50,"WEAK PASS",'
                        f'IF(E{row_num}>=45,"Needs Improvement",'
                        f'IF(E{row_num}>=40,"Poor","FAIL")))))))))'
                    )
                    remarks_cell.border = border

                    row_num += 1

                # Add hidden metadata for import
                ws["H1"] = "class_subject_id"
                ws["I1"] = class_subject.class_subject_id
                ws["H2"] = "assignment_id"
                ws["I2"] = assignment.assignment_id
                # Add school ID for multi-tenancy
                if user_school:
                    ws["H3"] = "school_id"
                    ws["I3"] = user_school.id

                # Hide the metadata columns
                ws.column_dimensions["H"].hidden = True
                ws.column_dimensions["I"].hidden = True

                # Set column widths
                ws.column_dimensions["A"].width = 15
                ws.column_dimensions["B"].width = 30
                ws.column_dimensions["C"].width = 15
                ws.column_dimensions["D"].width = 15
                ws.column_dimensions["E"].width = 15
                ws.column_dimensions["F"].width = 10
                ws.column_dimensions["G"].width = 20

                processed_assignments += 1

            except TeacherSubjectAssignment.DoesNotExist:
                # Skip this assignment
                continue

        # If no assignments were processed, return with an error
        if processed_assignments == 0:
            messages.error(request, "No valid assignments found to export.")
            return redirect("enter_scores")

        # Use BytesIO for in-memory file handling to avoid file corruption
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = output.getvalue()

        # Clean filename
        academic_year_name = "".join(
            e
            for e in str(current_academic_year.name)
            if e.isalnum() or e in ["_", "-", " "]
        )
        term_number = str(current_term.term_number)

        # Create response with proper content type
        response = HttpResponse(
            file_content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Set content disposition with cleaned filename
        response["Content-Disposition"] = (
            f'attachment; filename="batch_scores_{academic_year_name}_{term_number}.xlsx"'
        )

        return response

    except Exception as e:
        logger.error(f"Error generating batch Excel file: {str(e)}", exc_info=True)
        messages.error(request, f"Error generating batch Excel file: {str(e)}")
        return redirect("enter_scores")


@login_required
@csrf_exempt
def import_scores_batch(request):
    """Import scores for multiple classes from a batch Excel file with multiple sheets."""
    if request.user.role != "teacher":
        return JsonResponse(
            {"success": False, "error": "Permission denied"}, status=403
        )

    if request.method != "POST" or "batchFile" not in request.FILES:
        return JsonResponse({"success": False, "error": "No file uploaded"}, status=400)

    file = request.FILES["batchFile"]

    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    # Get current term for the user's school
    if user_school:
        current_term = Term.objects.filter(is_current=True, school=user_school).first()
    else:
        current_term = Term.objects.filter(is_current=True).first()

    if not current_term:
        return JsonResponse({"error": "No current term is set"}, status=400)

    # Get the teacher instance
    teacher = request.user.teacher_profile
    if not teacher:
        return JsonResponse(
            {"success": False, "error": "Teacher profile not found"}, status=400
        )

    # Validate the file extension
    if not (file.name.endswith(".xlsx") or file.name.endswith(".xls")):
        return JsonResponse(
            {"success": False, "error": "File must be an Excel spreadsheet"}, status=400
        )

    try:
        # Create a BytesIO object from the uploaded file
        from io import BytesIO

        file_content = BytesIO(file.read())

        # Load the workbook with data_only=True to get values instead of formulas
        wb = openpyxl.load_workbook(file_content, data_only=True)

        results = []
        errors = []

        # Process each sheet in the workbook
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Verify the file is for the correct school (for multi-tenancy)
            if user_school and "H3" in ws and ws["H3"].value == "school_id":
                file_school_id = ws["I3"].value
                if file_school_id and int(file_school_id) != user_school.id:
                    errors.append(f"Sheet '{sheet_name}' belongs to a different school")
                    continue

            # Verify the sheet has the expected metadata
            try:
                class_subject_id = ws["I1"].value
                file_assignment_id = ws["I2"].value

                if not class_subject_id or not file_assignment_id:
                    errors.append(f"Sheet '{sheet_name}' has invalid template format")
                    continue

                # Get the class subject with school context
                try:
                    class_subject_query = ClassSubject.objects.filter(
                        class_subject_id=class_subject_id
                    )

                    # Apply school filter for multi-tenancy
                    if user_school:
                        class_subject_query = class_subject_query.filter(
                            subject__school=user_school,
                            class_name__school=user_school,
                            academic_year__school=user_school,
                        )

                    class_subject = class_subject_query.get()
                except ClassSubject.DoesNotExist:
                    errors.append(
                        f"Sheet '{sheet_name}' references a class subject that doesn't exist"
                    )
                    continue

                # Verify the teacher is assigned to this class with school context
                try:
                    assignment_query = TeacherSubjectAssignment.objects.filter(
                        assignment_id=file_assignment_id,
                        teacher=teacher,
                        is_active=True,
                    )

                    # Apply school filter for multi-tenancy
                    if user_school:
                        assignment_query = assignment_query.filter(school=user_school)

                    assignment = assignment_query.get()
                except TeacherSubjectAssignment.DoesNotExist:
                    errors.append(
                        f"Sheet '{sheet_name}' references a class assignment that isn't yours"
                    )
                    continue

                # Start processing data from row 8 (after headers)
                updated_count = 0
                updated_assessments = []  # Store updated assessments for recalculating

                for row in range(8, ws.max_row + 1):
                    student_id = ws.cell(row=row, column=1).value
                    class_score_value = ws.cell(row=row, column=3).value
                    exam_score_value = ws.cell(row=row, column=4).value

                    # Skip if student ID or both scores are empty
                    if not student_id or (
                        class_score_value is None and exam_score_value is None
                    ):
                        continue

                    # Handle different data types for scores
                    if isinstance(class_score_value, str):
                        class_score_value = class_score_value.strip()
                        class_score_value = (
                            float(class_score_value) if class_score_value else None
                        )

                    if isinstance(exam_score_value, str):
                        exam_score_value = exam_score_value.strip()
                        exam_score_value = (
                            float(exam_score_value) if exam_score_value else None
                        )

                    # Try to convert scores to float
                    try:
                        class_score = (
                            float(class_score_value)
                            if class_score_value is not None
                            else None
                        )
                        exam_score = (
                            float(exam_score_value)
                            if exam_score_value is not None
                            else None
                        )
                    except (ValueError, TypeError):
                        continue

                    # Validate score ranges using dynamic configuration
                    scoring_config = None
                    if user_school:
                        scoring_config = ScoringConfiguration.get_active_config(
                            user_school
                        )

                    # Use dynamic maximum values if configuration exists, otherwise use defaults
                    max_class_score = (
                        scoring_config.max_class_score if scoring_config else 30
                    )
                    max_exam_score = (
                        scoring_config.max_exam_score if scoring_config else 70
                    )

                    if (
                        class_score is not None
                        and (class_score < 0 or class_score > max_class_score)
                    ) or (
                        exam_score is not None
                        and (exam_score < 0 or exam_score > max_exam_score)
                    ):
                        continue

                    # Find the student with school context
                    try:
                        student_query = Student.objects.filter(
                            admission_number=student_id
                        )

                        # Apply school filter for multi-tenancy
                        if user_school:
                            student_query = student_query.filter(school=user_school)

                        student = student_query.get()
                    except Student.DoesNotExist:
                        continue

                    # Check if student is in this class with school context
                    student_class_query = StudentClass.objects.filter(
                        student=student,
                        assigned_class=assignment.class_assigned,
                        is_active=True,
                    )

                    # Apply school filter for multi-tenancy
                    if user_school:
                        student_class_query = student_class_query.filter(
                            school=user_school
                        )

                    if not student_class_query.exists():
                        continue

                    # Update or create the assessment with school context
                    assessment_query = Assessment.objects.filter(
                        class_subject=class_subject,
                        student=student,
                        term=current_term,
                        assessment_type="exam_score",
                    )

                    assessment = assessment_query.first()

                    if assessment:
                        # Update existing assessment
                        # Clear individual components to prevent save method from overriding class_score
                        assessment.individual_score = None
                        assessment.class_test_score = None
                        assessment.project_score = None
                        assessment.group_work_score = None
                        assessment.class_score = class_score
                        assessment.exam_score = exam_score
                        assessment.recorded_by = request.user
                        assessment.save()
                    else:
                        # Create new assessment
                        assessment = Assessment.objects.create(
                            class_subject=class_subject,
                            student=student,
                            term=current_term,
                            assessment_type="exam_score",
                            individual_score=None,  # Clear to prevent auto-calculation
                            class_test_score=None,  # Clear to prevent auto-calculation
                            project_score=None,  # Clear to prevent auto-calculation
                            group_work_score=None,  # Clear to prevent auto-calculation
                            class_score=class_score,
                            exam_score=exam_score,
                            recorded_by=request.user,
                        )

                    # The Assessment model's save method will automatically calculate total_score, grade, and remarks
                    # No need for manual calculation here
                    updated_assessments.append(assessment)

                    updated_count += 1

                # Recalculate positions for all students in this class_subject
                if hasattr(Assessment, "calculate_positions"):
                    Assessment.calculate_positions(class_subject)
                else:
                    # Use optimized position calculation
                    calculate_positions_optimized(class_subject, current_term)

                # Add the result for this sheet
                results.append(
                    {
                        "class_name": assignment.class_assigned.name,
                        "subject_name": assignment.subject.subject_name,
                        "processed": updated_count,
                    }
                )

            except Exception as e:
                errors.append(f"Error processing sheet '{sheet_name}': {str(e)}")
                continue

        if not results:
            return JsonResponse(
                {
                    "success": False,
                    "error": "No scores were successfully imported",
                    "errors": errors,
                }
            )

        return JsonResponse({"success": True, "results": results, "errors": errors})

    except Exception as e:
        import traceback

        return JsonResponse(
            {"success": False, "error": str(e), "traceback": traceback.format_exc()},
            status=500,
        )
