"""
Mock Exam Export Views
Handles single and batch export of mock exam scores to Excel.
"""

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal, InvalidOperation
import logging
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from ..models import (
    User,
    Teacher,
    Student,
    StudentClass,
    TeacherSubjectAssignment,
    ClassSubject,
    Assessment,
    AcademicYear,
    MockExam,
    SchoolInformation,
    Class,
    Subject,
)

logger = logging.getLogger(__name__)


def get_user_school(user):
    """Get the school associated with the current user."""
    if hasattr(user, 'school') and user.school:
        return user.school
    
    # Try to get school from teacher profile
    if user.role == "teacher":
        try:
            teacher = Teacher.objects.get(user=user)
            if teacher.school:
                return teacher.school
        except Teacher.DoesNotExist:
            pass
    
    return None


def sanitize_excel_sheet_name(name):
    """
    Sanitize a name to be used as an Excel sheet name.
    Excel sheet names have the following restrictions:
    - Maximum 31 characters
    - Cannot contain: [ ] : * ? / \ 
    - Cannot start or end with apostrophe
    - Cannot be empty
    """
    if not name:
        return "Sheet"
    
    # Remove invalid characters
    invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
    for char in invalid_chars:
        name = name.replace(char, "_")
    
    # Remove leading/trailing apostrophes
    name = name.strip("'")
    
    # Limit to 31 characters
    if len(name) > 31:
        name = name[:31]
    
    # Ensure not empty
    if not name:
        name = "Sheet"
    
    return name


@login_required
def export_mock_exam_scores(request):
    """
    Export mock exam score template for a single class.
    Creates an Excel template that can be used for importing mock exam scores.
    """
    # Check if the user is authorized
    if request.user.role not in ["teacher", "admin"]:
        messages.error(request, "Only teachers and administrators can export mock exam scores.")
        return redirect("mock_exam_entry")
    
    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)
    
    if not user_school:
        messages.error(request, "No school associated with your account.")
        return redirect("mock_exam_entry")
    
    # Get the assignment ID from the request
    assignment_id = request.GET.get("assignment_id")
    if not assignment_id:
        messages.error(request, "No assignment selected for export.")
        return redirect("mock_exam_entry")
    
    # Get the teacher instance (for teachers)
    teacher = getattr(request.user, "teacher_profile", None)
    if request.user.role == "teacher" and not teacher:
        messages.error(request, "Teacher profile is not linked to this user.")
        return redirect("login")
    
    # Get current academic year
    current_academic_year = AcademicYear.objects.filter(
        school=user_school,
        is_current=True
    ).first()
    
    if not current_academic_year:
        messages.error(request, "No active academic year found.")
        return redirect("mock_exam_entry")
    
    # Get default active mock exam
    default_mock_exam = MockExam.objects.filter(
        school=user_school,
        academic_year=current_academic_year,
        is_active=True
    ).order_by('-exam_date', '-created_at').first()
    
    if not default_mock_exam:
        messages.error(request, "No active mock exam found. Please create an active mock exam first.")
        return redirect("mock_exam_entry")
    
    # Get the assignment with school context
    try:
        if request.user.role == "teacher" and teacher:
            assignment_query = TeacherSubjectAssignment.objects.filter(
                id=assignment_id, teacher=teacher, is_active=True
            )
        elif request.user.role == "admin":
            assignment_query = TeacherSubjectAssignment.objects.filter(
                id=assignment_id, is_active=True
            )
        else:
            messages.error(request, "Unauthorized access to assignment.")
            return redirect("mock_exam_entry")
        
        # Apply school filter for multi-tenancy
        if user_school:
            assignment_query = assignment_query.filter(school=user_school)
        
        assignment = assignment_query.get()
        
        # Check if the ClassSubject is active
        class_subject = ClassSubject.objects.filter(
            subject=assignment.subject,
            class_name=assignment.class_assigned,
            academic_year=assignment.academic_year,
            is_active=True
        ).first()
        
        if not class_subject:
            messages.error(request, f"Subject '{assignment.subject.subject_name}' is not currently assigned to class '{assignment.class_assigned.name}'. Cannot export scores for removed subjects.")
            return redirect("mock_exam_entry")
            
    except TeacherSubjectAssignment.DoesNotExist:
        messages.error(request, "Invalid assignment selected.")
        return redirect("mock_exam_entry")
    
    # Get students in class with school context
    student_classes_query = StudentClass.objects.filter(
        assigned_class=assignment.class_assigned,
        is_active=True,
        assigned_class__academic_year=current_academic_year,
        school=user_school,
        student__school=user_school
    ).select_related("student")
    
    student_classes = student_classes_query.order_by("student__full_name")
    
    try:
        # Create a BytesIO object for the Excel file
        output = BytesIO()
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Create sanitized sheet name
        sheet_name = (
            f"{assignment.class_assigned.name} - {assignment.subject.subject_name}"
        )
        ws.title = sanitize_excel_sheet_name(sheet_name)
        
        # Define styles
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        
        # Add school information
        ws["A1"] = f"School: {user_school.name}"
        ws["A2"] = f"Class: {assignment.class_assigned.name}"
        ws["A3"] = f"Subject: {assignment.subject.subject_name}"
        ws["A4"] = f"Academic Year: {assignment.academic_year.name}"
        ws["A5"] = f"Mock Exam: {default_mock_exam.name}"
        ws["A6"] = f"Exam Date: {default_mock_exam.exam_date.strftime('%B %d, %Y')}"
        
        # Add metadata for import validation
        ws["H1"] = "class_subject_id"
        ws["I1"] = class_subject.class_subject_id if class_subject else ""
        ws["H2"] = "assignment_id"
        ws["I2"] = assignment.assignment_id if hasattr(assignment, 'assignment_id') else assignment.id
        ws["H3"] = "school_id"
        ws["I3"] = user_school.id
        ws["H4"] = "mock_exam_id"
        ws["I4"] = default_mock_exam.id
        
        # Make headers bold
        for row in range(1, 7):
            ws.cell(row=row, column=1).font = Font(bold=True)
        
        # Add instructions
        instructions = (
            f"Instructions: Edit the Raw Score column (C) with values between 0 and 100. "
            f"Total Score, Grade, Remarks, and Position are calculated automatically."
        )
        ws["A7"] = instructions
        ws.merge_cells("A7:F7")
        ws["A7"].font = Font(bold=True, color="FF0000")
        ws["A7"].alignment = Alignment(horizontal="left", wrap_text=True)
        
        # Define column headers for mock exams
        headers = [
            "Student ID",
            "Student Name",
            "Raw Score (0-100)",
            "Total Score (Calculated)",
            "Grade (Calculated)",
            "Remarks (Calculated)",
            "Position (Calculated)",
        ]
        
        # Apply headers to row 8
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Get existing assessments for all students in this class and mock exam
        assessments = Assessment.objects.filter(
            class_subject=class_subject,
            student__in=[sc.student for sc in student_classes],
            assessment_type='mock_exam',
            mock_exam=default_mock_exam
        ).select_related("student")
        
        # Create a dictionary for quick lookup
        student_assessments = {
            assessment.student.id: assessment for assessment in assessments
        }
        
        # Add data rows
        row_num = 9
        for student_class in student_classes:
            student = student_class.student
            assessment = student_assessments.get(student.id)
            
            # Student ID and Name
            ws.cell(row=row_num, column=1, value=student.admission_number or student.student_id)
            ws.cell(row=row_num, column=2, value=student.full_name)
            
            # Raw score (editable)
            if assessment:
                ws.cell(row=row_num, column=3, value=assessment.raw_exam_score or "")
            else:
                ws.cell(row=row_num, column=3, value="")
            
            # Calculated fields (read-only, for reference)
            if assessment:
                ws.cell(row=row_num, column=4, value=assessment.total_score or "")
                ws.cell(row=row_num, column=5, value=assessment.grade or "")
                ws.cell(row=row_num, column=6, value=assessment.remarks or "")
                ws.cell(row=row_num, column=7, value=assessment.position or "")
            else:
                ws.cell(row=row_num, column=4, value="")
                ws.cell(row=row_num, column=5, value="")
                ws.cell(row=row_num, column=6, value="")
                ws.cell(row=row_num, column=7, value="")
            
            # Apply borders to all cells in the row
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).border = border
            
            row_num += 1
        
        # Adjust column widths
        column_widths = [15, 30, 18, 20, 15, 20, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Save the workbook to BytesIO
        wb.save(output)
        output.seek(0)
        
        # Create response
        filename = f"mock_exam_scores_{assignment.class_assigned.name}_{assignment.subject.subject_name}_{default_mock_exam.name.replace(' ', '_')}.xlsx"
        # Sanitize filename
        invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
        for char in invalid_chars:
            filename = filename.replace(char, "_")
        
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        
        return response
    
    except Exception as e:
        logger.error(f"Error exporting mock exam scores: {str(e)}", exc_info=True)
        messages.error(request, f"Error exporting mock exam scores: {str(e)}")
        return redirect("mock_exam_entry")


@login_required
def export_mock_exam_scores_batch(request):
    """
    Export mock exam score templates for multiple classes.
    Creates an Excel file with multiple sheets for batch import.
    """
    # Check if the user is authorized
    if request.user.role not in ["teacher", "admin"]:
        messages.error(request, "Only teachers and administrators can export mock exam scores.")
        return redirect("mock_exam_entry")
    
    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)
    
    if not user_school:
        messages.error(request, "No school associated with your account.")
        return redirect("mock_exam_entry")
    
    # Get the teacher instance (for teachers)
    teacher = getattr(request.user, "teacher_profile", None)
    if request.user.role == "teacher" and not teacher:
        messages.error(request, "Teacher profile is not linked to this user.")
        return redirect("login")
    
    # Get current academic year
    if user_school:
        current_academic_year = AcademicYear.objects.filter(
            is_current=True, school=user_school
        ).first()
    else:
        current_academic_year = AcademicYear.objects.filter(is_current=True).first()
    
    if not current_academic_year:
        messages.error(request, "No active academic year found.")
        return redirect("mock_exam_entry")
    
    # Get default active mock exam
    default_mock_exam = MockExam.objects.filter(
        school=user_school,
        academic_year=current_academic_year,
        is_active=True
    ).order_by('-exam_date', '-created_at').first()
    
    if not default_mock_exam:
        messages.error(request, "No active mock exam found. Please create an active mock exam first.")
        return redirect("mock_exam_entry")
    
    # Check if specific assignment IDs are provided
    assignment_ids = request.GET.get("assignment_ids")
    if assignment_ids:
        # Parse the comma-separated assignment IDs
        try:
            assignment_id_list = [int(id.strip()) for id in assignment_ids.split(",")]
        except ValueError:
            messages.error(request, "Invalid assignment IDs provided.")
            return redirect("mock_exam_entry")
        
        # Get specific assignments
        assignments_query = TeacherSubjectAssignment.objects.filter(
            id__in=assignment_id_list, is_active=True
        )
        
        # Apply authorization check
        if request.user.role == "teacher" and teacher:
            assignments_query = assignments_query.filter(teacher=teacher)
        elif request.user.role != "admin":
            messages.error(request, "Unauthorized access.")
            return redirect("mock_exam_entry")
    else:
        # Get all assignments for the teacher (original behavior)
        if request.user.role == "teacher" and teacher:
            assignments_query = TeacherSubjectAssignment.objects.filter(
                teacher=teacher, is_active=True, academic_year=current_academic_year
            )
        elif request.user.role == "admin":
            assignments_query = TeacherSubjectAssignment.objects.filter(
                is_active=True, academic_year=current_academic_year
            )
        else:
            messages.error(request, "Unauthorized access.")
            return redirect("mock_exam_entry")
    
    # Apply school filter for multi-tenancy
    if user_school:
        assignments_query = assignments_query.filter(school=user_school)
    
    # Filter out assignments where ClassSubject is not active
    # Get active class-subject combinations
    active_class_subjects = ClassSubject.objects.filter(
        academic_year=current_academic_year, is_active=True
    ).values_list('class_name_id', 'subject_id')
    
    # Filter assignments to only include those with active ClassSubject
    filtered_assignments = []
    for assignment in assignments_query.select_related("class_assigned", "subject", "academic_year"):
        if (assignment.class_assigned.id, assignment.subject.id) in active_class_subjects:
            filtered_assignments.append(assignment)
    
    assignments = filtered_assignments
    
    if not assignments:
        messages.error(request, "No assignments found for export.")
        return redirect("mock_exam_entry")
    
    try:
        # Create a BytesIO object for the Excel file
        output = BytesIO()
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Remove the default worksheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # Define styles
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        
        # Process each assignment
        processed_assignments = 0
        
        for assignment in assignments:
            # Create sanitized sheet name
            sheet_name = (
                f"{assignment.class_assigned.name} - {assignment.subject.subject_name}"
            )
            sanitized_sheet_name = sanitize_excel_sheet_name(sheet_name)
            
            # Create worksheet
            ws = wb.create_sheet(title=sanitized_sheet_name)
            
            # Add school information
            ws["A1"] = f"School: {user_school.name}"
            ws["A2"] = f"Class: {assignment.class_assigned.name}"
            ws["A3"] = f"Subject: {assignment.subject.subject_name}"
            ws["A4"] = f"Academic Year: {assignment.academic_year.name}"
            ws["A5"] = f"Mock Exam: {default_mock_exam.name}"
            ws["A6"] = f"Exam Date: {default_mock_exam.exam_date.strftime('%B %d, %Y')}"
            
            # Make headers bold
            for row in range(1, 7):
                ws.cell(row=row, column=1).font = Font(bold=True)
            
            # Add instructions
            instructions = (
                f"Instructions: Edit the Raw Score column (C) with values between 0 and 100. "
                f"Total Score, Grade, Remarks, and Position are calculated automatically."
            )
            ws["A7"] = instructions
            ws.merge_cells("A7:F7")
            ws["A7"].font = Font(bold=True, color="FF0000")
            ws["A7"].alignment = Alignment(horizontal="left", wrap_text=True)
            
            # Define column headers for mock exams
            headers = [
                "Student ID",
                "Student Name",
                "Raw Score (0-100)",
                "Total Score (Calculated)",
                "Grade (Calculated)",
                "Remarks (Calculated)",
                "Position (Calculated)",
            ]
            
            # Apply headers to row 8
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=8, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
            
            # Get or create ClassSubject
            try:
                class_subject_query = ClassSubject.objects.filter(
                    subject=assignment.subject,
                    class_name=assignment.class_assigned,
                    academic_year=assignment.academic_year,
                    is_active=True
                )
                
                # Apply school filter through related objects
                if user_school:
                    class_subject_query = class_subject_query.filter(
                        subject__school=user_school,
                        class_name__school=user_school,
                        academic_year__school=user_school,
                    )
                
                class_subject = class_subject_query.first()
                
                if not class_subject:
                    logger.error(f"No ClassSubject assignment found for {assignment}")
                    continue
            except Exception as e:
                logger.error(
                    f"Error creating/retrieving ClassSubject for {assignment}: {str(e)}"
                )
                continue
            
            # Add metadata for import validation (after class_subject is retrieved)
            ws["H1"] = "class_subject_id"
            ws["I1"] = class_subject.class_subject_id if class_subject else ""
            ws["H2"] = "assignment_id"
            ws["I2"] = assignment.assignment_id if hasattr(assignment, 'assignment_id') else assignment.id
            ws["H3"] = "school_id"
            ws["I3"] = user_school.id
            ws["H4"] = "mock_exam_id"
            ws["I4"] = default_mock_exam.id
            
            # Get students in class with school context
            student_classes_query = StudentClass.objects.filter(
                assigned_class=assignment.class_assigned,
                is_active=True,
                assigned_class__academic_year=current_academic_year,
                school=user_school,
                student__school=user_school
            ).select_related("student")
            
            student_classes = student_classes_query.order_by("student__full_name")
            
            # Get existing assessments for all students in this class and mock exam
            assessments = Assessment.objects.filter(
                class_subject=class_subject,
                student__in=[sc.student for sc in student_classes],
                assessment_type='mock_exam',
                mock_exam=default_mock_exam
            ).select_related("student")
            
            # Create a dictionary for quick lookup
            student_assessments = {
                assessment.student.id: assessment for assessment in assessments
            }
            
            # Add data rows
            row_num = 9
            for student_class in student_classes:
                student = student_class.student
                assessment = student_assessments.get(student.id)
                
                # Student ID and Name
                ws.cell(row=row_num, column=1, value=student.admission_number or student.student_id)
                ws.cell(row=row_num, column=2, value=student.full_name)
                
                # Raw score (editable)
                if assessment:
                    ws.cell(row=row_num, column=3, value=assessment.raw_exam_score or "")
                else:
                    ws.cell(row=row_num, column=3, value="")
                
                # Calculated fields (read-only, for reference)
                if assessment:
                    ws.cell(row=row_num, column=4, value=assessment.total_score or "")
                    ws.cell(row=row_num, column=5, value=assessment.grade or "")
                    ws.cell(row=row_num, column=6, value=assessment.remarks or "")
                    ws.cell(row=row_num, column=7, value=assessment.position or "")
                else:
                    ws.cell(row=row_num, column=4, value="")
                    ws.cell(row=row_num, column=5, value="")
                    ws.cell(row=row_num, column=6, value="")
                    ws.cell(row=row_num, column=7, value="")
                
                # Apply borders to all cells in the row
                for col in range(1, 8):
                    ws.cell(row=row_num, column=col).border = border
                
                row_num += 1
            
            # Adjust column widths
            column_widths = [15, 30, 18, 20, 15, 20, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            processed_assignments += 1
        
        # Save the workbook to BytesIO
        wb.save(output)
        output.seek(0)
        
        # Create response
        filename = f"mock_exam_scores_batch_{default_mock_exam.name.replace(' ', '_')}_{processed_assignments}_classes.xlsx"
        # Sanitize filename
        invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
        for char in invalid_chars:
            filename = filename.replace(char, "_")
        
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        
        return response
    
    except Exception as e:
        logger.error(f"Error exporting mock exam batch scores: {str(e)}", exc_info=True)
        messages.error(request, f"Error exporting batch scores: {str(e)}")
        return redirect("mock_exam_entry")


@login_required
@csrf_exempt
def import_mock_exam_scores(request):
    """
    Import mock exam scores from Excel file for a single class.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Only POST method allowed"}, status=405)

    # Check if the user is authorized
    if request.user.role not in ["teacher", "admin"]:
        return JsonResponse(
            {"error": "Only teachers and administrators can import mock exam scores."}, status=403
        )

    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    if not user_school:
        return JsonResponse({"error": "No school associated with your account."}, status=400)

    # Get current academic year
    current_academic_year = AcademicYear.objects.filter(
        school=user_school, is_current=True
    ).first()

    if not current_academic_year:
        return JsonResponse({"error": "No active academic year found."}, status=400)

    # Get default active mock exam
    default_mock_exam = MockExam.objects.filter(
        school=user_school,
        academic_year=current_academic_year,
        is_active=True
    ).order_by('-exam_date', '-created_at').first()

    if not default_mock_exam:
        return JsonResponse(
            {"error": "No active mock exam found. Please create an active mock exam first."}, status=400
        )

    # Get the assignment ID
    assignment_id = request.POST.get("assignment_id")
    if not assignment_id:
        return JsonResponse({"error": "No assignment selected."}, status=400)

    # Get the teacher instance (for teachers)
    teacher = getattr(request.user, "teacher_profile", None)
    if request.user.role == "teacher" and not teacher:
        return JsonResponse(
            {"error": "Teacher profile is not linked to your account."}, status=403
        )

    # Get the assignment with school context
    try:
        if request.user.role == "teacher" and teacher:
            assignment_query = TeacherSubjectAssignment.objects.filter(
                id=assignment_id, teacher=teacher, is_active=True
            )
        elif request.user.role == "admin":
            assignment_query = TeacherSubjectAssignment.objects.filter(
                id=assignment_id, is_active=True
            )
        else:
            return JsonResponse({"error": "Unauthorized access."}, status=403)

        # Apply school filter for multi-tenancy
        if user_school:
            assignment_query = assignment_query.filter(school=user_school)

        assignment = assignment_query.get()
    except TeacherSubjectAssignment.DoesNotExist:
        return JsonResponse({"error": "Invalid assignment selected."}, status=400)

    # Check if a file was uploaded
    if "scoreFile" not in request.FILES:
        return JsonResponse({"error": "No file was uploaded."}, status=400)

    file = request.FILES["scoreFile"]
    # Check if the file is an Excel file
    if not file.name.endswith((".xlsx", ".xls")):
        return JsonResponse(
            {"error": "Please upload an Excel file (.xlsx or .xls)."}, status=400
        )

    try:
        # Create a BytesIO object from the uploaded file
        file_content = BytesIO(file.read())

        # Load the workbook with data_only=True to get values instead of formulas
        wb = openpyxl.load_workbook(file_content, data_only=True)
        ws = wb.active

        # Verify the file is for the correct school and mock exam
        if user_school and "H3" in ws and ws["H3"].value == "school_id":
            file_school_id = ws["I3"].value
            if file_school_id and int(file_school_id) != user_school.id:
                return JsonResponse({"error": "This file belongs to a different school."}, status=400)

        # Check mock exam ID from file
        file_mock_exam_id = None
        if "H4" in ws and ws["H4"].value == "mock_exam_id":
            file_mock_exam_id = ws["I4"].value

        # Use mock exam from file if available, otherwise use default
        mock_exam = default_mock_exam
        if file_mock_exam_id:
            try:
                mock_exam = MockExam.objects.get(
                    id=file_mock_exam_id, school=user_school, is_active=True
                )
            except MockExam.DoesNotExist:
                # Use default if file mock exam not found
                pass

        # Get the ClassSubject instance
        class_subject = ClassSubject.objects.filter(
            subject=assignment.subject,
            class_name=assignment.class_assigned,
            academic_year=assignment.academic_year,
            is_active=True
        ).first()

        if not class_subject:
            return JsonResponse(
                {
                    "error": f"Subject '{assignment.subject.subject_name}' is not assigned to class '{assignment.class_assigned.name}'. Please assign this subject to the class first."
                },
                status=400,
            )

        # Find the data rows (starting from row 9, after headers)
        start_row = 9
        updated_assessments = []
        updated_count = 0
        errors = []

        # Process each row
        for row in range(start_row, ws.max_row + 1):
            # Get the student ID from column A
            student_id = ws.cell(row=row, column=1).value
            if not student_id:
                continue  # Skip empty rows

            # Find the student with school context
            try:
                student_query = Student.objects.filter(
                    admission_number=student_id
                )

                # Apply school filter for multi-tenancy
                if user_school:
                    student_query = student_query.filter(school=user_school)

                student = student_query.get()
            except Student.DoesNotExist:
                errors.append(f"Student with ID {student_id} not found.")
                continue

            # Check if the student is in the class
            student_class_query = StudentClass.objects.filter(
                student=student,
                assigned_class=assignment.class_assigned,
                is_active=True,
            )

            # Apply school filter for multi-tenancy
            if user_school:
                student_class_query = student_class_query.filter(school=user_school)

            if not student_class_query.exists():
                errors.append(
                    f"Student {student.full_name} ({student_id}) is not in this class."
                )
                continue

            # Get the raw score from column C
            try:
                raw_score_cell = ws.cell(row=row, column=3)
                raw_score_value = raw_score_cell.value

                # Handle different data types
                if isinstance(raw_score_value, str):
                    raw_score_value = raw_score_value.strip()
                    raw_score_value = (
                        float(raw_score_value) if raw_score_value else None
                    )

                # Convert to decimal or None
                raw_score = (
                    Decimal(str(raw_score_value))
                    if raw_score_value not in ("", None)
                    else None
                )

                # Validate score range (0-100 for mock exams)
                if raw_score is not None and (raw_score < 0 or raw_score > 100):
                    errors.append(
                        f"Invalid raw score for {student.full_name}: {raw_score}. Must be between 0 and 100."
                    )
                    continue

                # Skip if no score provided
                if raw_score is None:
                    continue

            except (ValueError, InvalidOperation) as e:
                errors.append(f"Invalid score format for {student.full_name}: {str(e)}")
                continue

            # Update or create the assessment
            assessment_query = Assessment.objects.filter(
                class_subject=class_subject,
                student=student,
                assessment_type="mock_exam",
                mock_exam=mock_exam,
            )

            assessment = assessment_query.first()

            if assessment:
                # Update existing assessment
                assessment.raw_exam_score = raw_score
                # For mock exams, total_score is the same as raw_exam_score
                assessment.total_score = raw_score
                assessment.recorded_by = request.user
                assessment.save()
            else:
                # Create new assessment
                # For mock exams, total_score is the same as raw_exam_score
                assessment = Assessment.objects.create(
                    class_subject=class_subject,
                    student=student,
                    assessment_type="mock_exam",
                    mock_exam=mock_exam,
                    raw_exam_score=raw_score,
                    total_score=raw_score,  # For mock exams, total = raw score
                    recorded_by=request.user,
                )

            # The Assessment model's save method will automatically calculate total_score, grade, and remarks
            updated_assessments.append(assessment)
            updated_count += 1

        # Calculate positions for all students in this class_subject for this mock exam
        if hasattr(Assessment, "calculate_mock_exam_positions"):
            Assessment.calculate_mock_exam_positions(class_subject, mock_exam)
        else:
            # Fallback: manual position calculation
            assessments = Assessment.objects.filter(
                class_subject=class_subject,
                assessment_type="mock_exam",
                mock_exam=mock_exam,
            ).order_by("-total_score")

            position = 1
            last_score = None
            for assessment in assessments:
                if assessment.total_score is None:
                    assessment.position = None
                    assessment.save(update_fields=["position"])
                    continue

                if last_score is not None and assessment.total_score != last_score:
                    position = position + 1

                assessment.position = position
                assessment.save(update_fields=["position"])
                last_score = assessment.total_score

        # Return success response
        if errors:
            return JsonResponse(
                {
                    "success": False,
                    "message": f"Import completed with {len(errors)} errors.",
                    "error_messages": errors[:10],  # First 10 errors
                    "total_errors": len(errors),
                    "updated_count": updated_count,
                }
            )
        else:
            return JsonResponse(
                {
                    "success": True,
                    "message": f"Successfully imported mock exam scores for {updated_count} students.",
                    "updated_count": updated_count,
                }
            )

    except Exception as e:
        logger.error(f"Error importing mock exam scores: {str(e)}", exc_info=True)
        return JsonResponse(
            {
                "success": False,
                "error": f"Error importing mock exam scores: {str(e)}",
                "error_messages": [str(e)],
            }
        )


@login_required
@csrf_exempt
def import_mock_exam_scores_batch(request):
    """
    Import mock exam scores from Excel file with multiple sheets for multiple classes.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Only POST method allowed"}, status=405)

    # Check if the user is authorized
    if request.user.role not in ["teacher", "admin"]:
        return JsonResponse(
            {"error": "Only teachers and administrators can import mock exam scores."}, status=403
        )

    # Check if a file was uploaded
    if "batchFile" not in request.FILES:
        return JsonResponse({"error": "No file was uploaded."}, status=400)

    file = request.FILES["batchFile"]
    # Check if the file is an Excel file
    if not file.name.endswith((".xlsx", ".xls")):
        return JsonResponse(
            {"error": "Please upload an Excel file (.xlsx or .xls)."}, status=400
        )

    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    if not user_school:
        return JsonResponse({"error": "No school associated with your account."}, status=400)

    # Get current academic year
    current_academic_year = AcademicYear.objects.filter(
        school=user_school, is_current=True
    ).first()

    if not current_academic_year:
        return JsonResponse({"error": "No active academic year found."}, status=400)

    # Get default active mock exam
    default_mock_exam = MockExam.objects.filter(
        school=user_school,
        academic_year=current_academic_year,
        is_active=True
    ).order_by('-exam_date', '-created_at').first()

    if not default_mock_exam:
        return JsonResponse(
            {"error": "No active mock exam found. Please create an active mock exam first."}, status=400
        )

    # Get the teacher instance (for teachers)
    teacher = getattr(request.user, "teacher_profile", None)
    if request.user.role == "teacher" and not teacher:
        return JsonResponse(
            {"error": "Teacher profile is not linked to your account."}, status=403
        )

    try:
        # Create a BytesIO object from the uploaded file
        file_content = BytesIO(file.read())

        # Load the workbook with data_only=True to get values instead of formulas
        wb = openpyxl.load_workbook(file_content, data_only=True)

        results = []
        total_updated = 0
        total_errors = []

        # Process each sheet (each sheet represents a class-subject combination)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Verify the file is for the correct school (for multi-tenancy)
            if user_school and "H3" in ws and ws["H3"].value == "school_id":
                file_school_id = ws["I3"].value
                if file_school_id and int(file_school_id) != user_school.id:
                    total_errors.append(f"Sheet '{sheet_name}' belongs to a different school")
                    continue

            # Check mock exam ID from file
            file_mock_exam_id = None
            if "H4" in ws and ws["H4"].value == "mock_exam_id":
                file_mock_exam_id = ws["I4"].value

            # Use mock exam from file if available, otherwise use default
            mock_exam = default_mock_exam
            if file_mock_exam_id:
                try:
                    mock_exam = MockExam.objects.get(
                        id=file_mock_exam_id, school=user_school, is_active=True
                    )
                except MockExam.DoesNotExist:
                    # Use default if file mock exam not found
                    pass

            # Try to get class_subject_id from metadata first (preferred method)
            class_subject = None
            file_class_subject_id = None
            if "H1" in ws and ws["H1"].value == "class_subject_id":
                file_class_subject_id = ws["I1"].value

            if file_class_subject_id:
                try:
                    class_subject_query = ClassSubject.objects.filter(
                        class_subject_id=file_class_subject_id
                    )
                    if user_school:
                        class_subject_query = class_subject_query.filter(
                            subject__school=user_school,
                            class_name__school=user_school,
                            academic_year__school=user_school,
                        )
                    class_subject = class_subject_query.first()
                except Exception as e:
                    logger.error(f"Error getting class_subject from ID: {str(e)}")

            # If class_subject not found from metadata, try to get from assignment_id
            if not class_subject:
                file_assignment_id = None
                if "H2" in ws and ws["H2"].value == "assignment_id":
                    file_assignment_id = ws["I2"].value

                if file_assignment_id:
                    try:
                        assignment_query = TeacherSubjectAssignment.objects.filter(
                            id=file_assignment_id, is_active=True
                        )
                        if request.user.role == "teacher" and teacher:
                            assignment_query = assignment_query.filter(teacher=teacher)
                        if user_school:
                            assignment_query = assignment_query.filter(school=user_school)
                        
                        assignment = assignment_query.first()
                        if assignment:
                            class_subject = ClassSubject.objects.filter(
                                subject=assignment.subject,
                                class_name=assignment.class_assigned,
                                academic_year=assignment.academic_year,
                                is_active=True
                            ).first()
                    except Exception as e:
                        logger.error(f"Error getting class_subject from assignment_id: {str(e)}")

            # Fallback: Get class and subject from sheet header
            if not class_subject:
                class_name = None
                subject_name = None

                # Try to get from header rows
                if ws["A2"].value:
                    class_name = str(ws["A2"].value).replace("Class: ", "").strip()
                if ws["A3"].value:
                    subject_name = str(ws["A3"].value).replace("Subject: ", "").strip()

                if not class_name or not subject_name:
                    total_errors.append(f"Sheet '{sheet_name}' has invalid format (missing class/subject info)")
                    continue

                # Find the class and subject
                try:
                    # Filter by academic year as well to avoid ambiguity
                    class_obj = Class.objects.filter(
                        name=class_name, 
                        school=user_school,
                        academic_year=current_academic_year
                    ).first()
                    
                    if not class_obj:
                        # Try without academic year filter as fallback
                        class_obj = Class.objects.filter(
                            name=class_name, 
                            school=user_school
                        ).first()
                    
                    subject_obj = Subject.objects.filter(
                        subject_name=subject_name, 
                        school=user_school
                    ).first()

                    if not class_obj or not subject_obj:
                        total_errors.append(
                            f"Sheet '{sheet_name}': Class '{class_name}' or Subject '{subject_name}' not found"
                        )
                        continue

                    # Verify the teacher is assigned to this class (for teachers)
                    if request.user.role == "teacher" and teacher:
                        assignment = TeacherSubjectAssignment.objects.filter(
                            teacher=teacher,
                            class_assigned=class_obj,
                            subject=subject_obj,
                            academic_year=current_academic_year,
                            is_active=True,
                            school=user_school,
                        ).first()

                        if not assignment:
                            total_errors.append(
                                f"Sheet '{sheet_name}': You are not assigned to {class_name} - {subject_name}"
                            )
                            continue

                    # Get the ClassSubject instance
                    class_subject = ClassSubject.objects.filter(
                        subject=subject_obj,
                        class_name=class_obj,
                        academic_year=current_academic_year,
                        is_active=True,
                    ).first()

                    if not class_subject:
                        total_errors.append(
                            f"Sheet '{sheet_name}': Subject '{subject_name}' is not assigned to class '{class_name}'"
                        )
                        continue
                except Exception as e:
                    total_errors.append(f"Sheet '{sheet_name}': Error processing class/subject - {str(e)}")
                    logger.error(f"Error processing sheet '{sheet_name}': {str(e)}", exc_info=True)
                    continue

            if not class_subject:
                total_errors.append(f"Sheet '{sheet_name}': Could not determine class-subject combination")
                logger.error(f"Sheet '{sheet_name}': class_subject is None after all attempts")
                continue

            # Start processing data from row 9 (after headers)
            updated_count = 0
            sheet_errors = []
            
            try:
                for row in range(9, ws.max_row + 1):
                    student_id = ws.cell(row=row, column=1).value
                    raw_score_value = ws.cell(row=row, column=3).value

                    # Skip if student ID or score is empty
                    if not student_id or raw_score_value is None:
                        continue

                    # Handle different data types for scores
                    if isinstance(raw_score_value, str):
                        raw_score_value = raw_score_value.strip()
                        raw_score_value = (
                            float(raw_score_value) if raw_score_value else None
                        )

                    # Try to convert score to float
                    try:
                        raw_score = (
                            float(raw_score_value)
                            if raw_score_value is not None
                            else None
                        )
                    except (ValueError, TypeError):
                        continue

                    # Validate score range (0-100 for mock exams)
                    if raw_score is not None and (raw_score < 0 or raw_score > 100):
                        continue

                    # Find the student
                    try:
                        student_query = Student.objects.filter(
                            admission_number=student_id
                        )

                        # Apply school filter for multi-tenancy
                        if user_school:
                            student_query = student_query.filter(school=user_school)

                        student = student_query.get()
                    except Student.DoesNotExist:
                        continue

                    # Check if student is in this class - use class from class_subject
                    student_class_query = StudentClass.objects.filter(
                        student=student,
                        assigned_class=class_subject.class_name,
                        is_active=True,
                    )

                    # Apply school filter for multi-tenancy
                    if user_school:
                        student_class_query = student_class_query.filter(school=user_school)

                    if not student_class_query.exists():
                        continue

                    # Update or create the assessment
                    assessment_query = Assessment.objects.filter(
                        class_subject=class_subject,
                        student=student,
                        assessment_type="mock_exam",
                        mock_exam=mock_exam,
                    )

                    assessment = assessment_query.first()

                    if assessment:
                        # Update existing assessment
                        assessment.raw_exam_score = Decimal(str(raw_score))
                        # For mock exams, total_score is the same as raw_exam_score
                        assessment.total_score = Decimal(str(raw_score))
                        assessment.recorded_by = request.user
                        assessment.save()
                    else:
                        # Create new assessment
                        # For mock exams, total_score is the same as raw_exam_score
                        assessment = Assessment.objects.create(
                            class_subject=class_subject,
                            student=student,
                            assessment_type="mock_exam",
                            mock_exam=mock_exam,
                            raw_exam_score=Decimal(str(raw_score)),
                            total_score=Decimal(str(raw_score)),  # For mock exams, total = raw score
                            recorded_by=request.user,
                        )

                    updated_count += 1

            except Exception as e:
                sheet_errors.append(f"Error processing students in sheet '{sheet_name}': {str(e)}")
                logger.error(f"Error processing students in sheet '{sheet_name}': {str(e)}", exc_info=True)
                total_errors.extend(sheet_errors)
                continue

            # Calculate positions for all students in this class_subject for this mock exam
            try:
                if hasattr(Assessment, "calculate_mock_exam_positions"):
                    Assessment.calculate_mock_exam_positions(class_subject, mock_exam)
                else:
                    # Fallback: manual position calculation
                    assessments = Assessment.objects.filter(
                        class_subject=class_subject,
                        assessment_type="mock_exam",
                        mock_exam=mock_exam,
                    ).order_by("-total_score")

                    position = 1
                    last_score = None
                    for assessment in assessments:
                        if assessment.total_score is None:
                            assessment.position = None
                            assessment.save(update_fields=["position"])
                            continue

                        if last_score is not None and assessment.total_score != last_score:
                            position = position + 1

                        assessment.position = position
                        assessment.save(update_fields=["position"])
                        last_score = assessment.total_score
            except Exception as e:
                logger.error(f"Error calculating positions for sheet '{sheet_name}': {str(e)}", exc_info=True)

            # Add the result for this sheet
            results.append(
                {
                    "class_name": class_name,
                    "subject_name": subject_name,
                    "processed": updated_count,
                }
            )
            total_updated += updated_count
            total_errors.extend(sheet_errors)

        if not results:
            return JsonResponse(
                {
                    "success": False,
                    "error": "No scores were successfully imported",
                    "errors": total_errors,
                }
            )

        return JsonResponse(
            {
                "success": True,
                "results": results,
                "total_updated": total_updated,
                "errors": total_errors[:10],  # First 10 errors
            }
        )

    except Exception as e:
        import traceback

        logger.error(f"Error importing mock exam batch scores: {str(e)}", exc_info=True)
        return JsonResponse(
            {"success": False, "error": str(e), "traceback": traceback.format_exc()},
            status=500,
        )

