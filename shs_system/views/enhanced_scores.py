"""
Enhanced Score Entry View
Handles individual score components (individual_score, class_test_score, project_score, group_work_score)
with real-time calculation and modern UI functionality.
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie, csrf_protect
from django.middleware.csrf import get_token
from django.views.decorators.http import require_POST
from django.db import transaction
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation
import logging
import json
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from ..models import (
    User,
    Teacher,
    Student,
    StudentClass,
    Class,
    Subject,
    TeacherSubjectAssignment,
    ClassSubject,
    Assessment,
    AcademicYear,
    Term,
    ScoringConfiguration,
    GradingSystem,
)

logger = logging.getLogger(__name__)


def find_assignment_for_import(class_name, subject_name, user, teacher, user_school):
    """
    Find assignment with improved matching logic to handle naming variations.
    """
    # Start with base query
    if user.role == "teacher" and teacher:
        base_query = TeacherSubjectAssignment.objects.filter(
            teacher=teacher,
            is_active=True,
        )
    elif user.role == "admin":
        base_query = TeacherSubjectAssignment.objects.filter(
            is_active=True,
        )
    else:
        return None

    # Apply school filter for multi-tenancy
    if user_school:
        base_query = base_query.filter(school=user_school)

    # Try exact match first
    assignment = base_query.filter(
        class_assigned__name=class_name,
        subject__subject_name=subject_name,
    ).first()

    if assignment:
        return assignment

    # Try case-insensitive match
    assignment = base_query.filter(
        class_assigned__name__iexact=class_name,
        subject__subject_name__iexact=subject_name,
    ).first()

    if assignment:
        return assignment

    # Try with trimmed whitespace variations
    assignment = base_query.filter(
        class_assigned__name__iexact=class_name.strip(),
        subject__subject_name__iexact=subject_name.strip(),
    ).first()

    if assignment:
        return assignment

    # Try with common abbreviations and variations
    # Handle common subject name variations
    subject_variations = [
        subject_name,
        subject_name.replace(" AND ", " & "),
        subject_name.replace(" & ", " AND "),
        subject_name.replace("TECHNO", "TECHNOLOGY"),
        subject_name.replace("TECHNOLOGY", "TECHNO"),
        subject_name.replace("EDU", "EDUCATION"),
        subject_name.replace("EDUCATION", "EDU"),
        subject_name.replace("MORAL", "MORALS"),
        subject_name.replace("MORALS", "MORAL"),
    ]

    for subject_var in subject_variations:
        assignment = base_query.filter(
            class_assigned__name__iexact=class_name.strip(),
            subject__subject_name__iexact=subject_var.strip(),
        ).first()

        if assignment:
            return assignment

    # Try with class name variations
    class_variations = [
        class_name,
        class_name.replace("JHS", "JUNIOR HIGH SCHOOL"),
        class_name.replace("JUNIOR HIGH SCHOOL", "JHS"),
        class_name.replace("SHS", "SENIOR HIGH SCHOOL"),
        class_name.replace("SENIOR HIGH SCHOOL", "SHS"),
    ]

    for class_var in class_variations:
        for subject_var in subject_variations:
            assignment = base_query.filter(
                class_assigned__name__iexact=class_var.strip(),
                subject__subject_name__iexact=subject_var.strip(),
            ).first()

            if assignment:
                return assignment

    return None


def get_available_assignments_for_debugging(user, teacher, user_school):
    """
    Get list of available assignments for debugging purposes.
    """
    try:
        if user.role == "teacher" and teacher:
            base_query = TeacherSubjectAssignment.objects.filter(
                teacher=teacher,
                is_active=True,
            )
        elif user.role == "admin":
            base_query = TeacherSubjectAssignment.objects.filter(
                is_active=True,
            )
        else:
            return "No assignments available for this user role"

        # Apply school filter for multi-tenancy
        if user_school:
            base_query = base_query.filter(school=user_school)

        assignments = base_query.select_related("class_assigned", "subject")[
            :10
        ]  # Limit to first 10

        assignment_list = []
        for assignment in assignments:
            assignment_list.append(
                f"'{assignment.class_assigned.name}' - '{assignment.subject.subject_name}'"
            )

        return "; ".join(assignment_list) if assignment_list else "No assignments found"

    except Exception as e:
        return f"Error retrieving assignments: {str(e)}"


def find_similar_assignments(class_name, subject_name, user, teacher, user_school):
    """
    Find assignments that are similar to the requested class and subject names.
    """
    try:
        if user.role == "teacher" and teacher:
            base_query = TeacherSubjectAssignment.objects.filter(
                teacher=teacher,
                is_active=True,
            )
        elif user.role == "admin":
            base_query = TeacherSubjectAssignment.objects.filter(
                is_active=True,
            )
        else:
            return "No assignments available for this user role"

        # Apply school filter for multi-tenancy
        if user_school:
            base_query = base_query.filter(school=user_school)

        # Find assignments with similar class names
        similar_class_assignments = base_query.filter(
            class_assigned__name__icontains=class_name.split()[
                0
            ]  # First word of class name
        ).select_related("class_assigned", "subject")[:5]

        # Find assignments with similar subject names
        similar_subject_assignments = base_query.filter(
            subject__subject_name__icontains=subject_name.split()[
                0
            ]  # First word of subject name
        ).select_related("class_assigned", "subject")[:5]

        similar_list = []
        for assignment in similar_class_assignments:
            similar_list.append(
                f"'{assignment.class_assigned.name}' - '{assignment.subject.subject_name}'"
            )

        for assignment in similar_subject_assignments:
            assignment_str = f"'{assignment.class_assigned.name}' - '{assignment.subject.subject_name}'"
            if assignment_str not in similar_list:
                similar_list.append(assignment_str)

        return (
            "; ".join(similar_list[:3])
            if similar_list
            else "No similar assignments found"
        )

    except Exception as e:
        return f"Error finding similar assignments: {str(e)}"


def sanitize_excel_sheet_name(name):
    """
    Sanitize a name to be used as an Excel sheet name.
    Excel sheet names have the following restrictions:
    - Maximum 31 characters
    - Cannot contain: [ ] : * ? / \ 
    - Cannot start or end with apostrophe
    - Cannot be empty
    """
    if not name:
        return "Sheet"

    # Remove or replace invalid characters
    invalid_chars = ["[", "]", ":", "*", "?", "/", "\\"]
    for char in invalid_chars:
        name = name.replace(char, "_")

    # Remove leading/trailing apostrophes and whitespace
    name = name.strip("'").strip()

    # Truncate to 31 characters (Excel limit)
    if len(name) > 31:
        name = name[:31]

    # Ensure it's not empty after sanitization
    if not name:
        name = "Sheet"

    return name


def calculate_enhanced_scores(
    individual_score,
    class_test_score,
    project_score,
    group_work_score,
    exam_score,
    scoring_config,
    user_school,
):
    """
    Calculate all enhanced score values including class score, total score, grade, and remarks.
    Returns a dictionary with all calculated values.
    """
    from decimal import Decimal

    result = {
        "class_score": None,
        "scaled_exam_score": None,
        "total_score": None,
        "grade": None,
        "remarks": None,
        "position": None,
    }

    # Calculate class score from individual components
    if any([individual_score, class_test_score, project_score, group_work_score]):
        if scoring_config:
            # Calculate weighted class score using scoring configuration
            individual_weighted = 0
            class_test_weighted = 0
            project_weighted = 0
            group_work_weighted = 0

            if individual_score is not None:
                individual_weighted = (
                    Decimal(str(individual_score))
                    / Decimal(str(scoring_config.max_individual_score))
                ) * Decimal(str(scoring_config.individual_score_percentage))

            if class_test_score is not None:
                class_test_weighted = (
                    Decimal(str(class_test_score))
                    / Decimal(str(scoring_config.max_class_test_score))
                ) * Decimal(str(scoring_config.class_test_score_percentage))

            if project_score is not None:
                project_weighted = (
                    Decimal(str(project_score))
                    / Decimal(str(scoring_config.max_project_score))
                ) * Decimal(str(scoring_config.project_score_percentage))

            if group_work_score is not None:
                group_work_weighted = (
                    Decimal(str(group_work_score))
                    / Decimal(str(scoring_config.max_group_work_score))
                ) * Decimal(str(scoring_config.group_work_score_percentage))

            result["class_score"] = float(
                individual_weighted
                + class_test_weighted
                + project_weighted
                + group_work_weighted
            )
        else:
            # Fallback: simple sum of all components
            total = 0
            if individual_score is not None:
                total += float(individual_score)
            if class_test_score is not None:
                total += float(class_test_score)
            if project_score is not None:
                total += float(project_score)
            if group_work_score is not None:
                total += float(group_work_score)
            result["class_score"] = total

    # Calculate scaled exam score
    if exam_score is not None:
        if scoring_config:
            result["scaled_exam_score"] = float(
                scoring_config.calculate_exam_score(exam_score)
            )
        else:
            result["scaled_exam_score"] = float(exam_score)

    # Calculate total score
    if result["class_score"] is not None and result["scaled_exam_score"] is not None:
        result["total_score"] = result["class_score"] + result["scaled_exam_score"]

    # Calculate grade and remarks
    if result["total_score"] is not None:
        grade_info = GradingSystem.get_grade_for_score(
            result["total_score"], user_school
        )
        if grade_info:
            result["grade"] = grade_info.grade_letter
            result["remarks"] = grade_info.remarks
        else:
            result["grade"] = "N/A"
            result["remarks"] = "Not Graded"

    return result


def get_user_school(user):
    """Get the school associated with the user for multi-tenancy support."""
    if hasattr(user, "school") and user.school:
        return user.school

    # For superadmins, return None to access all schools
    if user.is_superuser or getattr(user, "is_superadmin", False):
        return None

    return user.school


@login_required
@ensure_csrf_cookie
def enhanced_enter_scores(request):
    """
    Enhanced score entry view that handles individual score components
    and provides real-time calculation functionality.
    """
    # Ensure only teachers and admins can access this view
    if request.user.role not in ["teacher", "admin"]:
        messages.error(
            request,
            "Only teachers and administrators can access the score entry system.",
        )
        return render(request, "errors/403.html", status=403)

    teacher = getattr(request.user, "teacher_profile", None)

    # For admin users, they don't need a teacher profile
    if request.user.role == "teacher" and not teacher:
        messages.error(request, "Teacher profile is not linked to this user.")
        return redirect("login")

    # For admin users, we'll handle teacher assignments differently
    if request.user.role == "admin":
        teacher = None  # Admin users don't have teacher profiles

    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    # Get the current academic year and term for the user's school
    if user_school:
        current_academic_year = AcademicYear.objects.filter(
            is_current=True, school=user_school
        ).first()
        current_term = Term.objects.filter(is_current=True, school=user_school).first()
    else:
        # For superadmins, get any current academic year/term
        current_academic_year = AcademicYear.objects.filter(is_current=True).first()
        current_term = Term.objects.filter(is_current=True).first()

    if not current_academic_year or not current_term:
        messages.error(request, "No active academic year or term is set.")
        return redirect("dashboard")

    # Fetch only assignments for the current academic year and user's school
    assignments = []
    if current_academic_year and current_term:
        if request.user.role == "teacher" and teacher:
            # For teachers, get their specific assignments
            assignments_query = teacher.teachersubjectassignment_set.filter(
                is_active=True, academic_year=current_academic_year
            )

            # Apply school filter for multi-tenancy if not superadmin
            if user_school:
                assignments_query = assignments_query.filter(school=user_school)


            # Filter out assignments where ClassSubject is not active
            # Get active class-subject combinations
            active_class_subjects = ClassSubject.objects.filter(
                academic_year=current_academic_year, is_active=True
            ).values_list('class_name_id', 'subject_id')
            
            # Filter assignments to only include those with active ClassSubject
            filtered_assignments = []
            for assignment in assignments_query.select_related("class_assigned", "subject"):
                if (assignment.class_assigned.id, assignment.subject.id) in active_class_subjects:
                    filtered_assignments.append(assignment)
            
            assignments = filtered_assignments
        elif request.user.role == "admin":
            # For admin users, get all assignments in their school
            # Optionally filter by teacher if teacher_id is provided

            assignments_query = TeacherSubjectAssignment.objects.filter(
                is_active=True, academic_year=current_academic_year
            )


            # Apply teacher filter if provided (for admins to view specific teacher's classes)
            teacher_id = request.GET.get("teacher_id")
            if teacher_id:
                try:
                    teacher_filter = Teacher.objects.get(id=teacher_id, school=user_school)
                    assignments_query = assignments_query.filter(teacher=teacher_filter)
                except Teacher.DoesNotExist:
                    messages.warning(request, "Selected teacher not found.")
                    teacher_id = None


            # Apply school filter for multi-tenancy
            if user_school:
                assignments_query = assignments_query.filter(school=user_school)


            # Filter out assignments where ClassSubject is not active
            # Get active class-subject combinations
            active_class_subjects = ClassSubject.objects.filter(
                academic_year=current_academic_year, is_active=True
            ).values_list('class_name_id', 'subject_id')
            
            # Filter assignments to only include those with active ClassSubject
            filtered_assignments = []
            for assignment in assignments_query.select_related("class_assigned", "subject", "teacher"):
                if (assignment.class_assigned.id, assignment.subject.id) in active_class_subjects:
                    filtered_assignments.append(assignment)
            
            assignments = filtered_assignments
            
            # Get list of teachers for filter dropdown (admin only)
            if user_school:
                teachers_list = Teacher.objects.filter(school=user_school).order_by("full_name")
            else:
                teachers_list = Teacher.objects.all().order_by("full_name")
        else:
            # For teachers, no teachers list needed
            teachers_list = None


    # Get scoring configuration for dynamic display
    scoring_config = None
    if user_school:
        scoring_config = ScoringConfiguration.get_active_config(user_school)

    # If no configuration exists, create default values
    if not scoring_config:
        scoring_config = type(
            "ScoringConfig",
            (),
            {
                "exam_score_percentage": Decimal("70.0"),
                "class_score_percentage": Decimal("30.0"),
                "max_class_score": Decimal("30.0"),
                "individual_max_mark": Decimal("15.0"),
                "class_test_max_mark": Decimal("15.0"),
                "project_max_mark": Decimal("15.0"),
                "group_work_max_mark": Decimal("15.0"),
            },
        )()

    # Initialize context with common data
    context = {
        "assignments": assignments,
        "current_academic_year": current_academic_year,
        "current_term": current_term,
        "user_school": user_school,
        "scoring_config": scoring_config,
    }

    
    # Add teachers list for admin filter (if admin)
    if request.user.role == "admin":
        if 'teachers_list' in locals():
            context["teachers_list"] = teachers_list
        else:
            # Fallback if teachers_list wasn't created
            if user_school:
                context["teachers_list"] = Teacher.objects.filter(school=user_school).order_by("full_name")
            else:
                context["teachers_list"] = Teacher.objects.all().order_by("full_name")


    # Handle form submission for enhanced score entry
    if request.method == "POST":
        return handle_enhanced_score_submission(
            request, teacher, user_school, context, request.user.role
        )

    # Handle assignment selection and student display
    assignment_id = request.GET.get("assignment_id")
    if assignment_id:
        try:
            # Include school filter for multi-tenancy
            if request.user.role == "teacher" and teacher:
                assignment_query = TeacherSubjectAssignment.objects.filter(
                    id=assignment_id, teacher=teacher, is_active=True
                )
            elif request.user.role == "admin":
                # Admin can access any assignment in their school
                assignment_query = TeacherSubjectAssignment.objects.filter(
                    id=assignment_id, is_active=True
                )
            else:
                assignment = None

            if assignment_query and user_school:
                assignment_query = assignment_query.filter(school=user_school)

            assignment = (
                get_object_or_404(assignment_query) if assignment_query else None
            )

            if not current_term:
                messages.error(request, "No current term is set.")
                return redirect("enhanced_enter_scores")

            # Get or create the ClassSubject instance with school context
            try:
                class_subject_query = ClassSubject.objects.filter(
                    subject=assignment.subject,
                    class_name=assignment.class_assigned,
                    academic_year=assignment.academic_year,

                    is_active=True

                )

                # Apply school filter through related objects
                if user_school:
                    class_subject_query = class_subject_query.filter(
                        subject__school=user_school,
                        class_name__school=user_school,
                        academic_year__school=user_school,
                    )

                class_subject = class_subject_query.first()

                if not class_subject:

                    # Check if there's an inactive ClassSubject that we can reactivate
                    inactive_class_subject_query = ClassSubject.objects.filter(
                        subject=assignment.subject,
                        class_name=assignment.class_assigned,
                        academic_year=assignment.academic_year,
                        is_active=False
                    )

                    if user_school:
                        inactive_class_subject_query = inactive_class_subject_query.filter(
                            subject__school=user_school,
                            class_name__school=user_school,
                            academic_year__school=user_school,
                        )

                    inactive_class_subject = inactive_class_subject_query.first()

                    if inactive_class_subject:
                        # Reactivate the existing assignment
                        inactive_class_subject.is_active = True
                        inactive_class_subject.save()
                        class_subject = inactive_class_subject
                    else:
                        # Don't create automatically - this should be done through admin interface
                        messages.error(request, f"Subject '{assignment.subject.subject_name}' is not assigned to class '{assignment.class_assigned.name}'. Please assign this subject to the class first.")
                        return redirect("enhanced_enter_scores")

                logger.debug(
                    f"ClassSubject {'reactivated' if class_subject and not class_subject_query.exists() else 'found'}: {class_subject}"

                )
            except Exception as e:
                logger.error(f"Error creating/finding ClassSubject: {str(e)}")
                messages.error(request, f"Error setting up class subject: {str(e)}")
                return redirect("enhanced_enter_scores")

            # Get students in the selected class with school context
            student_classes_query = StudentClass.objects.filter(
                assigned_class=assignment.class_assigned, is_active=True
            ).select_related("student")

            if user_school:
                student_classes_query = student_classes_query.filter(
                    student__school=user_school, assigned_class__school=user_school
                )

            student_classes = student_classes_query.order_by("student__full_name")

            # Get existing assessments for all students in this class and subject

            # Exclude mock exam assessments - only get regular term assessments

            assessments = Assessment.objects.filter(
                class_subject=class_subject,
                student__in=[sc.student for sc in student_classes],
                term=current_term,

            ).exclude(assessment_type='mock_exam').select_related("student")


            # Create a dictionary for quick lookup
            student_assessments = {
                assessment.student.id: assessment for assessment in assessments
            }

            # Attach assessment data to student objects for template display
            for student_class in student_classes:
                student = student_class.student
                assessment = student_assessments.get(student.id)
                if assessment:
                    # Individual score components
                    student.individual_score = assessment.individual_score
                    student.class_test_score = assessment.class_test_score
                    student.project_score = assessment.project_score
                    student.group_work_score = assessment.group_work_score

                    # Calculated scores
                    student.class_score = assessment.class_score
                    student.exam_score = (
                        assessment.raw_exam_score
                    )  # Use raw for input field
                    student.total_score = assessment.total_score
                    student.grade = assessment.grade
                    student.remarks = assessment.remarks
                    student.position = assessment.position

                    # The exam_score field contains the scaled value for display
                    student.scaled_exam_score = assessment.exam_score
                else:
                    # Initialize empty values
                    student.individual_score = None
                    student.class_test_score = None
                    student.project_score = None
                    student.group_work_score = None
                    student.class_score = None
                    student.exam_score = None
                    student.scaled_exam_score = None
                    student.total_score = None
                    student.grade = None
                    student.remarks = None
                    student.position = None

            # Calculate class statistics
            class_stats = calculate_class_statistics(student_classes)

            context.update(
                {
                    "students_in_class": student_classes,
                    "selected_assignment": assignment,
                    "class_subject": class_subject,
                    "class_stats": class_stats,
                }
            )

        except Exception as e:
            logger.error(f"Error in enhanced score entry: {str(e)}")
            messages.error(request, f"An error occurred: {str(e)}")

    return render(request, "student/enhanced_enter_scores.html", context)


@transaction.atomic
def handle_enhanced_score_submission(request, teacher, user_school, context, user_role):
    """
    Handle the submission of enhanced score entry form with individual components.
    """
    try:
        assignment_id = request.POST.get("assignment_id")
        if not assignment_id:
            raise ValueError("No assignment ID provided")

        # Include school filter for multi-tenancy
        if user_role == "teacher" and teacher:
            assignment_query = TeacherSubjectAssignment.objects.filter(
                id=assignment_id, teacher=teacher, is_active=True
            )
        elif user_role == "admin":
            # Admin can access any assignment in their school
            assignment_query = TeacherSubjectAssignment.objects.filter(
                id=assignment_id, is_active=True
            )
        else:
            raise ValueError("Unauthorized access to assignment")

        if user_school:
            assignment_query = assignment_query.filter(school=user_school)

        assignment = get_object_or_404(assignment_query)

        # Get scoring configuration
        scoring_config = ScoringConfiguration.get_active_config(user_school)
        if not scoring_config:
            messages.error(
                request, "No scoring configuration found. Please contact administrator."
            )
            return redirect("enhanced_enter_scores")

        # Get or create ClassSubject
        class_subject = get_or_create_class_subject(assignment, user_school)

        # Get students in class
        student_classes_query = StudentClass.objects.filter(
            assigned_class=assignment.class_assigned, is_active=True
        ).select_related("student")

        if user_school:
            student_classes_query = student_classes_query.filter(
                student__school=user_school
            )

        student_classes = student_classes_query

        success_messages = []
        error_messages = []
        updated_students = []

        # Process each student's data
        for student_class in student_classes:
            student = student_class.student

            # Get individual score components
            individual_score = request.POST.get(f"individual_score_{student.id}")
            class_test_score = request.POST.get(f"class_test_score_{student.id}")
            project_score = request.POST.get(f"project_score_{student.id}")
            group_work_score = request.POST.get(f"group_work_score_{student.id}")
            exam_score = request.POST.get(f"exam_score_{student.id}")

            # Skip if no scores provided
            if not any(
                [
                    individual_score,
                    class_test_score,
                    project_score,
                    group_work_score,
                    exam_score,
                ]
            ):
                continue

            try:
                # Convert and validate scores
                individual_score = (
                    Decimal(individual_score) if individual_score else None
                )
                class_test_score = (
                    Decimal(class_test_score) if class_test_score else None
                )
                project_score = Decimal(project_score) if project_score else None
                group_work_score = (
                    Decimal(group_work_score) if group_work_score else None
                )
                exam_score = Decimal(exam_score) if exam_score else None

                # Validate score ranges
                validation_errors = validate_score_ranges(
                    individual_score,
                    class_test_score,
                    project_score,
                    group_work_score,
                    exam_score,
                    scoring_config,
                    student,
                )

                if validation_errors:
                    error_messages.extend(validation_errors)
                    continue

                # Update or create assessment
                assessment = update_or_create_assessment(
                    class_subject,
                    student,
                    individual_score,
                    class_test_score,
                    project_score,
                    group_work_score,
                    exam_score,
                    request.user,
                    context["current_term"],
                )

                updated_students.append(student.full_name)
                success_messages.append(f"Scores updated for {student.full_name}")

            except (ValueError, InvalidOperation) as e:
                error_message = (
                    f"Invalid score format for {student.full_name}: {str(e)}"
                )
                error_messages.append(error_message)
                logger.warning(error_message)
                continue

            except Exception as e:
                error_message = (
                    f"Error updating scores for {student.full_name}: {str(e)}"
                )
                error_messages.append(error_message)
                logger.error(error_message)
                continue

        # Provide feedback to user
        if success_messages:
            if len(updated_students) == 1:
                messages.success(
                    request, f"Scores successfully updated for {updated_students[0]}."
                )
            else:
                messages.success(
                    request,
                    f"Scores successfully updated for {len(updated_students)} students: "
                    f"{', '.join(updated_students[:3])}"
                    f"{'...' if len(updated_students) > 3 else ''}.",
                )

        if error_messages:
            for error in error_messages[:5]:  # Show max 5 errors
                messages.error(request, error)
            if len(error_messages) > 5:
                messages.error(
                    request, f"...and {len(error_messages) - 5} more errors."
                )

        # Redirect to prevent resubmission
        return redirect(f"{request.path}?assignment_id={assignment_id}")

    except Exception as e:
        logger.error(f"Unexpected error in enhanced score submission: {str(e)}")
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect("enhanced_enter_scores")


def get_or_create_class_subject(assignment, user_school):
    """Get or create ClassSubject with proper school context."""

    # First check for active ClassSubject

    class_subject_query = ClassSubject.objects.filter(
        subject=assignment.subject,
        class_name=assignment.class_assigned,
        academic_year=assignment.academic_year,

        is_active=True

    )

    if user_school:
        class_subject_query = class_subject_query.filter(
            subject__school=user_school,
            class_name__school=user_school,
            academic_year__school=user_school,
        )

    class_subject = class_subject_query.first()

    if not class_subject:

        # Check if there's an inactive ClassSubject that we can reactivate
        inactive_class_subject_query = ClassSubject.objects.filter(
            subject=assignment.subject,
            class_name=assignment.class_assigned,
            academic_year=assignment.academic_year,
            is_active=False
        )

        if user_school:
            inactive_class_subject_query = inactive_class_subject_query.filter(
                subject__school=user_school,
                class_name__school=user_school,
                academic_year__school=user_school,
            )

        inactive_class_subject = inactive_class_subject_query.first()

        if inactive_class_subject:
            # Reactivate the existing assignment
            inactive_class_subject.is_active = True
            inactive_class_subject.save()
            class_subject = inactive_class_subject
        else:
            # Only create if this is an explicit assignment (not automatic)
            # This should only happen through the admin interface
            raise ValueError(f"No ClassSubject assignment found for {assignment.subject.subject_name} in {assignment.class_assigned.name}. Please assign this subject to the class first.")


    return class_subject


def validate_score_ranges(
    individual_score,
    class_test_score,
    project_score,
    group_work_score,
    exam_score,
    scoring_config,
    student,
):
    """Validate that all scores are within acceptable ranges."""
    errors = []

    # Component scores should be within their respective max marks
    component_scores = [
        (individual_score, "Individual", scoring_config.individual_max_mark),
        (class_test_score, "Class Test", scoring_config.class_test_max_mark),
        (project_score, "Project", scoring_config.project_max_mark),
        (group_work_score, "Group Work", scoring_config.group_work_max_mark),
    ]

    for score, component_name, max_mark in component_scores:
        if score is not None and (score < 0 or score > max_mark):
            errors.append(
                f"Invalid {component_name.lower()} score for {student.full_name}: {score}. "
                f"Must be between 0 and {max_mark}."
            )

    # Exam score validation - always out of 100
    if exam_score is not None and (exam_score < 0 or exam_score > 100):
        errors.append(
            f"Invalid exam score for {student.full_name}: {exam_score}. "
            f"Must be between 0 and 100."
        )

    return errors


@transaction.atomic
def update_or_create_assessment(
    class_subject,
    student,
    individual_score,
    class_test_score,
    project_score,
    group_work_score,
    exam_score,
    recorded_by,
    term,
):
    """Update or create assessment with individual score components."""

    # Get existing assessment or create new one

    try:
        assessment, created = Assessment.objects.get_or_create(
            class_subject=class_subject,
            student=student,
            term=term,
            assessment_type="exam_score",  # Using exam_score as the main assessment type
            defaults={
                "recorded_by": recorded_by,
            },
        )
    except Exception as e:
        # Handle duplicate key errors or other database issues
        logger.error(f"Error creating/updating assessment for {student.full_name}: {str(e)}")
        
        # Try to get existing assessment if creation failed
        try:
            assessment = Assessment.objects.get(
                class_subject=class_subject,
                student=student,
                term=term,
                assessment_type="exam_score"
            )
            created = False
        except Assessment.DoesNotExist:
            # If we can't find existing assessment, re-raise the original error
            raise e


    # Update individual score components
    if individual_score is not None:
        assessment.individual_score = individual_score
    if class_test_score is not None:
        assessment.class_test_score = class_test_score
    if project_score is not None:
        assessment.project_score = project_score
    if group_work_score is not None:
        assessment.group_work_score = group_work_score

    # Get scoring configuration to calculate scaled exam score
    scoring_config = ScoringConfiguration.get_active_config(
        class_subject.class_name.school
    )

    # Update exam score - store both raw and scaled values
    if exam_score is not None:
        # Store the raw exam score
        assessment.raw_exam_score = exam_score

        if scoring_config:
            # Calculate and store the scaled exam score
            scaled_exam_score = scoring_config.calculate_exam_score(exam_score)
            assessment.exam_score = scaled_exam_score
        else:
            # Fallback to raw score if no configuration
            assessment.exam_score = exam_score

    # Update recorded_by for existing assessments
    assessment.recorded_by = recorded_by

    # Save the assessment (this will trigger automatic calculation in the model)
    assessment.save()

    return assessment


def calculate_class_statistics(student_classes):
    """Calculate statistics for the class performance."""
    if not student_classes:
        return None

    total_students = len(student_classes)
    graded_students = 0
    total_scores = []

    for student_class in student_classes:
        student = student_class.student
        if hasattr(student, "total_score") and student.total_score is not None:
            graded_students += 1
            total_scores.append(float(student.total_score))

    if not total_scores:
        return {
            "total_students": total_students,
            "graded_students": 0,
            "completion_percentage": 0,
            "highest_score": 0,
            "lowest_score": 0,
            "average_score": 0,
        }

    return {
        "total_students": total_students,
        "graded_students": graded_students,
        "completion_percentage": round((graded_students / total_students) * 100, 1),
        "highest_score": max(total_scores),
        "lowest_score": min(total_scores),
        "average_score": sum(total_scores) / len(total_scores),
    }


@login_required
@csrf_exempt
@require_POST
def save_individual_student_scores(request):
    """
    AJAX endpoint for saving individual student scores in real-time.
    """
    try:
        data = json.loads(request.body)
        student_id = data.get("student_id")
        assignment_id = data.get("assignment_id")

        # Validate required fields
        if not student_id or not assignment_id:
            return JsonResponse(
                {"error": "Missing student_id or assignment_id"}, status=400
            )

        # Get teacher and school context
        teacher = getattr(request.user, "teacher_profile", None)
        if not teacher:
            return JsonResponse({"error": "Teacher profile not found"}, status=403)

        user_school = get_user_school(request.user)

        # Get current term for the user's school
        if user_school:
            current_term = Term.objects.filter(
                is_current=True, school=user_school
            ).first()
        else:
            current_term = Term.objects.filter(is_current=True).first()

        if not current_term:
            return JsonResponse({"error": "No current term is set"}, status=400)

        # Get assignment with school context
        assignment_query = TeacherSubjectAssignment.objects.filter(
            id=assignment_id, teacher=teacher, is_active=True
        )
        if user_school:
            assignment_query = assignment_query.filter(school=user_school)

        assignment = assignment_query.first()
        if not assignment:
            return JsonResponse({"error": "Assignment not found"}, status=404)

        # Get student with school context
        student_query = Student.objects.filter(id=student_id)
        if user_school:
            student_query = student_query.filter(school=user_school)

        student = student_query.first()
        if not student:
            return JsonResponse({"error": "Student not found"}, status=404)

        # Get scoring configuration
        scoring_config = ScoringConfiguration.get_active_config(user_school)
        if not scoring_config:
            return JsonResponse({"error": "No scoring configuration found"}, status=400)

        # Get or create ClassSubject
        class_subject = get_or_create_class_subject(assignment, user_school)

        # Extract and validate scores
        individual_score = data.get("individual_score")
        class_test_score = data.get("class_test_score")
        project_score = data.get("project_score")
        group_work_score = data.get("group_work_score")
        exam_score = data.get("exam_score")

        # Convert to Decimal and validate
        try:
            individual_score = (
                Decimal(str(individual_score)) if individual_score is not None else None
            )
            class_test_score = (
                Decimal(str(class_test_score)) if class_test_score is not None else None
            )
            project_score = (
                Decimal(str(project_score)) if project_score is not None else None
            )
            group_work_score = (
                Decimal(str(group_work_score)) if group_work_score is not None else None
            )
            exam_score = Decimal(str(exam_score)) if exam_score is not None else None
        except (ValueError, InvalidOperation):
            return JsonResponse({"error": "Invalid score format"}, status=400)

        # Validate ranges
        validation_errors = validate_score_ranges(
            individual_score,
            class_test_score,
            project_score,
            group_work_score,
            exam_score,
            scoring_config,
            student,
        )

        if validation_errors:
            return JsonResponse({"error": validation_errors[0]}, status=400)

        # Update or create assessment
        with transaction.atomic():
            assessment = update_or_create_assessment(
                class_subject,
                student,
                individual_score,
                class_test_score,
                project_score,
                group_work_score,
                exam_score,
                request.user,
                current_term,
            )

        # Return calculated values
        response_data = {
            "success": True,
            "message": "Scores saved successfully",
            "calculated_values": {
                "class_score": (
                    float(assessment.class_score) if assessment.class_score else None
                ),
                "total_score": (
                    float(assessment.total_score) if assessment.total_score else None
                ),
                "grade": assessment.grade,
                "remarks": assessment.remarks,
                "position": assessment.position,
            },
        }

        return JsonResponse(response_data)

    except Exception as e:
        logger.error(f"Error in save_individual_student_scores: {str(e)}")
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def get_grading_info(request):
    """
    AJAX endpoint to get grading information for a given score.
    """
    try:
        score = request.GET.get("score")
        if not score:
            return JsonResponse({"error": "Score parameter required"}, status=400)

        score = Decimal(str(score))
        user_school = get_user_school(request.user)

        # Get grade information from grading system
        grade_info = GradingSystem.get_grade_for_score(score, user_school)

        if grade_info:
            return JsonResponse(
                {
                    "grade": grade_info.grade_letter,
                    "remarks": grade_info.remarks,
                    "min_score": float(grade_info.min_score),
                    "max_score": float(grade_info.max_score),
                }
            )
        else:
            return JsonResponse(
                {
                    "grade": "N/A",
                    "remarks": "Not Graded",
                    "min_score": 0,
                    "max_score": 100,
                }
            )

    except (ValueError, InvalidOperation):
        return JsonResponse({"error": "Invalid score format"}, status=400)
    except Exception as e:
        logger.error(f"Error in get_grading_info: {str(e)}")
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@csrf_exempt
def import_enhanced_scores(request):
    """
    Enhanced import function that handles individual score components
    (individual_score, class_test_score, project_score, group_work_score, exam_score)
    from Excel files.
    """

    logger.info(f"Import enhanced scores called by user: {request.user}")
    logger.info(f"Request method: {request.method}")
    logger.info(f"Request FILES: {list(request.FILES.keys())}")

    if request.method != "POST":
        return JsonResponse({"error": "Only POST method allowed"}, status=405)

    # Check if the user is authorized
    if request.user.role not in ["teacher", "admin"]:
        return JsonResponse(
            {"error": "Only teachers and administrators can import scores."}, status=403
        )

    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    # Get current term for the user's school
    if user_school:
        current_term = Term.objects.filter(is_current=True, school=user_school).first()
    else:
        current_term = Term.objects.filter(is_current=True).first()

    if not current_term:
        return JsonResponse({"error": "No current term is set"}, status=400)

    # Get the assignment ID (for class and subject)
    assignment_id = request.POST.get("assignment_id")
    if not assignment_id:
        return JsonResponse({"error": "No assignment selected."}, status=400)

    # Get the teacher instance (for teachers)
    teacher = getattr(request.user, "teacher_profile", None)
    if request.user.role == "teacher" and not teacher:
        return JsonResponse(
            {"error": "Teacher profile is not linked to your account."}, status=403
        )

    # Get the assignment with school context
    try:
        if request.user.role == "teacher" and teacher:
            assignment_query = TeacherSubjectAssignment.objects.filter(
                id=assignment_id, teacher=teacher, is_active=True
            )
        elif request.user.role == "admin":
            # Admin can access any assignment in their school
            assignment_query = TeacherSubjectAssignment.objects.filter(
                id=assignment_id, is_active=True
            )
        else:
            return JsonResponse(
                {"error": "Unauthorized access to assignment"}, status=403
            )

        # Apply school filter for multi-tenancy
        if user_school:
            assignment_query = assignment_query.filter(school=user_school)

        assignment = assignment_query.get()
    except TeacherSubjectAssignment.DoesNotExist:
        return JsonResponse({"error": "Invalid assignment selected."}, status=404)

    # Get current term with school context
    if user_school:
        current_term = Term.objects.filter(
            academic_year=assignment.academic_year, is_current=True, school=user_school
        ).first()
    else:
        current_term = Term.objects.filter(
            academic_year=assignment.academic_year, is_current=True
        ).first()

    if not current_term:
        return JsonResponse(
            {"error": "No active term found for this academic year."}, status=400
        )

    # Get the ClassSubject instance with school context
    try:
        class_subject_query = ClassSubject.objects.filter(
            subject=assignment.subject,
            class_name=assignment.class_assigned,
            academic_year=assignment.academic_year,

            is_active=True

        )

        # Apply school filter through related objects
        if user_school:
            class_subject_query = class_subject_query.filter(
                subject__school=user_school,
                class_name__school=user_school,
                academic_year__school=user_school,
            )

        class_subject = class_subject_query.first()

        if not class_subject:
            # Create with proper school association

                    # Don't create automatically - this should be done through admin interface
                    raise ValueError(f"No ClassSubject assignment found. Please assign this subject to the class first.")

    except Exception as e:
        logger.error(f"Error creating/retrieving ClassSubject: {str(e)}")
        return JsonResponse({"error": "Error setting up class subject."}, status=500)

    # Check if a file was uploaded
    if "scoreFile" not in request.FILES:
        return JsonResponse({"error": "No file was uploaded."}, status=400)

    file = request.FILES["scoreFile"]
    # Check if the file is an Excel file
    if not file.name.endswith((".xlsx", ".xls")):
        return JsonResponse(
            {"error": "Please upload an Excel file (.xlsx or .xls)."}, status=400
        )

    try:
        # Create a BytesIO object from the uploaded file
        file_content = BytesIO(file.read())

        # Load the workbook with data_only=True to get values instead of formulas
        wb = openpyxl.load_workbook(file_content, data_only=True)
        ws = wb.active

        # Verify the file is for the correct school (for multi-tenancy)
        if user_school and "H3" in ws and ws["H3"].value == "school_id":
            file_school_id = ws["I3"].value
            if file_school_id and int(file_school_id) != user_school.id:
                return JsonResponse(
                    {"error": "This file belongs to a different school."}, status=400
                )

        # Get scoring configuration
        scoring_config = ScoringConfiguration.get_active_config(user_school)
        if not scoring_config:
            return JsonResponse(
                {
                    "error": "No scoring configuration found. Please contact administrator."
                },
                status=400,
            )

        # Find the data rows (starting from row 8)
        start_row = 8
        updated_assessments = []
        updated_count = 0
        errors = []

        # Loop through the rows
        for row in range(start_row, ws.max_row + 1):
            # Get the student ID from column A
            student_id = ws.cell(row=row, column=1).value
            if not student_id:
                continue  # Skip empty rows

            # Find the student with school context
            try:
                student_query = Student.objects.filter(admission_number=student_id)

                # Apply school filter for multi-tenancy
                if user_school:
                    student_query = student_query.filter(school=user_school)

                student = student_query.get()
            except Student.DoesNotExist:
                errors.append(f"Student with ID {student_id} not found.")
                continue

            # Check if the student is in the class with school context
            student_class_query = StudentClass.objects.filter(
                student=student,
                assigned_class=assignment.class_assigned,
                is_active=True,
            )

            # Apply school filter for multi-tenancy
            if user_school:
                student_class_query = student_class_query.filter(school=user_school)

            if not student_class_query.exists():
                errors.append(
                    f"Student {student.full_name} ({student_id}) is not in this class."
                )
                continue

            # Get the enhanced score components
            try:
                # Enhanced scoring columns (adjust column numbers based on your Excel template)
                individual_score_cell = ws.cell(row=row, column=3)  # Column C
                class_test_score_cell = ws.cell(row=row, column=4)  # Column D
                project_score_cell = ws.cell(row=row, column=5)  # Column E
                group_work_score_cell = ws.cell(row=row, column=6)  # Column F
                exam_score_cell = ws.cell(row=row, column=7)  # Column G

                # Extract values
                individual_score_value = individual_score_cell.value
                class_test_score_value = class_test_score_cell.value
                project_score_value = project_score_cell.value
                group_work_score_value = group_work_score_cell.value
                exam_score_value = exam_score_cell.value

                # Handle different data types for scores
                def process_score_value(value):
                    if isinstance(value, str):
                        value = value.strip()
                        return float(value) if value else None
                    return value

                individual_score_value = process_score_value(individual_score_value)
                class_test_score_value = process_score_value(class_test_score_value)
                project_score_value = process_score_value(project_score_value)
                group_work_score_value = process_score_value(group_work_score_value)
                exam_score_value = process_score_value(exam_score_value)

                # Convert to decimal or None
                individual_score = (
                    Decimal(str(individual_score_value))
                    if individual_score_value not in ("", None)
                    else None
                )
                class_test_score = (
                    Decimal(str(class_test_score_value))
                    if class_test_score_value not in ("", None)
                    else None
                )
                project_score = (
                    Decimal(str(project_score_value))
                    if project_score_value not in ("", None)
                    else None
                )
                group_work_score = (
                    Decimal(str(group_work_score_value))
                    if group_work_score_value not in ("", None)
                    else None
                )
                exam_score = (
                    Decimal(str(exam_score_value))
                    if exam_score_value not in ("", None)
                    else None
                )

                # Skip if no scores provided
                if not any(
                    [
                        individual_score,
                        class_test_score,
                        project_score,
                        group_work_score,
                        exam_score,
                    ]
                ):
                    continue

                # Validate score ranges using scoring configuration
                validation_errors = validate_score_ranges(
                    individual_score,
                    class_test_score,
                    project_score,
                    group_work_score,
                    exam_score,
                    scoring_config,
                    student,
                )

                if validation_errors:
                    errors.extend(validation_errors)
                    continue

                # Update or create assessment using enhanced function
                assessment = update_or_create_assessment(
                    class_subject,
                    student,
                    individual_score,
                    class_test_score,
                    project_score,
                    group_work_score,
                    exam_score,
                    request.user,
                    current_term,
                )

                updated_assessments.append(assessment)
                updated_count += 1

            except (ValueError, InvalidOperation) as e:
                errors.append(f"Invalid score format for {student.full_name}: {str(e)}")
                continue

        # Calculate positions for all students in this class_subject
        if hasattr(Assessment, "calculate_positions"):
            Assessment.calculate_positions(class_subject, current_term)
        else:
            # Use optimized position calculation
            from ..views.scores import calculate_positions_optimized

            calculate_positions_optimized(class_subject, current_term)

        # Return success response
        if errors:
            return JsonResponse(
                {
                    "success": False,
                    "message": f"Import completed with {len(errors)} errors.",
                    "error_messages": errors[:10],  # First 10 errors
                    "total_errors": len(errors),
                    "updated_count": updated_count,
                }
            )
        else:
            return JsonResponse(
                {
                    "success": True,
                    "message": f"Successfully imported enhanced scores for {updated_count} students.",
                    "updated_count": updated_count,
                }
            )

    except Exception as e:
        logger.error(f"Error importing enhanced scores: {str(e)}", exc_info=True)
        return JsonResponse(
            {
                "success": False,
                "message": f"Error importing scores: {str(e)}",
                "error_messages": [str(e)],
            }
        )


@login_required
@csrf_exempt
def import_enhanced_scores_batch(request):
    """
    Enhanced batch import function that handles individual score components
    for multiple classes from Excel files with multiple sheets.
    """

    logger.info(f"Import enhanced scores batch called by user: {request.user}")
    logger.info(f"Request method: {request.method}")
    logger.info(f"Request FILES: {list(request.FILES.keys())}")

    if request.method != "POST":
        return JsonResponse({"error": "Only POST method allowed"}, status=405)

    # Check if the user is authorized
    if request.user.role not in ["teacher", "admin"]:
        return JsonResponse(
            {"error": "Only teachers and administrators can import scores."}, status=403
        )

    # Check if a file was uploaded
    if "batchFile" not in request.FILES:
        return JsonResponse({"error": "No file was uploaded."}, status=400)

    file = request.FILES["batchFile"]
    # Check if the file is an Excel file
    if not file.name.endswith((".xlsx", ".xls")):
        return JsonResponse(
            {"error": "Please upload an Excel file (.xlsx or .xls)."}, status=400
        )

    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    # Get current term for the user's school
    if user_school:
        current_term = Term.objects.filter(is_current=True, school=user_school).first()
    else:
        current_term = Term.objects.filter(is_current=True).first()

    if not current_term:
        return JsonResponse({"error": "No current term is set"}, status=400)

    # Get the teacher instance (for teachers)
    teacher = getattr(request.user, "teacher_profile", None)
    if request.user.role == "teacher" and not teacher:
        return JsonResponse(
            {"error": "Teacher profile is not linked to your account."}, status=403
        )

    try:
        # Create a BytesIO object from the uploaded file
        file_content = BytesIO(file.read())

        # Load the workbook with data_only=True to get values instead of formulas
        wb = openpyxl.load_workbook(file_content, data_only=True)

        # Get scoring configuration
        scoring_config = ScoringConfiguration.get_active_config(user_school)
        if not scoring_config:
            return JsonResponse(
                {
                    "error": "No scoring configuration found. Please contact administrator."
                },
                status=400,
            )

        results = []
        total_updated = 0
        total_errors = []

        # Process each sheet (each sheet represents a class-subject combination)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Skip if sheet is empty or doesn't contain expected data
            if ws.max_row < 8:
                continue

            # Extract class and subject information from sheet name or headers
            # Handle both original and sanitized sheet names
            # Format: "Class_Subject" or "Class - Subject" (may be truncated)

            # First, try to get the original names from the Excel file metadata
            class_name = None
            subject_name = None

            try:
                # Check if the sheet has metadata in specific cells
                if ws["A2"].value and "Class:" in str(ws["A2"].value):
                    class_name = str(ws["A2"].value).replace("Class:", "").strip()
                if ws["A3"].value and "Subject:" in str(ws["A3"].value):
                    subject_name = str(ws["A3"].value).replace("Subject:", "").strip()
            except:
                pass

            # If metadata not found, try to parse from sheet name
            if not class_name or not subject_name:
                class_subject_parts = sheet_name.replace("_", " - ").split(" - ")
                if len(class_subject_parts) >= 2:
                    class_name = class_subject_parts[0].strip()
                    subject_name = class_subject_parts[1].strip()
                else:
                    total_errors.append(
                        f"Invalid sheet name format: {sheet_name}. Expected 'Class - Subject'"
                    )
                    continue

            if not class_name or not subject_name:
                total_errors.append(
                    f"Could not extract class and subject from sheet: {sheet_name}"
                )
                continue

            # Find the assignment with improved matching
            try:
                assignment = find_assignment_for_import(
                    class_name, subject_name, request.user, teacher, user_school
                )

                if not assignment:
                    # Get available assignments for debugging
                    available_assignments = get_available_assignments_for_debugging(
                        request.user, teacher, user_school
                    )

                    # Also try to find similar assignments
                    similar_assignments = find_similar_assignments(
                        class_name, subject_name, request.user, teacher, user_school
                    )

                    error_msg = (
                        f"Assignment not found for '{class_name}' - '{subject_name}'"
                    )
                    if similar_assignments:
                        error_msg += (
                            f". Similar assignments found: {similar_assignments}"
                        )
                    error_msg += f". Available assignments: {available_assignments}"

                    total_errors.append(error_msg)
                    continue

            except Exception as e:
                total_errors.append(
                    f"Error finding assignment for {sheet_name}: {str(e)}"
                )
                continue

            # Get or create ClassSubject
            try:
                class_subject_query = ClassSubject.objects.filter(
                    subject=assignment.subject,
                    class_name=assignment.class_assigned,
                    academic_year=assignment.academic_year,

                    is_active=True

                )

                # Apply school filter through related objects
                if user_school:
                    class_subject_query = class_subject_query.filter(
                        subject__school=user_school,
                        class_name__school=user_school,
                        academic_year__school=user_school,
                    )

                class_subject = class_subject_query.first()

                if not class_subject:
                    # Create with proper school association

                    # Don't create automatically - this should be done through admin interface
                    raise ValueError(f"No ClassSubject assignment found. Please assign this subject to the class first.")

            except Exception as e:
                total_errors.append(
                    f"Error creating/retrieving ClassSubject for {sheet_name}: {str(e)}"
                )
                continue

            # Process the sheet data
            start_row = 8
            sheet_updated = 0
            sheet_errors = []

            # Loop through the rows
            for row in range(start_row, ws.max_row + 1):
                # Get the student ID from column A
                student_id = ws.cell(row=row, column=1).value
                if not student_id:
                    continue  # Skip empty rows

                # Find the student with school context
                try:
                    student_query = Student.objects.filter(admission_number=student_id)

                    # Apply school filter for multi-tenancy
                    if user_school:
                        student_query = student_query.filter(school=user_school)

                    student = student_query.get()
                except Student.DoesNotExist:
                    sheet_errors.append(
                        f"Student with ID {student_id} not found in {sheet_name}"
                    )
                    continue

                # Check if the student is in the class with school context
                student_class_query = StudentClass.objects.filter(
                    student=student,
                    assigned_class=assignment.class_assigned,
                    is_active=True,
                )

                # Apply school filter for multi-tenancy
                if user_school:
                    student_class_query = student_class_query.filter(school=user_school)

                if not student_class_query.exists():
                    sheet_errors.append(
                        f"Student {student.full_name} ({student_id}) is not in class {class_name}"
                    )
                    continue

                # Get the enhanced score components
                try:
                    # Enhanced scoring columns
                    individual_score_cell = ws.cell(row=row, column=3)  # Column C
                    class_test_score_cell = ws.cell(row=row, column=4)  # Column D
                    project_score_cell = ws.cell(row=row, column=5)  # Column E
                    group_work_score_cell = ws.cell(row=row, column=6)  # Column F
                    exam_score_cell = ws.cell(row=row, column=7)  # Column G

                    # Extract values
                    individual_score_value = individual_score_cell.value
                    class_test_score_value = class_test_score_cell.value
                    project_score_value = project_score_cell.value
                    group_work_score_value = group_work_score_cell.value
                    exam_score_value = exam_score_cell.value

                    # Handle different data types for scores
                    def process_score_value(value):
                        if isinstance(value, str):
                            value = value.strip()
                            return float(value) if value else None
                        return value

                    individual_score_value = process_score_value(individual_score_value)
                    class_test_score_value = process_score_value(class_test_score_value)
                    project_score_value = process_score_value(project_score_value)
                    group_work_score_value = process_score_value(group_work_score_value)
                    exam_score_value = process_score_value(exam_score_value)

                    # Convert to decimal or None
                    individual_score = (
                        Decimal(str(individual_score_value))
                        if individual_score_value not in ("", None)
                        else None
                    )
                    class_test_score = (
                        Decimal(str(class_test_score_value))
                        if class_test_score_value not in ("", None)
                        else None
                    )
                    project_score = (
                        Decimal(str(project_score_value))
                        if project_score_value not in ("", None)
                        else None
                    )
                    group_work_score = (
                        Decimal(str(group_work_score_value))
                        if group_work_score_value not in ("", None)
                        else None
                    )
                    exam_score = (
                        Decimal(str(exam_score_value))
                        if exam_score_value not in ("", None)
                        else None
                    )

                    # Skip if no scores provided
                    if not any(
                        [
                            individual_score,
                            class_test_score,
                            project_score,
                            group_work_score,
                            exam_score,
                        ]
                    ):
                        continue

                    # Validate score ranges using scoring configuration
                    validation_errors = validate_score_ranges(
                        individual_score,
                        class_test_score,
                        project_score,
                        group_work_score,
                        exam_score,
                        scoring_config,
                        student,
                    )

                    if validation_errors:
                        sheet_errors.extend(
                            [f"{sheet_name}: {error}" for error in validation_errors]
                        )
                        continue

                    # Update or create assessment using enhanced function
                    assessment = update_or_create_assessment(
                        class_subject,
                        student,
                        individual_score,
                        class_test_score,
                        project_score,
                        group_work_score,
                        exam_score,
                        request.user,
                        current_term,
                    )

                    sheet_updated += 1

                except (ValueError, InvalidOperation) as e:
                    sheet_errors.append(
                        f"Invalid score format for {student.full_name} in {sheet_name}: {str(e)}"
                    )
                    continue

            # Calculate positions for all students in this class_subject
            if hasattr(Assessment, "calculate_positions"):
                Assessment.calculate_positions(class_subject, current_term)
            else:
                # Use optimized position calculation
                from ..views.scores import calculate_positions_optimized

                calculate_positions_optimized(class_subject, current_term)

            # Add the result for this sheet
            results.append(
                {
                    "class_name": assignment.class_assigned.name,
                    "subject_name": assignment.subject.subject_name,
                    "processed": sheet_updated,
                    "sheet_name": sheet_name,
                    "updated_count": sheet_updated,
                    "errors": sheet_errors,
                }
            )
            total_updated += sheet_updated
            total_errors.extend(sheet_errors)

        # Return success response
        if total_errors:
            return JsonResponse(
                {
                    "success": False,
                    "message": f"Batch import completed with {len(total_errors)} errors.",
                    "error_messages": total_errors[:10],  # First 10 errors
                    "total_errors": len(total_errors),
                    "total_updated": total_updated,
                    "results": results,
                }
            )
        else:
            return JsonResponse(
                {
                    "success": True,
                    "message": f"Successfully imported enhanced scores for {total_updated} students across {len(results)} classes.",
                    "total_updated": total_updated,
                    "results": results,
                }
            )

    except Exception as e:
        logger.error(f"Error importing enhanced batch scores: {str(e)}", exc_info=True)
        return JsonResponse(
            {
                "success": False,
                "message": f"Error importing batch scores: {str(e)}",
                "error_messages": [str(e)],
            }
        )


@login_required
def export_enhanced_scores(request):
    """
    Export enhanced score template for a single class with individual score components.
    Creates an Excel template that can be used for importing enhanced scores.
    """
    # Check if the user is authorized
    if request.user.role not in ["teacher", "admin"]:
        messages.error(request, "Only teachers and administrators can export scores.")
        return redirect("enhanced_enter_scores")

    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    # Get the assignment ID from the request
    assignment_id = request.GET.get("assignment_id")
    if not assignment_id:

        messages.error(request, "Please select a class to export scores for a single class, or use batch export for multiple classes.")

        return redirect("enhanced_enter_scores")

    # Get the teacher instance (for teachers)
    teacher = getattr(request.user, "teacher_profile", None)
    if request.user.role == "teacher" and not teacher:
        messages.error(request, "Teacher profile is not linked to this user.")
        return redirect("login")

    # Get the assignment with school context
    try:
        if request.user.role == "teacher" and teacher:
            assignment_query = TeacherSubjectAssignment.objects.filter(
                id=assignment_id, teacher=teacher, is_active=True
            )
        elif request.user.role == "admin":
            assignment_query = TeacherSubjectAssignment.objects.filter(
                id=assignment_id, is_active=True
            )
        else:
            messages.error(request, "Unauthorized access to assignment.")
            return redirect("enhanced_enter_scores")

        # Apply school filter for multi-tenancy
        if user_school:
            assignment_query = assignment_query.filter(school=user_school)

        assignment = assignment_query.get()

        
        # Check if the ClassSubject is active
        class_subject = ClassSubject.objects.filter(
            subject=assignment.subject,
            class_name=assignment.class_assigned,
            academic_year=assignment.academic_year,
            is_active=True
        ).first()
        
        if not class_subject:
            messages.error(request, f"Subject '{assignment.subject.subject_name}' is not currently assigned to class '{assignment.class_assigned.name}'. Cannot export scores for removed subjects.")
            return redirect("enhanced_enter_scores")
            

    except TeacherSubjectAssignment.DoesNotExist:
        messages.error(request, "Invalid assignment selected.")
        return redirect("enhanced_enter_scores")

    # Get current term with school context
    if user_school:
        current_term = Term.objects.filter(
            academic_year=assignment.academic_year, is_current=True, school=user_school
        ).first()
    else:
        current_term = Term.objects.filter(
            academic_year=assignment.academic_year, is_current=True
        ).first()

    if not current_term:
        messages.error(request, "No active term found for this academic year.")
        return redirect("enhanced_enter_scores")

    # Get the ClassSubject instance with school context
    try:
        class_subject_query = ClassSubject.objects.filter(
            subject=assignment.subject,
            class_name=assignment.class_assigned,
            academic_year=assignment.academic_year,

            is_active=True

        )

        # Apply school filter through related objects
        if user_school:
            class_subject_query = class_subject_query.filter(
                subject__school=user_school,
                class_name__school=user_school,
                academic_year__school=user_school,
            )

        class_subject = class_subject_query.first()

        if not class_subject:
            # Create with proper school association

                    # Don't create automatically - this should be done through admin interface
                    raise ValueError(f"No ClassSubject assignment found. Please assign this subject to the class first.")

    except Exception as e:
        logger.error(f"Error creating/retrieving ClassSubject: {str(e)}")
        messages.error(request, "Error setting up class subject.")
        return redirect("enhanced_enter_scores")

    # Get students in class with school context
    student_classes_query = StudentClass.objects.filter(
        assigned_class=assignment.class_assigned, is_active=True
    ).select_related("student")

    if user_school:
        student_classes_query = student_classes_query.filter(
            student__school=user_school
        )

    student_classes = student_classes_query.order_by("student__full_name")

    # Get scoring configuration
    scoring_config = ScoringConfiguration.get_active_config(user_school)
    if not scoring_config:
        messages.error(
            request, "No scoring configuration found. Please contact administrator."
        )
        return redirect("enhanced_enter_scores")

    try:
        # Create a BytesIO object for the Excel file
        output = BytesIO()

        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Create sanitized sheet name
        sheet_name = (
            f"{assignment.class_assigned.name} - {assignment.subject.subject_name}"
        )
        ws.title = sanitize_excel_sheet_name(sheet_name)

        # Define styles
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Add school information
        if user_school:
            ws["A1"] = f"School: {user_school.name}"
            ws["A2"] = f"Class: {assignment.class_assigned.name}"
            ws["A3"] = f"Subject: {assignment.subject.subject_name}"
            ws["A4"] = f"Academic Year: {assignment.academic_year.name}"
            ws["A5"] = f"Term: {current_term.term_number}"
        else:
            ws["A1"] = f"Class: {assignment.class_assigned.name}"
            ws["A2"] = f"Subject: {assignment.subject.subject_name}"
            ws["A3"] = f"Academic Year: {assignment.academic_year.name}"
            ws["A4"] = f"Term: {current_term.term_number}"

        # Add school ID for multi-tenancy validation
        if user_school:
            ws["H3"] = "school_id"
            ws["I3"] = user_school.id

        # Make headers bold
        for row in range(1, 6):
            ws.cell(row=row, column=1).font = Font(bold=True)

        # Add instructions
        instructions = (
            f"Instructions: Edit the score columns (C-G) with values between 0 and their maximum values. "
            f"Individual Score (max: {scoring_config.individual_max_mark}), "
            f"Class Test Score (max: {scoring_config.class_test_max_mark}), "
            f"Project Score (max: {scoring_config.project_max_mark}), "
            f"Group Work Score (max: {scoring_config.group_work_max_mark}), "
            f"Exam Score (max: 100). "
            f"Class Score and Total Score are calculated automatically."
        )
        ws["A6"] = instructions
        ws.merge_cells("A6:K6")
        ws["A6"].font = Font(bold=True, color="FF0000")
        ws["A6"].alignment = Alignment(horizontal="left", wrap_text=True)

        # Define enhanced column headers
        headers = [
            "Student ID",
            "Student Name",
            f"Individual Score ({scoring_config.individual_max_mark})",
            f"Class Test Score ({scoring_config.class_test_max_mark})",
            f"Project Score ({scoring_config.project_max_mark})",
            f"Group Work Score ({scoring_config.group_work_max_mark})",
            "Exam Score (100)",
            "Class Score (Calculated)",
            "Total Score (Calculated)",
            "Grade (Calculated)",
            "Remarks (Calculated)",
            "Position (Calculated)",
        ]

        # Apply headers to row 7
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # Get existing assessments for all students in this class and subject

        # Exclude mock exam assessments - only get regular term assessments

        assessments = Assessment.objects.filter(
            class_subject=class_subject,
            student__in=[sc.student for sc in student_classes],
            term=current_term,

        ).exclude(assessment_type='mock_exam').select_related("student")


        # Create a dictionary for quick lookup
        student_assessments = {
            assessment.student.id: assessment for assessment in assessments
        }

        # Add data rows
        row_num = 8
        for student_class in student_classes:
            student = student_class.student
            assessment = student_assessments.get(student.id)

            # Student ID and Name
            ws.cell(row=row_num, column=1, value=student.admission_number)
            ws.cell(row=row_num, column=2, value=student.full_name)

            # Individual score components (editable)
            if assessment:
                ws.cell(row=row_num, column=3, value=assessment.individual_score or "")
                ws.cell(row=row_num, column=4, value=assessment.class_test_score or "")
                ws.cell(row=row_num, column=5, value=assessment.project_score or "")
                ws.cell(row=row_num, column=6, value=assessment.group_work_score or "")
                ws.cell(row=row_num, column=7, value=assessment.raw_exam_score or "")

                # Calculated fields (read-only, for reference)
                ws.cell(row=row_num, column=8, value=assessment.class_score or "")
                ws.cell(row=row_num, column=9, value=assessment.total_score or "")
                ws.cell(row=row_num, column=10, value=assessment.grade or "")
                ws.cell(row=row_num, column=11, value=assessment.remarks or "")
                ws.cell(row=row_num, column=12, value=assessment.position or "")
            else:
                # Empty cells for new students
                ws.cell(row=row_num, column=3, value="")
                ws.cell(row=row_num, column=4, value="")
                ws.cell(row=row_num, column=5, value="")
                ws.cell(row=row_num, column=6, value="")
                ws.cell(row=row_num, column=7, value="")

                # Empty calculated fields
                ws.cell(row=row_num, column=8, value="")
                ws.cell(row=row_num, column=9, value="")
                ws.cell(row=row_num, column=10, value="")
                ws.cell(row=row_num, column=11, value="")
                ws.cell(row=row_num, column=12, value="")

            # Apply borders to all cells in the row
            for col in range(1, 13):
                ws.cell(row=row_num, column=col).border = border

            row_num += 1

        # Adjust column widths
        column_widths = [15, 30, 20, 20, 15, 20, 15, 20, 20, 10, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Save the workbook to BytesIO
        wb.save(output)
        output.seek(0)

        # Create response
        filename = f"enhanced_scores_{assignment.class_assigned.name}_{assignment.subject.subject_name}_{current_term.term_number}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        logger.error(f"Error exporting enhanced scores: {str(e)}", exc_info=True)
        messages.error(request, f"Error exporting scores: {str(e)}")
        return redirect("enhanced_enter_scores")


@login_required
def export_enhanced_scores_batch(request):
    """
    Export enhanced score templates for multiple classes with individual score components.
    Creates an Excel file with multiple sheets for batch import.
    """
    # Check if the user is authorized
    if request.user.role not in ["teacher", "admin"]:
        messages.error(request, "Only teachers and administrators can export scores.")
        return redirect("enhanced_enter_scores")

    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    # Get the teacher instance (for teachers)
    teacher = getattr(request.user, "teacher_profile", None)
    if request.user.role == "teacher" and not teacher:
        messages.error(request, "Teacher profile is not linked to this user.")
        return redirect("login")

    # Get current academic year and term with school context
    if user_school:
        current_academic_year = AcademicYear.objects.filter(
            is_current=True, school=user_school
        ).first()
        current_term = Term.objects.filter(
            academic_year=current_academic_year, is_current=True, school=user_school
        ).first()
    else:
        current_academic_year = AcademicYear.objects.filter(is_current=True).first()
        current_term = Term.objects.filter(
            academic_year=current_academic_year, is_current=True
        ).first()

    if not current_academic_year or not current_term:
        messages.error(request, "No active academic year or term found.")
        return redirect("enhanced_enter_scores")

    # Check if specific assignment IDs are provided
    assignment_ids = request.GET.get("assignment_ids")
    if assignment_ids:
        # Parse the comma-separated assignment IDs
        try:
            assignment_id_list = [int(id.strip()) for id in assignment_ids.split(",")]
        except ValueError:
            messages.error(request, "Invalid assignment IDs provided.")
            return redirect("enhanced_enter_scores")

        # Get specific assignments
        assignments_query = TeacherSubjectAssignment.objects.filter(
            id__in=assignment_id_list, is_active=True
        )

        # Apply authorization check
        if request.user.role == "teacher" and teacher:
            assignments_query = assignments_query.filter(teacher=teacher)
        elif request.user.role != "admin":
            messages.error(request, "Unauthorized access.")
            return redirect("enhanced_enter_scores")
    else:
        # Get all assignments for the teacher (original behavior)
        if request.user.role == "teacher" and teacher:
            assignments_query = TeacherSubjectAssignment.objects.filter(
                teacher=teacher, is_active=True, academic_year=current_academic_year
            )
        elif request.user.role == "admin":
            assignments_query = TeacherSubjectAssignment.objects.filter(
                is_active=True, academic_year=current_academic_year
            )
        else:
            messages.error(request, "Unauthorized access.")
            return redirect("enhanced_enter_scores")

    # Apply school filter for multi-tenancy
    if user_school:
        assignments_query = assignments_query.filter(school=user_school)


    # Filter out assignments where ClassSubject is not active
    # Get active class-subject combinations
    active_class_subjects = ClassSubject.objects.filter(
        academic_year=current_academic_year, is_active=True
    ).values_list('class_name_id', 'subject_id')
    
    # Filter assignments to only include those with active ClassSubject
    filtered_assignments = []
    for assignment in assignments_query.select_related("class_assigned", "subject", "academic_year"):
        if (assignment.class_assigned.id, assignment.subject.id) in active_class_subjects:
            filtered_assignments.append(assignment)
    
    assignments = filtered_assignments


    if not assignments:
        messages.error(request, "No assignments found for export.")
        return redirect("enhanced_enter_scores")

    # Get scoring configuration
    scoring_config = ScoringConfiguration.get_active_config(user_school)
    if not scoring_config:
        messages.error(
            request, "No scoring configuration found. Please contact administrator."
        )
        return redirect("enhanced_enter_scores")

    try:
        # Create a BytesIO object for the Excel file
        output = BytesIO()

        # Create a new workbook
        wb = openpyxl.Workbook()

        # Remove the default worksheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Define styles
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Process each assignment
        processed_assignments = 0

        for assignment in assignments:
            # Create sanitized sheet name
            sheet_name = (
                f"{assignment.class_assigned.name} - {assignment.subject.subject_name}"
            )
            sanitized_sheet_name = sanitize_excel_sheet_name(sheet_name)

            # Create worksheet
            ws = wb.create_sheet(title=sanitized_sheet_name)

            # Add school information
            if user_school:
                ws["A1"] = f"School: {user_school.name}"
                ws["A2"] = f"Class: {assignment.class_assigned.name}"
                ws["A3"] = f"Subject: {assignment.subject.subject_name}"
                ws["A4"] = f"Academic Year: {assignment.academic_year.name}"
                ws["A5"] = f"Term: {current_term.term_number}"
            else:
                ws["A1"] = f"Class: {assignment.class_assigned.name}"
                ws["A2"] = f"Subject: {assignment.subject.subject_name}"
                ws["A3"] = f"Academic Year: {assignment.academic_year.name}"
                ws["A4"] = f"Term: {current_term.term_number}"

            # Add school ID for multi-tenancy validation
            if user_school:
                ws["H3"] = "school_id"
                ws["I3"] = user_school.id

            # Make headers bold
            for row in range(1, 6):
                ws.cell(row=row, column=1).font = Font(bold=True)

            # Add instructions
            instructions = (
                f"Instructions: Edit the score columns (C-G) with values between 0 and their maximum values. "
                f"Individual Score (max: {scoring_config.individual_max_mark}), "
                f"Class Test Score (max: {scoring_config.class_test_max_mark}), "
                f"Project Score (max: {scoring_config.project_max_mark}), "
                f"Group Work Score (max: {scoring_config.group_work_max_mark}), "
                f"Exam Score (max: 100). "
                f"Class Score and Total Score are calculated automatically."
            )
            ws["A6"] = instructions
            ws.merge_cells("A6:K6")
            ws["A6"].font = Font(bold=True, color="FF0000")
            ws["A6"].alignment = Alignment(horizontal="left", wrap_text=True)

            # Define enhanced column headers
            headers = [
                "Student ID",
                "Student Name",
                f"Individual Score ({scoring_config.individual_max_mark})",
                f"Class Test Score ({scoring_config.class_test_max_mark})",
                f"Project Score ({scoring_config.project_max_mark})",
                f"Group Work Score ({scoring_config.group_work_max_mark})",
                "Exam Score (100)",
                "Class Score (Calculated)",
                "Total Score (Calculated)",
                "Grade (Calculated)",
                "Remarks (Calculated)",
                "Position (Calculated)",
            ]

            # Apply headers to row 7
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

            # Get or create ClassSubject
            try:
                class_subject_query = ClassSubject.objects.filter(
                    subject=assignment.subject,
                    class_name=assignment.class_assigned,
                    academic_year=assignment.academic_year,

                    is_active=True

                )

                # Apply school filter through related objects
                if user_school:
                    class_subject_query = class_subject_query.filter(
                        subject__school=user_school,
                        class_name__school=user_school,
                        academic_year__school=user_school,
                    )

                class_subject = class_subject_query.first()

                if not class_subject:
                    # Create with proper school association

                    # Don't create automatically - this should be done through admin interface
                    raise ValueError(f"No ClassSubject assignment found. Please assign this subject to the class first.")

            except Exception as e:
                logger.error(
                    f"Error creating/retrieving ClassSubject for {assignment}: {str(e)}"
                )
                continue

            # Get students in class with school context
            student_classes_query = StudentClass.objects.filter(
                assigned_class=assignment.class_assigned, is_active=True
            ).select_related("student")

            if user_school:
                student_classes_query = student_classes_query.filter(
                    student__school=user_school
                )

            student_classes = student_classes_query.order_by("student__full_name")

            # Get existing assessments for all students in this class and subject

            # Exclude mock exam assessments - only get regular term assessments

            assessments = Assessment.objects.filter(
                class_subject=class_subject,
                student__in=[sc.student for sc in student_classes],
                term=current_term,

            ).exclude(assessment_type='mock_exam').select_related("student")


            # Create a dictionary for quick lookup
            student_assessments = {
                assessment.student.id: assessment for assessment in assessments
            }

            # Add data rows
            row_num = 8
            for student_class in student_classes:
                student = student_class.student
                assessment = student_assessments.get(student.id)

                # Student ID and Name
                ws.cell(row=row_num, column=1, value=student.admission_number)
                ws.cell(row=row_num, column=2, value=student.full_name)

                # Individual score components (editable)
                if assessment:
                    ws.cell(
                        row=row_num, column=3, value=assessment.individual_score or ""
                    )
                    ws.cell(
                        row=row_num, column=4, value=assessment.class_test_score or ""
                    )
                    ws.cell(row=row_num, column=5, value=assessment.project_score or "")
                    ws.cell(
                        row=row_num, column=6, value=assessment.group_work_score or ""
                    )
                    ws.cell(
                        row=row_num, column=7, value=assessment.raw_exam_score or ""
                    )
                else:
                    # Empty cells for new students
                    ws.cell(row=row_num, column=3, value="")
                    ws.cell(row=row_num, column=4, value="")
                    ws.cell(row=row_num, column=5, value="")
                    ws.cell(row=row_num, column=6, value="")
                    ws.cell(row=row_num, column=7, value="")

                # Calculated fields (read-only, for reference)
                if assessment:
                    ws.cell(row=row_num, column=8, value=assessment.class_score or "")
                    ws.cell(row=row_num, column=9, value=assessment.total_score or "")
                    ws.cell(row=row_num, column=10, value=assessment.grade or "")
                    ws.cell(row=row_num, column=11, value=assessment.remarks or "")
                    ws.cell(row=row_num, column=12, value=assessment.position or "")
                else:
                    ws.cell(row=row_num, column=8, value="")
                    ws.cell(row=row_num, column=9, value="")
                    ws.cell(row=row_num, column=10, value="")
                    ws.cell(row=row_num, column=11, value="")
                    ws.cell(row=row_num, column=12, value="")

                # Apply borders to all cells in the row
                for col in range(1, 13):
                    ws.cell(row=row_num, column=col).border = border

                row_num += 1

            # Adjust column widths
            column_widths = [15, 30, 20, 20, 15, 20, 15, 20, 20, 10, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            processed_assignments += 1

        # Save the workbook to BytesIO
        wb.save(output)
        output.seek(0)

        # Create response
        filename = f"enhanced_scores_batch_{current_term.term_number}_{processed_assignments}_classes.xlsx"
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        logger.error(f"Error exporting enhanced batch scores: {str(e)}", exc_info=True)
        messages.error(request, f"Error exporting batch scores: {str(e)}")
        return redirect("enhanced_enter_scores")
