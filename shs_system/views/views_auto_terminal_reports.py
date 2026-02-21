"""
Automatic Terminal Report Card Generation System

This module provides automatic generation of terminal report cards based on
student data without requiring manual bulk generation. It includes flexible
filtering, sorting, and both bulk and individual printing capabilities.
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Q, Avg, Count, F, Sum
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.contrib import messages
from django.utils import timezone
from django.db import transaction
from django.template.loader import render_to_string
from django.conf import settings
import json
from datetime import datetime, timedelta

from ..models import (
    Student, Class, Term, AcademicYear, Assessment, 
    ReportCard, StudentClass, StudentTermRemarks,
    Subject, ClassSubject, GradingSystem, PerformanceRequirement,
    SchoolInformation, SchoolAuthoritySignature
)
from ..utils import get_user_school
from .school_info import set_sweet_alert
from ..utils.pdf_generator import generate_pdf_from_html


def apply_school_filter(queryset, user_school, school_field='school'):
    """
    Apply school filter to a queryset if user_school is not None.
    """
    if user_school:
        return queryset.filter(**{school_field: user_school})
    return queryset


def auto_generate_reports_for_term(user_school, term, generated_by_user):
    """
    Automatically generate report cards for all students in a term.
    This runs in the background without user intervention.
    """
    try:
        # Get all students for this term
        if user_school:
            student_classes = StudentClass.objects.filter(
                school=user_school
            ).select_related('student', 'assigned_class')
        else:
            student_classes = StudentClass.objects.all().select_related('student', 'assigned_class')
        
        generated_count = 0
        updated_count = 0
        
        with transaction.atomic():
            for student_class in student_classes:
                try:
                    # Check if report card already exists
                    report_card_query = ReportCard.objects.filter(
                        student=student_class.student,
                        academic_year=term.academic_year,
                        term=term,
                        class_assigned=student_class.assigned_class,
                    )
                    
                    if user_school:
                        report_card_query = report_card_query.filter(school=user_school)
                    
                    report_card = report_card_query.first()
                    
                    if not report_card:
                        # Create new report card
                        report_card = ReportCard.objects.create(
                            student=student_class.student,
                            academic_year=term.academic_year,
                            term=term,
                            class_assigned=student_class.assigned_class,
                            school=user_school,
                            generated_by=generated_by_user
                        )
                        generated_count += 1
                    else:
                        updated_count += 1
                    
                    # Auto-calculate scores, attendance, and positions
                    report_card.calculate_totals()
                    report_card.calculate_attendance()
                    report_card.calculate_position()
                    report_card.save()
                    
                except Exception as e:
                    print(f"Error processing student {student_class.student.full_name}: {str(e)}")
                    continue
        
        print(f"Auto-generated {generated_count} new report cards, updated {updated_count} existing ones")
        return generated_count, updated_count
        
    except Exception as e:
        print(f"Error in auto_generate_reports_for_term: {str(e)}")
        return 0, 0


def ensure_report_card_exists(student, academic_year, term, class_assigned, user_school, generated_by_user):
    """
    Ensure a report card exists for a specific student, term, and class.
    Creates it if it doesn't exist, otherwise returns the existing one.
    This is called lazily when a report card is actually needed.
    """
    try:
        # Check if report card already exists
        report_card_query = ReportCard.objects.filter(
            student=student,
            academic_year=academic_year,
            term=term,
            class_assigned=class_assigned,
        )
        
        if user_school:
            report_card_query = report_card_query.filter(school=user_school)
        
        report_card = report_card_query.first()
        
        if not report_card:
            # Create new report card
            report_card = ReportCard.objects.create(
                student=student,
                academic_year=academic_year,
                term=term,
                class_assigned=class_assigned,
                school=user_school,
                generated_by=generated_by_user
            )
            
            # Auto-calculate scores, attendance, and positions
            report_card.calculate_totals()
            report_card.calculate_attendance()
            report_card.calculate_position()
            report_card.save()
        
        return report_card
        
    except Exception as e:
        print(f"Error creating report card for {student.full_name}: {str(e)}")
        return None


@login_required
def auto_terminal_reports_dashboard(request):
    """
    Main dashboard for automatic terminal report card generation.
    Shows existing report cards without generating new ones.
    """
    user_school = get_user_school(request.user)
    
    # Get current academic year and term
    if user_school:
        current_academic_year = AcademicYear.objects.filter(
            is_current=True,
            school=user_school
        ).first()
        
        current_term = None
        if current_academic_year:
            current_term = Term.objects.filter(
                academic_year=current_academic_year,
                is_current=True,
                school=user_school
            ).first()
        
        # Get statistics
        stats = {
            'total_students': Student.objects.filter(school=user_school).count(),
            'total_classes': Class.objects.filter(school=user_school).count(),
            'total_terms': Term.objects.filter(school=user_school).count(),
            'generated_reports': ReportCard.objects.filter(school=user_school).count(),
        }
        
        # Get recent activity
        recent_reports = ReportCard.objects.filter(
            school=user_school
        ).select_related('student', 'term', 'class_assigned').order_by('-date_generated')[:10]
    else:
        current_academic_year = AcademicYear.objects.filter(
            is_current=True
        ).first()
        
        current_term = None
        if current_academic_year:
            current_term = Term.objects.filter(
                academic_year=current_academic_year,
                is_current=True
            ).first()
        
        # Get statistics
        stats = {
            'total_students': Student.objects.count(),
            'total_classes': Class.objects.count(),
            'total_terms': Term.objects.count(),
            'generated_reports': ReportCard.objects.count(),
        }
        
        # Get recent activity
        recent_reports = ReportCard.objects.select_related('student', 'term', 'class_assigned').order_by('-date_generated')[:10]
    
    context = {
        'current_academic_year': current_academic_year,
        'current_term': current_term,
        'stats': stats,
        'recent_reports': recent_reports,
        'user_school': user_school,
    }
    
    return render(request, 'terminal_reports/auto_dashboard.html', context)


@login_required
def auto_generate_terminal_reports(request):
    """
    Automatically generate terminal report cards based on student data.
    This eliminates the need for manual bulk generation.
    """
    user_school = get_user_school(request.user)
    
    if request.method == 'POST':
        # Get parameters from form
        academic_year_id = request.POST.get('academic_year')
        term_id = request.POST.get('term')
        class_id = request.POST.get('class')
        auto_calculate = request.POST.get('auto_calculate', 'true') == 'true'
        overwrite_existing = request.POST.get('overwrite_existing', 'false') == 'true'
        
        try:
            # Get objects
            academic_year = get_object_or_404(AcademicYear, id=academic_year_id, school=user_school)
            term = get_object_or_404(Term, id=term_id, school=user_school)
            
            # Get students to process
            if class_id:
                # Generate for specific class
                class_obj = get_object_or_404(Class, id=class_id, school=user_school)
                student_classes = StudentClass.objects.filter(
                    assigned_class=class_obj,
                    school=user_school
                ).select_related('student')
                students = [sc.student for sc in student_classes]
            else:
                # Generate for all students in the academic year
                student_classes = StudentClass.objects.filter(
                    school=user_school
                ).select_related('student', 'assigned_class')
                students = [sc.student for sc in student_classes]
            
            # Process each student
            generated_count = 0
            updated_count = 0
            errors = []
            
            with transaction.atomic():
                for student in students:
                    try:
                        # Get student's class assignment for this term
                        student_class = StudentClass.objects.filter(
                            student=student,
                            school=user_school
                        ).order_by('-date_assigned').first()
                        
                        if not student_class:
                            errors.append(f"No class assignment found for {student.full_name}")
                            continue
                        
                        # Check if report card already exists
                        report_card = ReportCard.objects.filter(
                            student=student,
                            academic_year=academic_year,
                            term=term,
                            school=user_school
                        ).first()
                        
                        if report_card and not overwrite_existing:
                            continue
                        
                        # Create or update report card
                        if not report_card:
                            report_card = ReportCard.objects.create(
                                student=student,
                                academic_year=academic_year,
                                term=term,
                                class_assigned=student_class.assigned_class,
                                school=user_school,
                                generated_by=request.user
                            )
                            generated_count += 1
                        else:
                            updated_count += 1
                        
                        # Auto-calculate if requested
                        if auto_calculate:
                            report_card.calculate_totals()
                            report_card.calculate_attendance()
                            report_card.calculate_position()
                            report_card.save()
                        
                    except Exception as e:
                        errors.append(f"Error processing {student.full_name}: {str(e)}")
                        continue
            
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                return JsonResponse({
                    'success': True,
                    'generated_count': generated_count,
                    'updated_count': updated_count,
                    'errors': errors,
                    'message': f"Generated {generated_count} new report cards and updated {updated_count} existing ones."
                })
            
            # Prepare response for regular form submission
            if errors:
                set_sweet_alert(
                    request,
                    "Generation Completed with Errors",
                    f"Generated: {generated_count}, Updated: {updated_count}. Errors: {len(errors)}",
                    "warning"
                )
            else:
                set_sweet_alert(
                    request,
                    "Generation Successful",
                    f"Generated: {generated_count}, Updated: {updated_count} report cards",
                    "success"
                )
            
            return redirect('auto_terminal_reports_list')
            
        except Exception as e:
            # Check if this is an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                return JsonResponse({
                    'success': False,
                    'message': f'An error occurred: {str(e)}'
                })
            
            set_sweet_alert(
                request,
                "Generation Failed",
                f"Error: {str(e)}",
                "error"
            )
            return redirect('auto_generate_terminal_reports')
    
    # GET request - show form
    academic_years = AcademicYear.objects.filter(
        is_archived=False,
        school=user_school
    ).order_by('-name')
    
    terms = Term.objects.filter(school=user_school).order_by('-term_number')
    classes = Class.objects.filter(school=user_school).order_by('name')
    
    context = {
        'academic_years': academic_years,
        'terms': terms,
        'classes': classes,
        'user_school': user_school,
    }
    
    return render(request, 'terminal_reports/auto_generate_form.html', context)


@login_required
def auto_terminal_reports_list(request):
    """
    List terminal report cards with advanced filtering and sorting options.
    Shows existing report cards without generating new ones.
    """
    user_school = get_user_school(request.user)
    
    # Get filter parameters
    academic_year_id = request.GET.get('academic_year')
    term_id = request.GET.get('term')
    class_id = request.GET.get('class')
    student_name = request.GET.get('student_name')
    status = request.GET.get('status')
    sort_by = request.GET.get('sort_by', 'student__full_name')
    sort_order = request.GET.get('sort_order', 'asc')
    
    # Clean up filter parameters - handle 'None' string and empty values
    if student_name in ['None', '', None]:
        student_name = None
    if academic_year_id in ['', None]:
        academic_year_id = None
    if term_id in ['', None]:
        term_id = None
    if class_id in ['', None]:
        class_id = None
    if status in ['', None]:
        status = None
    
    # Build query
    query = Q()
    
    # Apply school filter if user_school exists (skip for superadmins)
    if user_school:
        query &= Q(school=user_school)
    
    if academic_year_id:
        query &= Q(academic_year_id=academic_year_id)
    if term_id:
        query &= Q(term_id=term_id)
    if class_id:
        query &= Q(class_assigned_id=class_id)
    if student_name:
        query &= Q(student__full_name__icontains=student_name)
    if status:
        if status == 'approved':
            query &= Q(is_approved=True)
        elif status == 'pending':
            query &= Q(is_approved=False)
    
    # Apply sorting first
    if sort_order == 'desc':
        sort_by = f'-{sort_by}'
    else:
        sort_by = f'{sort_by}'
    
    # Get report cards with pagination limit for performance
    report_cards = ReportCard.objects.filter(query).select_related(
        'student', 'academic_year', 'term', 'class_assigned'
    ).order_by(sort_by)[:1000]  # Limit to 1000 results for performance
    
    # Pagination
    paginator = Paginator(report_cards, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter options
    if user_school:
        academic_years = AcademicYear.objects.filter(school=user_school).order_by('-name')
        terms = Term.objects.filter(school=user_school).order_by('-term_number')
        classes = Class.objects.filter(school=user_school).order_by('name')
    else:
        academic_years = AcademicYear.objects.all().order_by('-name')
        terms = Term.objects.all().order_by('-term_number')
        classes = Class.objects.all().order_by('name')
    
    # Get selected values for form
    selected_academic_year = None
    selected_term = None
    selected_class = None
    
    if academic_year_id:
        try:
            if user_school:
                selected_academic_year = AcademicYear.objects.get(id=academic_year_id, school=user_school)
            else:
                selected_academic_year = AcademicYear.objects.get(id=academic_year_id)
        except AcademicYear.DoesNotExist:
            pass
    
    if term_id:
        try:
            if user_school:
                selected_term = Term.objects.get(id=term_id, school=user_school)
            else:
                selected_term = Term.objects.get(id=term_id)
        except Term.DoesNotExist:
            pass
    
    if class_id:
        try:
            if user_school:
                selected_class = Class.objects.get(id=class_id, school=user_school)
            else:
                selected_class = Class.objects.get(id=class_id)
        except Class.DoesNotExist:
            pass
    
    context = {
        'page_obj': page_obj,
        'academic_years': academic_years,
        'terms': terms,
        'classes': classes,
        'selected_academic_year': selected_academic_year,
        'selected_term': selected_term,
        'selected_class': selected_class,
        'student_name': student_name,
        'status': status,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'user_school': user_school,
    }
    
    return render(request, 'terminal_reports/auto_reports_list.html', context)


@login_required
def auto_terminal_report_detail(request, report_card_id):
    """
    View detailed terminal report card with the same template as batch print.
    """
    user_school = get_user_school(request.user)
    
    # Fetch report card, applying school filter if user_school is present
    if user_school:
        report_card = get_object_or_404(
            ReportCard.objects.select_related(
                'student', 'academic_year', 'term', 'class_assigned'
            ),
            id=report_card_id,
            school=user_school
        )
    else:
        report_card = get_object_or_404(
            ReportCard.objects.select_related(
                'student', 'academic_year', 'term', 'class_assigned'
            ),
            id=report_card_id
        )
    
    # Get subject scores for this student in this term
    # First, get all class subjects for this class and academic year (only active ones)
    class_subjects = ClassSubject.objects.filter(
        class_name=report_card.class_assigned,
        academic_year=report_card.academic_year,
        is_active=True
    )
    
    # Then get assessments for these class subjects
    # Exclude mock exam assessments from report card calculations
    subject_scores_query = Assessment.objects.filter(
        student=report_card.student,
        class_subject__in=class_subjects,
        term=report_card.term,
    ).exclude(assessment_type='mock_exam').select_related("class_subject__subject", "class_subject")
    
    if user_school:
        subject_scores_query = subject_scores_query.filter(school=user_school)
    
    subject_scores = subject_scores_query.all()
    
    # Get teacher remarks for this student in this term
    teacher_remarks_query = StudentTermRemarks.objects.filter(
        student=report_card.student,
        term=report_card.term,
        academic_year=report_card.academic_year,
    )
    
    if user_school and hasattr(StudentTermRemarks, "school"):
        teacher_remarks_query = teacher_remarks_query.filter(school=user_school)
    
    teacher_remarks = teacher_remarks_query.first()
    
    # Get class size for student's class (including archived students)
    class_size_query = StudentClass.objects.filter(
        assigned_class=report_card.class_assigned,
    )
    
    if user_school:
        class_size_query = class_size_query.filter(school=user_school)
    
    class_size = class_size_query.count()
    
    # Get related data
    if user_school:
        school_info = user_school  # user_school is already a SchoolInformation object
    else:
        school_info = SchoolInformation.objects.filter(is_active=True).first()
    
    # Get grading system
    if user_school:
        grading_system = GradingSystem.objects.filter(school=user_school).first()
        grades = GradingSystem.objects.filter(school=user_school).order_by('-min_score')
    else:
        grading_system = GradingSystem.objects.filter(is_active=True).first()
        grades = GradingSystem.objects.filter(is_active=True).order_by('-min_score')
    
    # Get performance requirements
    if user_school:
        performance_requirements = PerformanceRequirement.objects.filter(school=user_school).first()
    else:
        performance_requirements = PerformanceRequirement.objects.filter(is_active=True).first()
    
    # Get scoring configuration for dynamic percentages
    scoring_config = None
    if user_school:
        from ..models import ScoringConfiguration
        scoring_config = ScoringConfiguration.get_active_config(user_school)
    
    # Get authority signatures
    if user_school:
        authority_signatures = SchoolAuthoritySignature.objects.filter(school=user_school)
    else:
        authority_signatures = SchoolAuthoritySignature.objects.filter(is_active=True)
    
    # Add the missing data to the report card object for template access
    report_card.subject_scores = subject_scores
    report_card.teacher_remarks = teacher_remarks
    report_card.class_size = class_size
    report_card.authority_signatures = authority_signatures
    
    context = {
        'report_cards': [report_card],  # Use same template structure as batch print
        'is_individual_view': True,
        'school_info': school_info,
        'grades': grades,
        'performance_requirements': performance_requirements,
        'scoring_config': scoring_config,
        'authority_signatures': authority_signatures,
        'user_school': user_school,
    }
    
    return render(request, 'reports/batch_print_report_cards.html', context)


@login_required
def auto_bulk_print_terminal_reports(request):
    """
    Bulk print terminal report cards with filtering options.
    """
    user_school = get_user_school(request.user)
    
    if request.method == 'POST':
        # Get selected report card IDs
        selected_ids = request.POST.getlist('selected_reports')
        
        if not selected_ids:
            set_sweet_alert(
                request,
                "No Reports Selected",
                "Please select at least one report card to print.",
                "warning"
            )
            return redirect('auto_terminal_reports_list')
        
        # Get report cards
        report_cards_query = ReportCard.objects.filter(
            id__in=selected_ids
        ).select_related(
            'student', 'academic_year', 'term', 'class_assigned'
        )
        
        if user_school:
            report_cards_query = report_cards_query.filter(school=user_school)
        
        report_cards = list(report_cards_query)
        
        if not report_cards:
            set_sweet_alert(
                request,
                "No Reports Found",
                "No report cards found for the selected criteria.",
                "error"
            )
            return redirect('auto_terminal_reports_list')
        
        # Get authority signatures once for all report cards
        if user_school:
            authority_signatures = SchoolAuthoritySignature.objects.filter(school=user_school)
        else:
            authority_signatures = SchoolAuthoritySignature.objects.filter(is_active=True)
        
        # Process each report card to add missing data
        for report_card in report_cards:
            # Get subject scores for this student in this term (only active ones)
            class_subjects = ClassSubject.objects.filter(
                class_name=report_card.class_assigned,
                academic_year=report_card.academic_year,
                is_active=True
            )
            
            # Exclude mock exam assessments from report card calculations
            subject_scores_query = Assessment.objects.filter(
                student=report_card.student,
                class_subject__in=class_subjects,
                term=report_card.term,
            ).exclude(assessment_type='mock_exam').select_related("class_subject__subject", "class_subject")
            
            if user_school:
                subject_scores_query = subject_scores_query.filter(school=user_school)
            
            subject_scores = subject_scores_query.all()
            
            # Get teacher remarks for this student in this term
            teacher_remarks_query = StudentTermRemarks.objects.filter(
                student=report_card.student,
                term=report_card.term,
                academic_year=report_card.academic_year,
            )
            
            if user_school and hasattr(StudentTermRemarks, "school"):
                teacher_remarks_query = teacher_remarks_query.filter(school=user_school)
            
            teacher_remarks = teacher_remarks_query.first()
            
            # Get class size for student's class
            class_size_query = StudentClass.objects.filter(
                assigned_class=report_card.class_assigned,
            )
            
            if user_school:
                class_size_query = class_size_query.filter(school=user_school)
            
            class_size = class_size_query.count()
            
            # Add the missing data to the report card object
            report_card.subject_scores = subject_scores
            report_card.teacher_remarks = teacher_remarks
            report_card.class_size = class_size
            report_card.authority_signatures = authority_signatures
        
        # Get related data
        if user_school:
            school_info = user_school  # user_school is already a SchoolInformation object
        else:
            school_info = SchoolInformation.objects.filter(is_active=True).first()
        
        # Get grading system
        if user_school:
            grading_system = GradingSystem.objects.filter(school=user_school).first()
            grades = GradingSystem.objects.filter(school=user_school).order_by('-min_score')
        else:
            grading_system = GradingSystem.objects.filter(is_active=True).first()
            grades = GradingSystem.objects.filter(is_active=True).order_by('-min_score')
        
        # Get performance requirements
        if user_school:
            performance_requirements = PerformanceRequirement.objects.filter(school=user_school).first()
        else:
            performance_requirements = PerformanceRequirement.objects.filter(is_active=True).first()
        
        # Get scoring configuration for dynamic percentages
        scoring_config = None
        if user_school:
            from ..models import ScoringConfiguration
            scoring_config = ScoringConfiguration.get_active_config(user_school)
        
        context = {
            'report_cards': report_cards,
            'is_individual_view': False,
            'school_info': school_info,
            'grades': grades,
            'performance_requirements': performance_requirements,
            'scoring_config': scoring_config,
            'authority_signatures': authority_signatures,
            'user_school': user_school,
        }
        
        return render(request, 'reports/batch_print_report_cards.html', context)
    
    # GET request - show selection form
    return redirect('auto_terminal_reports_list')


@login_required
def auto_terminal_reports_analytics(request):
    """
    Analytics dashboard for terminal report cards.
    """
    user_school = get_user_school(request.user)
    
    # Get filter parameters
    academic_year_id = request.GET.get('academic_year')
    term_id = request.GET.get('term')
    class_id = request.GET.get('class')
    
    # Build query
    query = Q(school=user_school)
    
    if academic_year_id:
        query &= Q(academic_year_id=academic_year_id)
    if term_id:
        query &= Q(term_id=term_id)
    if class_id:
        query &= Q(class_assigned_id=class_id)
    
    # Get analytics data
    report_cards = ReportCard.objects.filter(query)
    
    # Calculate statistics
    total_reports = report_cards.count()
    approved_reports = report_cards.filter(is_approved=True).count()
    pending_reports = report_cards.filter(is_approved=False).count()
    
    # Average scores by class
    class_stats = report_cards.values('class_assigned__name').annotate(
        avg_score=Avg('total_score'),
        count=Count('id')
    ).order_by('-avg_score')
    
    # Performance distribution
    performance_distribution = {
        'excellent': report_cards.filter(total_score__gte=80).count(),
        'good': report_cards.filter(total_score__gte=70, total_score__lt=80).count(),
        'average': report_cards.filter(total_score__gte=60, total_score__lt=70).count(),
        'below_average': report_cards.filter(total_score__lt=60).count(),
    }
    
    # Get filter options
    academic_years = AcademicYear.objects.filter(school=user_school).order_by('-name')
    terms = Term.objects.filter(school=user_school).order_by('-term_number')
    classes = Class.objects.filter(school=user_school).order_by('name')
    
    context = {
        'total_reports': total_reports,
        'approved_reports': approved_reports,
        'pending_reports': pending_reports,
        'class_stats': class_stats,
        'performance_distribution': performance_distribution,
        'academic_years': academic_years,
        'terms': terms,
        'classes': classes,
        'user_school': user_school,
    }
    
    return render(request, 'terminal_reports/analytics.html', context)


@login_required
def auto_terminal_reports_export(request):
    """
    Export terminal report cards data to various formats.
    """
    user_school = get_user_school(request.user)
    
    if request.method == 'POST':
        export_format = request.POST.get('export_format', 'csv')
        selected_ids = request.POST.getlist('selected_reports')
        
        if not selected_ids:
            set_sweet_alert(
                request,
                "No Reports Selected",
                "Please select at least one report card to export.",
                "warning"
            )
            return redirect('auto_terminal_reports_list')
        
        # Get report cards
        report_cards = ReportCard.objects.filter(
            id__in=selected_ids,
            school=user_school
        ).select_related(
            'student', 'academic_year', 'term', 'class_assigned'
        )
        
        if export_format == 'csv':
            # Generate CSV
            import csv
            from django.http import HttpResponse
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="terminal_reports.csv"'
            
            writer = csv.writer(response)
            writer.writerow([
                'Student Name', 'Admission Number', 'Class', 'Academic Year', 'Term',
                'Total Score', 'Position', 'Status', 'Date Generated'
            ])
            
            for report in report_cards:
                writer.writerow([
                    report.student.full_name,
                    report.student.admission_number,
                    report.class_assigned.name,
                    report.academic_year.name,
                    report.term.get_term_number_display(),
                    report.total_score,
                    report.position,
                    'Approved' if report.is_approved else 'Pending',
                    report.date_generated.strftime('%Y-%m-%d')
                ])
            
            return response
        
        elif export_format == 'excel':
            # Generate Excel (you may need to install openpyxl)
            try:
                import openpyxl
                from openpyxl import Workbook
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Terminal Reports"
                
                # Add headers
                headers = [
                    'Student Name', 'Admission Number', 'Class', 'Academic Year', 'Term',
                    'Total Score', 'Position', 'Status', 'Date Generated'
                ]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # Add data
                for row, report in enumerate(report_cards, 2):
                    ws.cell(row=row, column=1, value=report.student.full_name)
                    ws.cell(row=row, column=2, value=report.student.admission_number)
                    ws.cell(row=row, column=3, value=report.class_assigned.name)
                    ws.cell(row=row, column=4, value=report.academic_year.name)
                    ws.cell(row=row, column=5, value=report.term.get_term_number_display())
                    ws.cell(row=row, column=6, value=report.total_score)
                    ws.cell(row=row, column=7, value=report.position)
                    ws.cell(row=row, column=8, value='Approved' if report.is_approved else 'Pending')
                    ws.cell(row=row, column=9, value=report.date_generated.strftime('%Y-%m-%d'))
                
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="terminal_reports.xlsx"'
                
                wb.save(response)
                return response
                
            except ImportError:
                set_sweet_alert(
                    request,
                    "Export Error",
                    "Excel export requires openpyxl package. Please install it or use CSV export.",
                    "error"
                )
                return redirect('auto_terminal_reports_list')
    
    return redirect('auto_terminal_reports_list')


# API endpoints for AJAX requests
@login_required
def get_terms_for_auto_generation(request, academic_year_id):
    """
    Get terms for a specific academic year for auto generation form.
    """
    user_school = get_user_school(request.user)
    
    try:
        # Get academic year - superadmins can access any academic year
        if user_school:
            academic_year = get_object_or_404(AcademicYear, id=academic_year_id, school=user_school)
        else:
            # Superadmin - can access any academic year
            academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        
        # Get terms for this academic year
        terms_query = Term.objects.filter(academic_year=academic_year)
        if user_school:
            terms_query = terms_query.filter(school=user_school)
        
        terms = terms_query.order_by('term_number')
        
        terms_data = []
        for term in terms:
            terms_data.append({
                'id': term.id,
                'name': term.name,
                'display_name': term.get_term_number_display(),
                'start_date': term.start_date.strftime('%Y-%m-%d') if term.start_date else '',
                'end_date': term.end_date.strftime('%Y-%m-%d') if term.end_date else '',
            })
        
        return JsonResponse({'success': True, 'terms': terms_data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
def get_classes_for_auto_generation(request, academic_year_id):
    """
    Get classes for a specific academic year for auto generation form.
    """
    user_school = get_user_school(request.user)
    
    try:
        # Get academic year - superadmins can access any academic year
        if user_school:
            academic_year = get_object_or_404(AcademicYear, id=academic_year_id, school=user_school)
        else:
            # Superadmin - can access any academic year
            academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        
        # Get classes that have subjects in this academic year (only active ones)
        class_subjects_query = ClassSubject.objects.filter(academic_year=academic_year, is_active=True)
        if user_school:
            class_subjects_query = class_subjects_query.filter(school=user_school)
        
        class_subjects = class_subjects_query.values_list('class_name_id', flat=True).distinct()
        
        classes_query = Class.objects.filter(id__in=class_subjects)
        if user_school:
            classes_query = classes_query.filter(school=user_school)
        
        classes = classes_query.order_by('name')
        
        classes_data = []
        for class_obj in classes:
            classes_data.append({
                'id': class_obj.id,
                'name': class_obj.name,
                'form': class_obj.form.name if class_obj.form else '',
            })
        
        return JsonResponse({'success': True, 'classes': classes_data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
def recalculate_report_card(request, report_card_id):
    """
    Recalculate a specific report card's scores, positions, and attendance.
    """
    user_school = get_user_school(request.user)
    
    try:
        report_card = get_object_or_404(
            ReportCard,
            id=report_card_id,
            school=user_school
        )
        
        # Recalculate all values
        report_card.calculate_totals()
        report_card.calculate_attendance()
        report_card.calculate_position()
        report_card.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Report card recalculated successfully',
            'data': {
                'total_score': float(report_card.total_score or 0),
                'position': report_card.position,
                'days_present': report_card.days_present,
                'total_school_days': report_card.total_school_days,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error recalculating report card: {str(e)}'
        }, status=400)
