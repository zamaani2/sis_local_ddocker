from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Prefetch, Avg, Max, Min

from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods

from django.utils import timezone
from datetime import datetime, timedelta
import numpy as np
import traceback

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO


from shs_system.models import (
    Teacher,
    Class,
    Student,
    StudentClass,
    StudentTermRemarks,
    AcademicYear,
    Term,
    SchoolInformation,
    AttendanceRecord,
    ClassSubject,
    Assessment,
    ClassTeacher,
    ScoringConfiguration,
)
from shs_system.forms import StudentTermRemarksForm

# Set to True to enable detailed debug logging
DEBUG_MODE = True


def debug_log(message):
    """Helper function for debug logging"""
    if DEBUG_MODE:
        print(f"[DEBUG] {message}")


# Helper function to get user's school
def get_user_school(user):
    """
    Get the school associated with the current user.
    For superadmins, this will return None (they can access all schools).
    For other users, it will return their associated school.
    """
    if user.is_superadmin:
        return None  # Superadmins can access all schools
    return user.school


# Utility functions
def get_current_academic_data(user_school=None):
    """
    Get current academic year and term for a specific school or globally

    Args:
        user_school: Optional school to filter by

    Returns:
        tuple: (current_academic_year, current_term)
    """
    if user_school:
        # Get school-specific academic year and term
        current_academic_year = AcademicYear.objects.filter(
            is_current=True, school=user_school
        ).first()

        if current_academic_year:
            current_term = Term.objects.filter(
                is_current=True, academic_year=current_academic_year, school=user_school
            ).first()
        else:
            current_term = None
    else:
        # Get any current academic year and term (for superadmins)
        current_academic_year = AcademicYear.objects.filter(is_current=True).first()
        current_term = Term.objects.filter(is_current=True).first()

    return current_academic_year, current_term


def calculate_weekdays(start_date, end_date):
    """Calculate number of weekdays between two dates"""
    # Convert to numpy datetime for faster calculation
    start = np.datetime64(start_date)
    end = np.datetime64(end_date)
    # Count business days between start and end
    days = np.busday_count(start, end)
    return int(days)


def calculate_student_attendance(student, start_date, end_date, school=None):
    """
    Calculate attendance for a student within a date range

    Args:
        student: Student object
        start_date: Start date for attendance range
        end_date: End date for attendance range
        school: Optional school to filter by

    Returns:
        int: Number of days present
    """
    query = AttendanceRecord.objects.filter(
        student=student,
        date__range=[start_date, end_date],
        is_present=True,
    )

    # Apply school filter for multi-tenancy
    if school:
        query = query.filter(school=school)

    return query.count()


def update_student_remarks(
    remarks, form_data, student=None, term=None, prefix="", school=None
):
    """
    Update student remarks from form data

    Args:
        remarks: StudentTermRemarks object to update
        form_data: Form data with remarks information
        student: Optional Student object
        term: Optional Term object
        prefix: Form field prefix
        school: Optional school to filter by
    """
    try:
        # Track if any fields were updated
        fields_updated = []

        # Update remarks text - check both naming patterns
        interest = form_data.get(
            f"{prefix}interest_remarks", form_data.get(f"{prefix}interest", "")
        )
        if remarks.interest_remarks != interest:
            remarks.interest_remarks = interest
            fields_updated.append("interest_remarks")

        conduct = form_data.get(
            f"{prefix}conduct_remarks", form_data.get(f"{prefix}conduct", "")
        )
        if remarks.conduct_remarks != conduct:
            remarks.conduct_remarks = conduct
            fields_updated.append("conduct_remarks")

        attitude = form_data.get(
            f"{prefix}attitude_remarks", form_data.get(f"{prefix}attitude", "")
        )
        if remarks.attitude_remarks != attitude:
            remarks.attitude_remarks = attitude
            fields_updated.append("attitude_remarks")

        # For general remarks, check multiple possible field names
        general_remarks = form_data.get(
            f"{prefix}general_remarks",
            form_data.get(f"{prefix}general", form_data.get(f"{prefix}remarks", "")),
        )

        if remarks.general_remarks != general_remarks:
            remarks.general_remarks = general_remarks
            fields_updated.append("general_remarks")

        # Handle attendance
        auto_calculate_key = f"{prefix}auto_calculate_attendance"
        if prefix:  # for batch updates the key is different
            auto_calculate_key = f"{prefix}_auto_calculate"
            # Also check alternative naming
            if auto_calculate_key not in form_data:
                auto_calculate_key = f"{prefix}auto_calculate"

        auto_calculate = form_data.get(auto_calculate_key, "true").lower() == "true"
        if remarks.auto_calculate_attendance != auto_calculate:
            remarks.auto_calculate_attendance = auto_calculate
            fields_updated.append("auto_calculate_attendance")

        if auto_calculate and student and term:
            # Calculate attendance from records with school context
            days_present = calculate_student_attendance(
                student, term.start_date, term.end_date, school
            )
            if remarks.days_present != days_present:
                remarks.days_present = days_present
                fields_updated.append("days_present")
        else:
            # Use manually entered attendance
            days_present_key = (
                f"{prefix}days_present" if not prefix else f"{prefix}_days_present"
            )
            # Check alternative naming if the first pattern isn't found
            if days_present_key not in form_data and prefix:
                days_present_key = f"{prefix}days_present"

            days_absent_key = (
                f"{prefix}days_absent" if not prefix else f"{prefix}_days_absent"
            )
            # Check alternative naming if the first pattern isn't found
            if days_absent_key not in form_data and prefix:
                days_absent_key = f"{prefix}days_absent"

            total_days_key = (
                "total_school_days" if not prefix else "total_attendance_days"
            )
            # Also check for class-specific total days
            if total_days_key not in form_data:
                total_days_key = f"{prefix}total_school_days"

            # Safely convert values to integers
            try:
                days_present = int(
                    form_data.get(days_present_key, remarks.days_present)
                )
                if remarks.days_present != days_present:
                    remarks.days_present = days_present
                    fields_updated.append("days_present")
            except (ValueError, TypeError):
                # Keep existing value if conversion fails
                pass

            try:
                days_absent = int(form_data.get(days_absent_key, remarks.days_absent))
                if remarks.days_absent != days_absent:
                    remarks.days_absent = days_absent
                    fields_updated.append("days_absent")
            except (ValueError, TypeError):
                # Keep existing value if conversion fails
                pass

            try:
                # If no specific value is provided, don't override the existing value
                total_days = form_data.get(total_days_key)
                if total_days and remarks.total_school_days != int(total_days):
                    remarks.total_school_days = int(total_days)
                    fields_updated.append("total_school_days")
            except (ValueError, TypeError):
                # Keep the existing value if conversion fails
                pass

        # Only save if something actually changed
        if fields_updated:
            print(f"Updating fields for student remarks: {fields_updated}")
            remarks.save(update_fields=fields_updated if fields_updated else None)
        else:
            print("No fields updated for student remarks")

        return remarks

    except Exception as e:
        print(f"Error updating student remarks: {str(e)}")
        import traceback

        print(traceback.format_exc())
        # Return the original remarks object without changes
        return remarks


def get_attendance_percentage(days_present, total_days):
    """Calculate attendance percentage"""
    if total_days > 0:
        return round((days_present / total_days) * 100, 1)
    return 0


def verify_teacher_class_access(
    teacher, class_obj, redirect_url="class_teacher_remarks", school=None
):
    """
    Verify teacher has access to the class

    Args:
        teacher: Teacher object
        class_obj: Class object
        redirect_url: URL to redirect to if access is denied
        school: Optional school to filter by

    Returns:
        bool: True if teacher has access, False otherwise
    """
    # Build query with school context
    query = ClassTeacher.objects.filter(
        teacher=teacher, class_assigned=class_obj, is_active=True
    )

    # Apply school filter for multi-tenancy
    if school:
        query = query.filter(school=school)

    return query.exists()


@login_required
def class_teacher_remarks(request):
    # Get the logged-in teacher
    teacher = get_object_or_404(Teacher, user=request.user)

    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    # Get current academic year and term for user's school
    current_academic_year, current_term = get_current_academic_data(user_school)

    if not current_academic_year or not current_term:
        messages.error(request, "No active academic year or term found.")
        return redirect("teacher_dashboard")

    # Get assigned classes for the teacher using the ClassTeacher model with school context
    class_teacher_query = ClassTeacher.objects.filter(
        teacher=teacher, academic_year=current_academic_year, is_active=True
    )

    # Apply school filter for multi-tenancy
    if user_school:
        class_teacher_query = class_teacher_query.filter(school=user_school)

    class_teacher_assignments = class_teacher_query.select_related("class_assigned")

    assigned_classes = [ct.class_assigned for ct in class_teacher_assignments]

    if not assigned_classes:
        messages.warning(
            request, "You are not assigned as a class teacher to any class."
        )
        return redirect("teacher_dashboard")

    # Get selected class or default to first assigned class
    selected_class_id = request.GET.get("class_id")
    if selected_class_id:
        # First check if this class is among the teacher's assigned classes
        selected_class = None
        for class_obj in assigned_classes:
            if class_obj.class_id == selected_class_id:
                selected_class = class_obj
                break

        if not selected_class:
            messages.error(request, "You are not authorized to view this class.")
            return redirect("class_teacher_remarks")
    else:
        selected_class = assigned_classes[0]

    # Calculate total school days for the term
    term_start_date = current_term.start_date
    term_end_date = current_term.end_date
    total_school_days = calculate_weekdays(term_start_date, term_end_date)

    # Get students in the selected class with school context
    student_classes_query = StudentClass.objects.filter(
        assigned_class=selected_class, is_active=True
    )

    # Apply school filter for multi-tenancy
    if user_school:
        student_classes_query = student_classes_query.filter(school=user_school)

    student_classes = student_classes_query.select_related("student")

    students = [sc.student for sc in student_classes]

    # Get or create remarks for each student
    students_with_remarks = []
    for student in students:
        try:
            # Get or create remarks query with school context
            remarks_query = StudentTermRemarks.objects.filter(
                student=student,
                academic_year=current_academic_year,
                term=current_term,
                class_assigned=selected_class,
            )

            # Apply school filter for multi-tenancy if student has school association
            if student.school:
                remarks_query = remarks_query.filter(student__school=student.school)

            remarks = remarks_query.first()

            if not remarks:
                # Create new remarks with proper school association
                remarks = StudentTermRemarks.objects.create(
                    student=student,
                    academic_year=current_academic_year,
                    term=current_term,
                    class_assigned=selected_class,
                    class_teacher=teacher,
                    total_school_days=total_school_days,
                    auto_calculate_attendance=True,
                )
                created = True
            else:
                created = False

            # Ensure class_teacher is set for existing records
            if not remarks.class_teacher:
                remarks.class_teacher = teacher
                remarks.save(update_fields=["class_teacher"])

            # Update total school days if needed
            if remarks.total_school_days != total_school_days:
                remarks.total_school_days = total_school_days
                remarks.save(update_fields=["total_school_days"])

            # Calculate attendance if auto-calculate is enabled
            if remarks.auto_calculate_attendance:
                remarks.days_present = calculate_student_attendance(
                    student, term_start_date, term_end_date, user_school
                )
                remarks.save(update_fields=["days_present"])

            form = StudentTermRemarksForm(instance=remarks)
            students_with_remarks.append(
                {"student": student, "remarks": remarks, "form": form}
            )
        except Exception as e:
            # Log error but continue with other students
            print(f"Error processing student {student.id}: {str(e)}")
            continue

    context = {
        "assigned_classes": assigned_classes,
        "selected_class": selected_class,
        "students_with_remarks": students_with_remarks,
        "current_academic_year": current_academic_year,
        "current_term": current_term,
        "total_school_days": total_school_days,
        "term_start_date": term_start_date,
        "term_end_date": term_end_date,
        "user_school": user_school,  # Add user's school to context
    }

    return render(request, "teacher/class_teacher_remarks.html", context)


@login_required
@require_POST
def save_student_remarks_ajax(request):
    try:
        debug_log(
            f"save_student_remarks_ajax called - POST keys: {list(request.POST.keys())}"
        )

        student_id = request.POST.get("student_id")
        class_id = request.POST.get("class_id")

        if not student_id or not class_id:
            debug_log("Missing required parameters")
            return JsonResponse(
                {"status": "error", "message": "Missing required parameters"},
                status=400,
            )

        # Get user's school for multi-tenancy
        user_school = get_user_school(request.user)

        # Get student with school context
        student_query = Student.objects.filter(id=student_id)
        if user_school:
            student_query = student_query.filter(school=user_school)
        student = get_object_or_404(student_query)

        # Get class with school context
        class_query = Class.objects.filter(class_id=class_id)
        if user_school:
            class_query = class_query.filter(school=user_school)
        class_obj = get_object_or_404(class_query)

        # Get teacher with school context
        teacher_query = Teacher.objects.filter(user=request.user)
        if user_school:
            teacher_query = teacher_query.filter(school=user_school)
        teacher = get_object_or_404(teacher_query)

        debug_log(
            f"Processing student {student.id} - {student.full_name} for class {class_obj.name}"
        )

        # Verify teacher has access to this class with school context
        if not verify_teacher_class_access(teacher, class_obj, school=user_school):
            debug_log("Teacher not authorized for this class")
            return JsonResponse(
                {
                    "status": "error",
                    "message": "You are not authorized to update this class",
                },
                status=403,
            )

        # Get current academic year and term for user's school
        current_academic_year, current_term = get_current_academic_data(user_school)

        if not current_academic_year or not current_term:
            debug_log("No active academic year or term found")
            return JsonResponse(
                {"status": "error", "message": "No active academic year or term found"},
                status=400,
            )

        # Get or create remarks with proper defaults and school context
        remarks_query = StudentTermRemarks.objects.filter(
            student=student,
            academic_year=current_academic_year,
            term=current_term,
            class_assigned=class_obj,
        )

        if user_school:
            remarks_query = remarks_query.filter(
                student__school=user_school,
                academic_year__school=user_school,
                term__school=user_school,
                class_assigned__school=user_school,
            )

        remarks = remarks_query.first()

        if not remarks:
            # Create new remarks with proper school association
            remarks = StudentTermRemarks.objects.create(
                student=student,
                academic_year=current_academic_year,
                term=current_term,
                class_assigned=class_obj,
                class_teacher=teacher,
            )
            created = True
        else:
            created = False

        debug_log(
            f"Remarks record {'created' if created else 'retrieved'} (ID: {remarks.id})"
        )

        # Ensure class_teacher is set
        if not remarks.class_teacher:
            debug_log("Setting class_teacher on existing record")
            remarks.class_teacher = teacher

        # Update remarks from form data with school context
        debug_log("Updating remarks from form data")
        remarks = update_student_remarks(
            remarks, request.POST, student, current_term, school=user_school
        )

        # Calculate attendance percentage
        attendance_percentage = get_attendance_percentage(
            remarks.days_present, remarks.total_school_days
        )

        debug_log(
            f"Update complete - days_present: {remarks.days_present}, percentage: {attendance_percentage}%"
        )

        return JsonResponse(
            {
                "status": "success",
                "message": f"Remarks saved for {student.full_name}",
                "attendance_percentage": attendance_percentage,
                "days_present": remarks.days_present,
            }
        )

    except Exception as e:
        error_trace = traceback.format_exc()
        debug_log(f"Error in save_student_remarks_ajax: {str(e)}")
        debug_log(error_trace)
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@login_required
@require_POST
def save_all_remarks_ajax(request):
    try:
        debug_log(
            f"save_all_remarks_ajax called - POST keys: {list(request.POST.keys())}"
        )

        class_id = request.POST.get("class_id")
        if not class_id:
            debug_log("Missing class ID")
            return JsonResponse(
                {"status": "error", "message": "Missing class ID"}, status=400
            )

        # Get user's school for multi-tenancy
        user_school = get_user_school(request.user)

        # Get class with school context
        class_query = Class.objects.filter(class_id=class_id)
        if user_school:
            class_query = class_query.filter(school=user_school)
        class_obj = get_object_or_404(class_query)

        # Get teacher with school context
        teacher_query = Teacher.objects.filter(user=request.user)
        if user_school:
            teacher_query = teacher_query.filter(school=user_school)
        teacher = get_object_or_404(teacher_query)

        debug_log(
            f"Processing bulk update for class {class_obj.name} by teacher {teacher.full_name}"
        )

        # Verify teacher has access to this class with school context
        if not verify_teacher_class_access(teacher, class_obj, school=user_school):
            debug_log("Teacher not authorized for this class")
            return JsonResponse(
                {
                    "status": "error",
                    "message": "You are not authorized to update this class",
                },
                status=403,
            )

        # Get current academic year and term for user's school
        current_academic_year, current_term = get_current_academic_data(user_school)

        if not current_academic_year or not current_term:
            debug_log("No active academic year or term found")
            return JsonResponse(
                {"status": "error", "message": "No active academic year or term found"},
                status=400,
            )

        # Get all students in the class with school context
        student_classes_query = StudentClass.objects.filter(
            assigned_class=class_obj, is_active=True
        )

        # Apply school filter for multi-tenancy
        if user_school:
            student_classes_query = student_classes_query.filter(school=user_school)

        student_classes = student_classes_query.select_related("student")

        # Get the total attendance days entered
        total_school_days = int(request.POST.get("total_attendance_days", 0))

        # Debug information
        debug_log(
            f"Processing {len(student_classes)} students with total_school_days={total_school_days}"
        )
        debug_log(f"Form data keys: {list(request.POST.keys())}")

        updated_count = 0
        error_count = 0
        skipped_count = 0
        for student_class in student_classes:
            student = student_class.student
            try:
                # Get or create remarks with school context
                remarks_query = StudentTermRemarks.objects.filter(
                    student=student,
                    academic_year=current_academic_year,
                    term=current_term,
                    class_assigned=class_obj,
                )

                # Apply school filter for multi-tenancy
                if user_school:
                    remarks_query = remarks_query.filter(
                        student__school=user_school,
                        academic_year__school=user_school,
                        term__school=user_school,
                        class_assigned__school=user_school,
                    )

                remarks = remarks_query.first()

                if not remarks:
                    # Create new remarks with proper school association
                    remarks = StudentTermRemarks.objects.create(
                        student=student,
                        academic_year=current_academic_year,
                        term=current_term,
                        class_assigned=class_obj,
                        class_teacher=teacher,
                        total_school_days=total_school_days,
                    )
                    created = True
                else:
                    created = False

                debug_log(
                    f"Student {student.id} - {student.full_name}: Record {'created' if created else 'retrieved'}"
                )

                # Set class_teacher explicitly if it wasn't set (to handle any existing records)
                if not remarks.class_teacher:
                    remarks.class_teacher = teacher

                # Set total school days
                if remarks.total_school_days != total_school_days:
                    remarks.total_school_days = total_school_days

                # Important: Save these changes before proceeding
                remarks.save(update_fields=["class_teacher", "total_school_days"])

                prefix = f"student_{student.id}_"

                # Debug for this specific student
                student_fields = [
                    k for k in request.POST.keys() if k.startswith(prefix)
                ]
                debug_log(
                    f"Student {student.id} fields found in form: {student_fields}"
                )

                if not student_fields:
                    debug_log(f"No fields found for student {student.id} - skipping")
                    skipped_count += 1
                    continue

                # Now update the remarks directly from the form data with school context
                has_updates = False

                # Get interest remarks
                interest_field = f"{prefix}interest_remarks"
                if interest_field in request.POST:
                    interest_value = request.POST.get(interest_field, "")
                    if remarks.interest_remarks != interest_value:
                        remarks.interest_remarks = interest_value
                        has_updates = True
                        debug_log(f"Updated interest remarks to: {interest_value}")

                # Get conduct remarks
                conduct_field = f"{prefix}conduct_remarks"
                if conduct_field in request.POST:
                    conduct_value = request.POST.get(conduct_field, "")
                    if remarks.conduct_remarks != conduct_value:
                        remarks.conduct_remarks = conduct_value
                        has_updates = True
                        debug_log(f"Updated conduct remarks to: {conduct_value}")

                # Get attitude remarks
                attitude_field = f"{prefix}attitude_remarks"
                if attitude_field in request.POST:
                    attitude_value = request.POST.get(attitude_field, "")
                    if remarks.attitude_remarks != attitude_value:
                        remarks.attitude_remarks = attitude_value
                        has_updates = True
                        debug_log(f"Updated attitude remarks to: {attitude_value}")

                # Get general remarks
                general_field = f"{prefix}general_remarks"
                if general_field in request.POST:
                    general_value = request.POST.get(general_field, "")
                    if remarks.general_remarks != general_value:
                        remarks.general_remarks = general_value
                        has_updates = True
                        debug_log(f"Updated general remarks to: {general_value}")

                # Get attendance data
                days_present_field = f"{prefix}days_present"
                if days_present_field in request.POST:
                    try:
                        days_present_value = int(
                            request.POST.get(days_present_field, 0)
                        )
                        if remarks.days_present != days_present_value:
                            remarks.days_present = days_present_value
                            has_updates = True
                            debug_log(f"Updated days present to: {days_present_value}")
                    except (ValueError, TypeError) as e:
                        debug_log(f"Error converting days present: {e}")

                # Calculate days absent based on total school days
                if (
                    has_updates
                    and hasattr(remarks, "days_present")
                    and remarks.total_school_days
                ):
                    remarks.days_absent = max(
                        0, remarks.total_school_days - remarks.days_present
                    )
                    debug_log(f"Updated days absent to: {remarks.days_absent}")

                # Auto-calculate setting
                auto_calc_field = f"{prefix}auto_calculate"
                if auto_calc_field in request.POST:
                    auto_calc_value = (
                        request.POST.get(auto_calc_field, "false").lower() == "true"
                    )
                    if remarks.auto_calculate_attendance != auto_calc_value:
                        remarks.auto_calculate_attendance = auto_calc_value
                        has_updates = True
                        debug_log(f"Updated auto calculate to: {auto_calc_value}")

                        # If auto-calculate is enabled, recalculate attendance with school context
                        if auto_calc_value:
                            days_present = calculate_student_attendance(
                                student,
                                current_term.start_date,
                                current_term.end_date,
                                user_school,
                            )
                            if remarks.days_present != days_present:
                                remarks.days_present = days_present
                                has_updates = True
                                debug_log(
                                    f"Auto-calculated days present to: {days_present}"
                                )

                if has_updates:
                    # Save the updated remarks
                    remarks.save()
                    debug_log(f"Saved updates for student {student.id}")
                    updated_count += 1
                else:
                    debug_log(f"No changes detected for student {student.id}")
                    skipped_count += 1

            except Exception as e:
                # Log individual student errors but continue with the rest
                error_trace = traceback.format_exc()
                debug_log(f"Error updating remarks for student {student.id}: {str(e)}")
                debug_log(error_trace)
                error_count += 1
                continue

        debug_log(
            f"Bulk update complete: {updated_count} updated, {error_count} errors, {skipped_count} skipped"
        )

        if updated_count > 0:
            return JsonResponse(
                {
                    "status": "success",
                    "message": f"Remarks saved for {updated_count} students",
                    "updated_count": updated_count,
                    "error_count": error_count,
                    "skipped_count": skipped_count,
                }
            )
        else:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "No student remarks were updated",
                    "error_count": error_count,
                },
                status=400,
            )

    except Exception as e:
        # Log the full exception for debugging
        error_trace = traceback.format_exc()
        debug_log(f"Error in save_all_remarks_ajax: {str(e)}")
        debug_log(error_trace)
        return JsonResponse({"status": "error", "message": str(e)}, status=500)


@login_required

def export_remarks(request):
    """Export student remarks to Excel file with dropdown lists for remarks."""
    try:
        class_id = request.GET.get("class_id")
        if not class_id:
            messages.error(request, "No class selected for export.")
            return redirect("class_teacher_remarks")

        # Get user's school for multi-tenancy
        user_school = get_user_school(request.user)

        # Get class with school context
        class_query = Class.objects.filter(class_id=class_id)
        if user_school:
            class_query = class_query.filter(school=user_school)
        class_obj = get_object_or_404(class_query)

        # Get teacher with school context
        teacher_query = Teacher.objects.filter(user=request.user)
        if user_school:
            teacher_query = teacher_query.filter(school=user_school)
        teacher = get_object_or_404(teacher_query)

        # Verify teacher has access to this class
        if not verify_teacher_class_access(teacher, class_obj, school=user_school):
            messages.error(request, "You are not authorized to export remarks for this class.")
            return redirect("class_teacher_remarks")

        # Get current academic year and term
        current_academic_year, current_term = get_current_academic_data(user_school)
        if not current_academic_year or not current_term:
            messages.error(request, "No active academic year or term found.")
            return redirect("class_teacher_remarks")

        # Get students in the class
        student_classes_query = StudentClass.objects.filter(
            assigned_class=class_obj, is_active=True
        )
        if user_school:
            student_classes_query = student_classes_query.filter(school=user_school)
        student_classes = student_classes_query.select_related("student")
        students = [sc.student for sc in student_classes]

        # Define all available remark options
        general_remarks_options = [
            "",
            "More room for improvement",
            "Could do better than this",
            "Needs extra tuition at home",
            "Good performance, keep it up",
            "Keep it up",
            "Must sit up",
            "Advised to be repeated",
            "Must work hard",
            "It would be better for him/ her to be repeated",
            "Good work done, Keep it up",
            "Has improved",
            "Hard Working",
            "Must Buck up",
        ]

        interest_remarks_options = [
            "",
            "Has a well-developed sense of humor",
            "Holds many varied interests",
            "Has a keen interest that has been shared with the class",
            "Displays and talks about personal related items from home",
            "Provides background knowledge about topics of particular interest",
            "Impressive understanding and depth of knowledge about his interests",
            "Seeks additional information independently about classroom topics",
            "Reads extensively for enjoyment",
            "Frequently discusses concepts about which they have read",
            "Is a gifted performer",
            "Is a talented artist",
            "Has a flair for dramatic reading and acting",
            "Enjoys sharing their musical talent with the class",
            "Running",
            "Reading",
            "Drumming",
            "Football",
            "Volleyball",
            "Singing",
            "Dancing",
            "Games",
            "In-door games",
        ]

        conduct_remarks_options = [
            "",
            "Shows respect for teachers and peers",
            "Treats school property and the belongings of others with care and respect",
            "Honest and trustworthy in dealing with others",
            "Displays good citizenship by assisting other students",
            "Joins in school community projects",
            "Concerned about the feelings of peers",
            "Faithfully performs classroom tasks",
            "Can be depended on to do what they are asked to do",
            "Respectful",
            "Well-Behaved",
            "Disrespectful",
            "Calm",
            "Humble",
            "Approachable (Sociable)",
            "Bully",
            "Truant",
            "Satisfactory",
            "Aggressive",
        ]

        attitude_remarks_options = [
            "",
            "An enthusiastic learner who seems to enjoy school",
            "Exhibits a positive outlook and attitude in the classroom",
            "Appears well rested and ready for each day's activities",
            "Shows enthusiasm for classroom activities",
            "Shows initiative and looks for new ways to get involved",
            "Uses instincts to deal with matters independently and in a positive way",
            "Strives to reach their full potential",
            "Needs to work on his talking in class",
            "Has been disrupting the class lately",
            "Needs to improve him/ her attention span",
            "Is steadily improving",
            "Needs to pay more attention",
            "Hard Working",
            "Dependable",
            "Not Serious in class",
            "Lazy",
            "Aggressive",
            "Calm",
            "Slow in learning",
        ]

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Clean sheet name
        sheet_name = f"{class_obj.name}_Remarks"
        invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, "_")
        ws.title = sheet_name[:31]

        # Define styles
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Get school information
        if user_school:
            school_name = user_school.name
        else:
            active_school = SchoolInformation.get_active()
            school_name = active_school.name if active_school else "SCHOOL NAME"

        # Write header information
        ws["A1"] = f"School Name: {school_name}"
        ws["A2"] = f"Class: {class_obj.name}"
        ws["A3"] = f"Term: {current_term.name}"
        ws["A4"] = f"Academic Year: {current_academic_year.name}"
        ws["A5"] = f"Export Date: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Make headers bold
        for row in range(1, 6):
            ws.cell(row=row, column=1).font = Font(bold=True)

        # Define column headers
        headers = [
            "Admission Number",
            "Student Name",
            "Days Present",
            "Total School Days",
            "Attendance %",
            "General Remarks",
            "Interest Remarks",
            "Conduct Remarks",
            "Attitude Remarks",
        ]

        # Apply headers to row 7
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add instructions
        instruction_text = "Instructions: 1) If dropdown arrows appear, use them to select remarks. 2) Otherwise, go to 'Options' sheet, copy a remark, and paste into remarks columns. Do NOT modify Admission Number or Student Name."
        ws["A6"] = instruction_text
        ws.merge_cells("A6:I6")
        ws["A6"].font = Font(bold=True, color="FF0000")
        ws["A6"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[6].height = 40  # Make row taller for wrapped text

        # Add data rows
        row_num = 8
        max_row = 8  # Track the last row for data validation
        for student in students:
            # Get remarks for this student
            remarks_query = StudentTermRemarks.objects.filter(
                student=student,
                academic_year=current_academic_year,
                term=current_term,
                class_assigned=class_obj,
            )
            if user_school:
                remarks_query = remarks_query.filter(student__school=user_school)
            remarks = remarks_query.first()

            # Write student data
            ws.cell(row=row_num, column=1, value=student.admission_number).border = border
            ws.cell(row=row_num, column=2, value=student.full_name).border = border

            if remarks:
                ws.cell(row=row_num, column=3, value=remarks.days_present or 0).border = border
                ws.cell(row=row_num, column=4, value=remarks.total_school_days or 0).border = border
                
                # Calculate attendance percentage
                attendance_pct = get_attendance_percentage(
                    remarks.days_present or 0, remarks.total_school_days or 0
                )
                ws.cell(row=row_num, column=5, value=attendance_pct).border = border
                ws.cell(row=row_num, column=5).number_format = "0.0"
                
                ws.cell(row=row_num, column=6, value=remarks.general_remarks or "").border = border
                ws.cell(row=row_num, column=7, value=remarks.interest_remarks or "").border = border
                ws.cell(row=row_num, column=8, value=remarks.conduct_remarks or "").border = border
                ws.cell(row=row_num, column=9, value=remarks.attitude_remarks or "").border = border
            else:
                # Empty values if no remarks exist
                for col in range(3, 10):
                    ws.cell(row=row_num, column=col, value="").border = border

            row_num += 1
            max_row = row_num - 1

        # Create a helper sheet for dropdown options
        # Use a simple sheet name without spaces for better compatibility
        helper_sheet_name = "Options"
        helper_sheet = wb.create_sheet(helper_sheet_name)
        # Ensure helper sheet is created before main sheet data validation
        # Keep it visible (don't hide it) for Excel 2019 compatibility
        
        # Write options to helper sheet
        # Filter out empty strings and ensure all options are valid
        # Column A: General Remarks
        helper_sheet["A1"] = "General Remarks"
        row_idx = 2
        for option in general_remarks_options:
            if option:  # Only add non-empty options
                helper_sheet[f"A{row_idx}"] = str(option).strip()
                row_idx += 1
        # Add empty option at the end for clearing
        helper_sheet[f"A{row_idx}"] = ""
        general_last_row = row_idx
        
        # Column B: Interest Remarks
        helper_sheet["B1"] = "Interest Remarks"
        row_idx = 2
        for option in interest_remarks_options:
            if option:  # Only add non-empty options
                helper_sheet[f"B{row_idx}"] = str(option).strip()
                row_idx += 1
        helper_sheet[f"B{row_idx}"] = ""
        interest_last_row = row_idx
        
        # Column C: Conduct Remarks
        helper_sheet["C1"] = "Conduct Remarks"
        row_idx = 2
        for option in conduct_remarks_options:
            if option:  # Only add non-empty options
                helper_sheet[f"C{row_idx}"] = str(option).strip()
                row_idx += 1
        helper_sheet[f"C{row_idx}"] = ""
        conduct_last_row = row_idx
        
        # Column D: Attitude Remarks
        helper_sheet["D1"] = "Attitude Remarks"
        row_idx = 2
        for option in attitude_remarks_options:
            if option:  # Only add non-empty options
                helper_sheet[f"D{row_idx}"] = str(option).strip()
                row_idx += 1
        helper_sheet[f"D{row_idx}"] = ""
        attitude_last_row = row_idx
        
        # Format helper sheet headers
        for col in ["A1", "B1", "C1", "D1"]:
            helper_sheet[col].font = Font(bold=True)
        helper_sheet.column_dimensions["A"].width = 50
        helper_sheet.column_dimensions["B"].width = 50
        helper_sheet.column_dimensions["C"].width = 50
        helper_sheet.column_dimensions["D"].width = 50

        # Add data validation using INDIRECT function for better Excel 2019 compatibility
        # This approach is more reliable across different Excel versions
        from openpyxl.worksheet.datavalidation import DataValidation

        if max_row >= 8:
            # Get the actual sheet title
            helper_sheet_title = helper_sheet.title
            
            # Use INDIRECT to reference the helper sheet - this works better in Excel 2019
            # Format: INDIRECT("'SheetName'!$A$2:$A$15")
            general_formula = f'INDIRECT("\'{helper_sheet_title}\'!$A$2:$A${general_last_row}")'
            interest_formula = f'INDIRECT("\'{helper_sheet_title}\'!$B$2:$B${interest_last_row}")'
            conduct_formula = f'INDIRECT("\'{helper_sheet_title}\'!$C$2:$C${conduct_last_row}")'
            attitude_formula = f'INDIRECT("\'{helper_sheet_title}\'!$D$2:$D${attitude_last_row}")'
            
            # General Remarks dropdown (column 6)
            general_dv = DataValidation(
                type="list",
                formula1=general_formula,
                allow_blank=True,
                showDropDown=True,
            )
            general_dv.error = "Please select from the dropdown list"
            general_dv.errorTitle = "Invalid Entry"
            general_dv.prompt = "Select a general remark from the list"
            general_dv.promptTitle = "General Remarks"
            ws.add_data_validation(general_dv)
            general_dv.add(f"F8:F{max_row}")

            # Interest Remarks dropdown (column 7)
            interest_dv = DataValidation(
                type="list",
                formula1=interest_formula,
                allow_blank=True,
                showDropDown=True,
            )
            interest_dv.error = "Please select from the dropdown list"
            interest_dv.errorTitle = "Invalid Entry"
            interest_dv.prompt = "Select an interest remark from the list"
            interest_dv.promptTitle = "Interest Remarks"
            ws.add_data_validation(interest_dv)
            interest_dv.add(f"G8:G{max_row}")

            # Conduct Remarks dropdown (column 8)
            conduct_dv = DataValidation(
                type="list",
                formula1=conduct_formula,
                allow_blank=True,
                showDropDown=True,
            )
            conduct_dv.error = "Please select from the dropdown list"
            conduct_dv.errorTitle = "Invalid Entry"
            conduct_dv.prompt = "Select a conduct remark from the list"
            conduct_dv.promptTitle = "Conduct Remarks"
            ws.add_data_validation(conduct_dv)
            conduct_dv.add(f"H8:H{max_row}")

            # Attitude Remarks dropdown (column 9)
            attitude_dv = DataValidation(
                type="list",
                formula1=attitude_formula,
                allow_blank=True,
                showDropDown=True,
            )
            attitude_dv.error = "Please select from the dropdown list"
            attitude_dv.errorTitle = "Invalid Entry"
            attitude_dv.prompt = "Select an attitude remark from the list"
            attitude_dv.promptTitle = "Attitude Remarks"
            ws.add_data_validation(attitude_dv)
            attitude_dv.add(f"I8:I{max_row}")
            
            debug_log(f"Data validation added using INDIRECT for rows 8 to {max_row}")
            debug_log(f"Helper sheet: {helper_sheet_title}")

        # Auto-adjust column widths
        column_widths = [18, 30, 12, 15, 12, 40, 40, 40, 40]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Create response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{class_obj.name}_Remarks_{current_term.name}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        # Save workbook to response
        wb.save(response)
        return response

    except Exception as e:
        error_trace = traceback.format_exc()
        debug_log(f"Error in export_remarks: {str(e)}")
        debug_log(error_trace)
        messages.error(request, f"Error exporting remarks: {str(e)}")
        return redirect("class_teacher_remarks")


@login_required
@require_POST
def import_remarks(request):
    """Import student remarks from Excel file."""
    try:
        class_id = request.POST.get("class_id")
        if not class_id:
            return JsonResponse(
                {"status": "error", "message": "Missing class ID"}, status=400
            )

        if "file" not in request.FILES:
            return JsonResponse(
                {"status": "error", "message": "No file uploaded"}, status=400
            )

        file = request.FILES["file"]
        if not file.name.endswith((".xlsx", ".xls")):
            return JsonResponse(
                {"status": "error", "message": "Invalid file format. Please upload an Excel file."},
                status=400,
            )

        # Get user's school for multi-tenancy
        user_school = get_user_school(request.user)

        # Get class with school context
        class_query = Class.objects.filter(class_id=class_id)
        if user_school:
            class_query = class_query.filter(school=user_school)
        class_obj = get_object_or_404(class_query)

        # Get teacher with school context
        teacher_query = Teacher.objects.filter(user=request.user)
        if user_school:
            teacher_query = teacher_query.filter(school=user_school)
        teacher = get_object_or_404(teacher_query)

        # Verify teacher has access
        if not verify_teacher_class_access(teacher, class_obj, school=user_school):
            return JsonResponse(
                {"status": "error", "message": "You are not authorized to import remarks for this class."},
                status=403,
            )

        # Get current academic year and term
        current_academic_year, current_term = get_current_academic_data(user_school)
        if not current_academic_year or not current_term:
            return JsonResponse(
                {"status": "error", "message": "No active academic year or term found."},
                status=400,
            )

        # Read Excel file
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # Find header row (should be row 7)
        header_row = 7
        headers = []
        for col in range(1, 10):
            cell_value = ws.cell(row=header_row, column=col).value
            headers.append(str(cell_value).strip() if cell_value else "")

        # Validate headers
        expected_headers = [
            "Admission Number",
            "Student Name",
            "Days Present",
            "Total School Days",
            "Attendance %",
            "General Remarks",
            "Interest Remarks",
            "Conduct Remarks",
            "Attitude Remarks",
        ]
        
        if headers != expected_headers:
            return JsonResponse(
                {
                    "status": "error",
                    "message": "Invalid file format. Please use the exported template.",
                },
                status=400,
            )

        # Process data rows (starting from row 8)
        updated_count = 0
        error_count = 0
        errors = []

        for row_num in range(8, ws.max_row + 1):
            admission_number = ws.cell(row=row_num, column=1).value
            if not admission_number:
                continue  # Skip empty rows

            try:
                # Get student by admission number
                student_query = Student.objects.filter(
                    admission_number=str(admission_number).strip()
                )
                if user_school:
                    student_query = student_query.filter(school=user_school)
                student = student_query.first()

                if not student:
                    error_count += 1
                    errors.append(f"Row {row_num}: Student with admission number {admission_number} not found")
                    continue

                # Verify student is in the class
                student_class_query = StudentClass.objects.filter(
                    student=student, assigned_class=class_obj, is_active=True
                )
                if user_school:
                    student_class_query = student_class_query.filter(school=user_school)
                if not student_class_query.exists():
                    error_count += 1
                    errors.append(f"Row {row_num}: Student {admission_number} is not in this class")
                    continue

                # Get or create remarks
                remarks_query = StudentTermRemarks.objects.filter(
                    student=student,
                    academic_year=current_academic_year,
                    term=current_term,
                    class_assigned=class_obj,
                )
                if user_school:
                    remarks_query = remarks_query.filter(student__school=user_school)
                remarks = remarks_query.first()

                if not remarks:
                    remarks = StudentTermRemarks.objects.create(
                        student=student,
                        academic_year=current_academic_year,
                        term=current_term,
                        class_assigned=class_obj,
                        class_teacher=teacher,
                    )

                # Update fields from Excel
                days_present = ws.cell(row=row_num, column=3).value
                total_days = ws.cell(row=row_num, column=4).value
                general_remarks = ws.cell(row=row_num, column=6).value
                interest_remarks = ws.cell(row=row_num, column=7).value
                conduct_remarks = ws.cell(row=row_num, column=8).value
                attitude_remarks = ws.cell(row=row_num, column=9).value
                
                # Calculate days absent from days present and total days
                if days_present is not None and total_days is not None:
                    try:
                        days_absent = int(float(total_days)) - int(float(days_present))
                        days_absent = max(0, days_absent)  # Ensure non-negative
                    except (ValueError, TypeError):
                        days_absent = None
                else:
                    days_absent = None

                # Convert to appropriate types
                if days_present is not None:
                    try:
                        remarks.days_present = int(float(days_present))
                    except (ValueError, TypeError):
                        pass

                if days_absent is not None:
                    try:
                        remarks.days_absent = int(float(days_absent))
                    except (ValueError, TypeError):
                        pass

                if total_days is not None:
                    try:
                        remarks.total_school_days = int(float(total_days))
                    except (ValueError, TypeError):
                        pass

                # Update remarks - handle all cases including empty strings and None
                # General Remarks
                if general_remarks is not None:
                    general_value = str(general_remarks).strip()
                    remarks.general_remarks = general_value if general_value else None
                elif general_remarks == "":
                    remarks.general_remarks = None

                # Interest Remarks - explicitly handle all cases
                if interest_remarks is not None:
                    interest_value = str(interest_remarks).strip()
                    remarks.interest_remarks = interest_value if interest_value else None
                    debug_log(f"Row {row_num} - Student {admission_number}: Setting interest_remarks to: '{interest_value}'")
                elif interest_remarks == "":
                    remarks.interest_remarks = None
                    debug_log(f"Row {row_num} - Student {admission_number}: Interest remarks is empty string, setting to None")
                else:
                    debug_log(f"Row {row_num} - Student {admission_number}: Interest remarks is None, keeping existing value")

                # Conduct Remarks
                if conduct_remarks is not None:
                    conduct_value = str(conduct_remarks).strip()
                    remarks.conduct_remarks = conduct_value if conduct_value else None
                elif conduct_remarks == "":
                    remarks.conduct_remarks = None

                # Attitude Remarks
                if attitude_remarks is not None:
                    attitude_value = str(attitude_remarks).strip()
                    remarks.attitude_remarks = attitude_value if attitude_value else None
                elif attitude_remarks == "":
                    remarks.attitude_remarks = None

                # Ensure class_teacher is set
                if not remarks.class_teacher:
                    remarks.class_teacher = teacher

                # Save with explicit field update to ensure interest_remarks is saved
                remarks.save(update_fields=[
                    'days_present', 'days_absent', 'total_school_days',
                    'general_remarks', 'interest_remarks', 'conduct_remarks', 'attitude_remarks',
                    'class_teacher'
                ])
                debug_log(f"Row {row_num} - Student {admission_number}: Saved remarks. Interest remarks value: '{remarks.interest_remarks}'")
                updated_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f"Row {row_num}: {str(e)}")
                debug_log(f"Error processing row {row_num}: {str(e)}")
                continue

        # Prepare response
        response_data = {
            "status": "success",
            "message": f"Remarks imported successfully. {updated_count} students updated.",
            "updated_count": updated_count,
            "error_count": error_count,
        }

        if errors and error_count <= 10:  # Include errors if not too many
            response_data["errors"] = errors

        return JsonResponse(response_data)

    except Exception as e:
        error_trace = traceback.format_exc()
        debug_log(f"Error in import_remarks: {str(e)}")
        debug_log(error_trace)
        return JsonResponse(
            {"status": "error", "message": f"Error importing remarks: {str(e)}"},
            status=500,
        )


@login_required

def print_class_list(request, class_id):
    # Get user's school for multi-tenancy
    user_school = get_user_school(request.user)

    # Get class with school context
    class_query = Class.objects.filter(class_id=class_id)
    if user_school:
        class_query = class_query.filter(school=user_school)
    selected_class = get_object_or_404(class_query)

    # Get teacher with school context
    teacher_query = Teacher.objects.filter(user=request.user)
    if user_school:
        teacher_query = teacher_query.filter(school=user_school)
    teacher = get_object_or_404(teacher_query)

    # Verify teacher has access to this class with school context
    if not verify_teacher_class_access(teacher, selected_class, school=user_school):
        messages.error(request, "You are not authorized to view this class list.")
        return redirect("class_teacher_remarks")

    # Get active school information with school context
    school_info_query = SchoolInformation.objects.filter(is_active=True)
    if user_school:
        school_info_query = school_info_query.filter(school=user_school)
    school_info = school_info_query.first()

    # Get students with school context
    student_classes_query = StudentClass.objects.filter(
        assigned_class=selected_class, is_active=True
    )
    if user_school:
        student_classes_query = student_classes_query.filter(school=user_school)
    student_classes = student_classes_query.select_related("student")

    # Get current academic year and term for user's school
    current_academic_year, current_term = get_current_academic_data(user_school)

    context = {
        "class_obj": selected_class,
        "students": student_classes,
        "student_count": student_classes.count(),
        "teacher": teacher,
        "current_academic_year": current_academic_year,
        "current_term": current_term,
        "school": school_info,
        "user_school": user_school,  # Add user's school to context
    }

    return render(request, "teacher/print_class_list.html", context)


@login_required
def student_term_performance(request, class_id, student_id):
    try:
        # Get user's school for multi-tenancy
        user_school = get_user_school(request.user)

        # Get class with school context
        class_query = Class.objects.filter(class_id=class_id)
        if user_school:
            class_query = class_query.filter(school=user_school)
        selected_class = get_object_or_404(class_query)

        # Get teacher with school context
        teacher_query = Teacher.objects.filter(user=request.user)
        if user_school:
            teacher_query = teacher_query.filter(school=user_school)
        teacher = get_object_or_404(teacher_query)

        # Get student with school context
        student_query = Student.objects.filter(id=student_id)
        if user_school:
            student_query = student_query.filter(school=user_school)
        student = get_object_or_404(student_query)

        # Verify teacher has access to this class with school context
        if not verify_teacher_class_access(teacher, selected_class, school=user_school):
            messages.error(
                request, "You are not authorized to view this student's performance."
            )
            return redirect("class_teacher_remarks")

        # Get current academic year and term for user's school
        current_academic_year, current_term = get_current_academic_data(user_school)


        # Get subjects and assessments with school context (only active ones)
        class_subjects_query = ClassSubject.objects.filter(
            class_name=selected_class,
            academic_year=current_academic_year,
            is_active=True

        )
        if user_school:
            class_subjects_query = class_subjects_query.filter(
                class_name__school=user_school, academic_year__school=user_school
            )
        class_subjects = class_subjects_query.select_related("subject")

        # Get assessments with school context

        # Exclude mock exam assessments from student performance display
        assessments_query = Assessment.objects.filter(
            student=student, class_subject__in=class_subjects
        ).exclude(assessment_type='mock_exam')

        if user_school:
            assessments_query = assessments_query.filter(
                student__school=user_school,
                class_subject__class_name__school=user_school,
            )
        assessments = assessments_query.select_related("class_subject__subject")

        # Process assessment data
        subject_assessments = {}
        total_score = 0
        subjects_count = 0

        for assessment in assessments:
            subject = assessment.class_subject.subject
            if assessment.total_score:
                total_score += assessment.total_score
                subjects_count += 1

            subject_assessments[subject.id] = {
                "subject_name": subject.subject_name,
                "class_score": assessment.class_score,
                "exam_score": assessment.exam_score,
                "total_score": assessment.total_score,
                "grade": assessment.grade,
                "remarks": assessment.remarks,
                "position": assessment.position,
            }

        # Calculate statistics
        average_score = (
            round(total_score / subjects_count, 2) if subjects_count > 0 else 0
        )
        passed_subjects = sum(
            1 for a in assessments if a.total_score and a.total_score >= 40
        )
        failed_subjects = sum(
            1 for a in assessments if a.total_score and a.total_score < 40
        )

        # Get scoring configuration
        scoring_config = None
        if user_school:
            scoring_config = ScoringConfiguration.get_active_config(user_school)

        context = {
            "student": student,
            "class_obj": selected_class,
            "current_academic_year": current_academic_year,
            "current_term": current_term,
            "subject_assessments": subject_assessments,
            "total_subjects": len(class_subjects),
            "subjects_assessed": subjects_count,
            "average_score": average_score,
            "passed_subjects": passed_subjects,
            "failed_subjects": failed_subjects,
            "user_school": user_school,  # Add user's school to context
            "scoring_config": scoring_config,
        }

        return render(request, "teacher/student_term_performance.html", context)

    except Exception as e:
        messages.error(request, f"Error loading student performance: {str(e)}")
        return redirect("class_teacher_remarks")


@login_required
def auto_generate_remarks(request):
    """Automatically generate teacher remarks for a student based on performance data."""
    if request.method == "POST":
        student_id = request.POST.get("student_id")
        term_id = request.POST.get("term_id")
        include_trends = request.POST.get("include_trends") == "true"
        include_personality = request.POST.get("include_personality") == "true"
        include_suggestions = request.POST.get("include_suggestions") == "true"

        if not student_id or not term_id:
            return JsonResponse(
                {"success": False, "message": "Student and term must be provided"}
            )

        try:
            # Get user's school for multi-tenancy
            user_school = get_user_school(request.user)

            # Get student with school context
            student_query = Student.objects.filter(pk=student_id)
            if user_school:
                student_query = student_query.filter(school=user_school)
            student = get_object_or_404(student_query)

            # Get term with school context
            term_query = Term.objects.filter(pk=term_id)
            if user_school:
                term_query = term_query.filter(school=user_school)
            term = get_object_or_404(term_query)

            # Get student's assessments for this term with school context

            # Exclude mock exam assessments from term performance display
            assessments_query = Assessment.objects.filter(
                student=student, class_subject__academic_year=term.academic_year
            ).exclude(assessment_type='mock_exam')


            if user_school:
                assessments_query = assessments_query.filter(
                    student__school=user_school,
                    class_subject__academic_year__school=user_school,
                )

            assessments = assessments_query.select_related("class_subject__subject")

            if not assessments.exists():
                return JsonResponse(
                    {
                        "success": False,
                        "message": "No assessment data found for this student in this term",
                    }
                )

            # Calculate average score
            total_score = sum(a.total_score or 0 for a in assessments)
            num_subjects = assessments.count()
            average_score = total_score / num_subjects if num_subjects > 0 else 0

            # Generate base remarks based on performance
            base_remarks = ""

            # Basic performance remarks
            if average_score >= 80:
                base_remarks += f"{student.full_name} has demonstrated exceptional academic performance this term. "
            elif average_score >= 70:
                base_remarks += f"{student.full_name} has shown very good academic performance this term. "
            elif average_score >= 60:
                base_remarks += (
                    f"{student.full_name} has achieved good results this term. "
                )
            elif average_score >= 50:
                base_remarks += (
                    f"{student.full_name} has shown satisfactory academic progress. "
                )
            else:
                base_remarks += (
                    f"{student.full_name} needs to improve academic performance. "
                )

            # Add subject-specific remarks for best and worst subjects
            if assessments.count() > 1:
                best_subject = max(assessments, key=lambda a: a.total_score or 0)
                worst_subject = min(assessments, key=lambda a: a.total_score or 0)

                if best_subject.total_score and best_subject.total_score > 70:
                    base_remarks += f"Particularly strong in {best_subject.class_subject.subject.name}. "

                if worst_subject.total_score and worst_subject.total_score < 50:
                    base_remarks += f"Needs to focus more on {worst_subject.class_subject.subject.name}. "

            # Generate specialized remarks for different fields
            interest_remarks = "Shows interest in learning activities."
            if average_score >= 70:
                interest_remarks = "Demonstrates keen interest in academic activities and class participation."
            elif average_score >= 50:
                interest_remarks = (
                    "Shows adequate interest in school work and learning activities."
                )
            else:
                interest_remarks = (
                    "Needs to develop more interest in academic activities."
                )

            # Generate conduct remarks
            conduct_remarks = "Displays appropriate conduct in the school environment."
            if average_score >= 70:
                conduct_remarks = (
                    "Exhibits excellent behavior and follows school rules consistently."
                )
            elif average_score >= 50:
                conduct_remarks = "Generally follows school rules and demonstrates acceptable behavior."
            else:
                conduct_remarks = "Needs improvement in following school rules and maintaining proper conduct."

            # Generate attitude remarks
            attitude_remarks = "Has a positive attitude towards school work."
            if average_score >= 70:
                attitude_remarks = "Maintains a highly positive attitude towards learning and school activities."
            elif average_score >= 50:
                attitude_remarks = (
                    "Shows a generally positive attitude towards academic tasks."
                )
            else:
                attitude_remarks = "Should work on developing a more positive attitude towards school work."

            # Create teacher remarks (more comprehensive)
            teacher_remarks = base_remarks

            # Add trend information if requested
            if include_trends:
                teacher_remarks += "Showing consistent effort in coursework. "

            # Add personality traits if requested
            if include_personality:
                teacher_remarks += "Demonstrates a positive attitude towards learning. "

            # Add suggestions if requested
            if include_suggestions:
                if average_score < 60:
                    teacher_remarks += "Would benefit from additional study time and seeking help when needed. "
                else:
                    teacher_remarks += (
                        "Encouraged to continue with the current level of dedication. "
                    )

            # Return all the generated remarks
            return JsonResponse(
                {
                    "success": True,
                    "interest_remarks": interest_remarks.strip(),
                    "conduct_remarks": conduct_remarks.strip(),
                    "attitude_remarks": attitude_remarks.strip(),
                    "teacher_remarks": teacher_remarks.strip(),
                }
            )

        except (Student.DoesNotExist, Term.DoesNotExist) as e:
            return JsonResponse({"success": False, "message": str(e)})
        except Exception as e:
            return JsonResponse(
                {"success": False, "message": f"Error generating remarks: {str(e)}"}
            )

    # If not POST request, return an error
    return JsonResponse({"success": False, "message": "Only POST requests are allowed"})
