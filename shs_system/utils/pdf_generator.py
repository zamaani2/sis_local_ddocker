import os
import logging
from io import BytesIO
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import render_to_string

logger = logging.getLogger(__name__)

try:
    import weasyprint

    # Test if WeasyPrint can actually be used (not just imported)
    # This catches OSError from missing system dependencies
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError) as e:
    WEASYPRINT_AVAILABLE = False
    # Log at debug level since fallback is available and working
    logger.debug(
        f"WeasyPrint not available (error: {type(e).__name__}: {str(e)}). "
        f"PDF generation will use ReportLab fallback method."

    )

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning(
        "OpenPyXL not available. Excel generation will use alternative method."
    )


def generate_pdf_from_html(html_content, css_content=None):
    """
    Generate PDF from HTML content using WeasyPrint or alternative method.
    """
    try:
        if WEASYPRINT_AVAILABLE:
            # Use WeasyPrint for high-quality PDF generation
            html_doc = weasyprint.HTML(string=html_content)

            # Add CSS if provided
            if css_content:
                css_doc = weasyprint.CSS(string=css_content)
                pdf_content = html_doc.write_pdf(stylesheets=[css_doc])
            else:
                pdf_content = html_doc.write_pdf()

            return pdf_content
        else:

            # Fallback to ReportLab for PDF generation
            logger.debug("Using ReportLab fallback for PDF generation")

            return generate_pdf_fallback(html_content)

    except Exception as e:
        logger.error(f"Error generating PDF: {str(e)}")
        # Return a simple error PDF
        return generate_error_pdf(str(e))


def generate_pdf_fallback(html_content):
    """
    Fallback PDF generation method when WeasyPrint is not available.
    """
    try:
        # This is a basic fallback - in production, you might want to use
        # a different library like reportlab or wkhtmltopdf
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()

        # Convert HTML to simple text for reportlab
        import re

        text_content = re.sub(r"<[^>]+>", "", html_content)

        story = []
        story.append(Paragraph("Score Sheet Report", styles["Title"]))
        story.append(Spacer(1, 12))
        story.append(Paragraph(text_content, styles["Normal"]))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    except ImportError:
        logger.error("ReportLab not available for fallback PDF generation")
        return generate_error_pdf("PDF generation not available")


def generate_error_pdf(error_message):
    """
    Generate a simple error PDF when PDF generation fails.
    """
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()

        story = []
        story.append(Paragraph("Error Generating PDF", styles["Title"]))
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"An error occurred: {error_message}", styles["Normal"]))
        story.append(Spacer(1, 12))
        story.append(
            Paragraph("Please contact the system administrator.", styles["Normal"])
        )

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    except Exception:
        # Ultimate fallback - return empty bytes
        return b""


def generate_excel_from_data(excel_data):
    """
    Generate Excel file from structured data.
    """
    try:
        if OPENPYXL_AVAILABLE:
            return generate_excel_openpyxl(excel_data)
        else:
            logger.warning("OpenPyXL not available, using fallback Excel generation")
            return generate_excel_fallback(excel_data)

    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}")
        return generate_error_excel(str(e))


def generate_excel_openpyxl(excel_data):
    """
    Generate Excel file using OpenPyXL.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Score Sheet"

    # Define styles
    title_font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=11)

    title_fill = PatternFill(
        start_color="007bff", end_color="007bff", fill_type="solid"
    )
    header_fill = PatternFill(
        start_color="6c757d", end_color="6c757d", fill_type="solid"
    )

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    # Add title
    worksheet.merge_cells("A1:Z1")
    title_cell = worksheet["A1"]
    title_cell.value = excel_data.get("title", "Score Sheet")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment

    # Add subtitle
    if "subtitle" in excel_data:
        worksheet.merge_cells("A2:Z2")
        subtitle_cell = worksheet["A2"]
        subtitle_cell.value = excel_data["subtitle"]
        subtitle_cell.font = Font(name="Arial", size=12, italic=True)
        subtitle_cell.alignment = center_alignment

    # Add headers
    headers = excel_data.get("headers", [])
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # Add data
    data = excel_data.get("data", [])
    for row_idx, row_data in enumerate(data, 5):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = cell_value
            cell.font = data_font
            cell.border = border

            # Special formatting for different columns
            if col_idx == 1:  # Position column
                cell.alignment = center_alignment
                cell.font = Font(name="Arial", size=11, bold=True, color="007bff")
            elif col_idx == 2:  # Student name column
                cell.alignment = left_alignment
                cell.font = Font(name="Arial", size=11, bold=True)
            else:
                cell.alignment = center_alignment

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 20)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_fallback(excel_data):
    """
    Fallback Excel generation when OpenPyXL is not available.
    """
    try:
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)

        # Write title
        writer.writerow([excel_data.get("title", "Score Sheet")])
        writer.writerow([excel_data.get("subtitle", "")])
        writer.writerow([])  # Empty row

        # Write headers
        writer.writerow(excel_data.get("headers", []))

        # Write data
        for row in excel_data.get("data", []):
            writer.writerow(row)

        # Convert to bytes
        csv_content = output.getvalue()
        output.close()

        return csv_content.encode("utf-8")

    except Exception as e:
        logger.error(f"Error in Excel fallback generation: {str(e)}")
        return generate_error_excel(str(e))


def generate_error_excel(error_message):
    """
    Generate a simple error Excel file when Excel generation fails.
    """
    try:
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(["Error Generating Excel"])
        writer.writerow([f"An error occurred: {error_message}"])
        writer.writerow(["Please contact the system administrator."])

        csv_content = output.getvalue()
        output.close()

        return csv_content.encode("utf-8")

    except Exception:
        return b"Error generating Excel file"


def generate_report_card_pdf(
    report_card,
    subject_scores,
    teacher_remarks,
    grades,
    authority_signatures,
    school_info,
    class_size,
):
    """
    Generate PDF for report card.
    This function is used by the report card system.
    """
    try:
        from django.template.loader import render_to_string

        # Render the report card template
        html_content = render_to_string(
            "reports/report_card_pdf.html",
            {
                "report_card": report_card,
                "subject_scores": subject_scores,
                "teacher_remarks": teacher_remarks,
                "grades": grades,
                "authority_signatures": authority_signatures,
                "school_info": school_info,
                "class_size": class_size,
            },
        )

        # Generate PDF using the main PDF generation function
        return generate_pdf_from_html(html_content)

    except Exception as e:
        logger.error(f"Error generating report card PDF: {str(e)}")
        return generate_error_pdf(str(e))
